#!/usr/bin/python2.7
import qrcode
import labels
import requests
import os
from reportlab.graphics import shapes
from reportlab.platypus import Image
from reportlab.lib.units import inch
#from qrcode.image.pure import PymagingImage
from reportlab.pdfgen import canvas
from reportlab.graphics.barcode import code39, code128, code93
from reportlab.graphics.barcode import eanbc, qr, usps
from reportlab.graphics.shapes import Drawing
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas
from reportlab.graphics import renderPDF
from datetime import datetime
import uuid
testImg = 'pic.png'

class generateProductOrderQRCode():
    def __init__(self):
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_H,
            box_size=10,
            border=4,
        )
        qr.add_data('123456789')
        qr.make(fit=True)
        img = qr.make_image()
        image_file = open('media/ProductOrderQR/piiic.png','wb')
        img.save(image_file)

        image_file.close()

class GenerateQRLabelDoc():

    def __init__(self):
        self.specs = None
        self.sheet = None
        self.labelCount = 1
        self.printTemplates = []
        self.pageWidth = 0
        self.pageHeight = 0
        self.columnCount = 0
        self.rowCount = 0
        self.labelLength = 0
        self.labelHeight = 0
        self.row_gap = 0

    # Create an A4 portrait (210mm x 297mm) sheets with 2 columns and 8 rows of
    # labels. Each label is 90mm x 25mm with a 5mm rounded corner. The margins are
    # automatically calculated.

    # Create a function to draw each label. This will be given the ReportLab drawing
    # object to draw on, the dimensions (NB. these will be in points, the unit
    # ReportLab uses) of the label, and the object to render.
    def draw_label(label, width, height, obj):
        # Just convert the object to a string and print this at the bottom left of
        # the label.
        label.add(shapes.String(2, 2, str(obj), fontName="Helvetica", fontSize=10))

    def draw_image(self, label, width, height, obj):
        # Just convert the object to a string and print this at the bottom left of
        # the label.
        #label.add(shapes.String(2, 2, str(obj), fontName="Helvetica", fontSize=40)) #new
        label.add(obj)

    def getLabelCount(self):
        return self.labelCount

    def createBarCode(self, uniqueID, productName, upc, sku):
        """
        Create barcode examples and embed in a PDF
        """
        #c = canvas.Canvas("barcodes.pdf", pagesize=A4)
        # draw a QR code
        qr_code = qr.QrCodeWidget(uniqueID)
        bounds = qr_code.getBounds()
        width = bounds[2] - bounds[0]
        height = bounds[3] - bounds[1]
        d = Drawing(40, 40, transform=[70 / width, 0, 0, 70 / height, 0, 0])
        d.add(qr_code)
        print('0-----------------')
        sku = sku if sku else ' '

        upc = upc if upc else ' '

        labelStr = "#"+str(self.getLabelCount())
        d.add(shapes.String(90, 70, str(productName), fontName="Helvetica", fontSize=10))
        d.add(shapes.String(210, 70, str(labelStr), fontName="Helvetica", fontSize=10))
        d.add(shapes.String(90, 55, str('XSPACE ID: '+uniqueID.split('-')[0]), fontName="Helvetica", fontSize=10))
        d.add(shapes.String(90, 35, str('SKU: '+ sku), fontName="Helvetica", fontSize=10))
        d.add(shapes.String(90, 15, str('UPC: ' + upc), fontName="Helvetica", fontSize=10))

        return d

    def createBarCode2(self, uniqueID, productName, upc, sku):
        """
        Create barcode examples and embed in a PDF
        """
        #c = canvas.Canvas("barcodes.pdf", pagesize=A4)
        # draw a QR code
        qr_code = qr.QrCodeWidget(uniqueID)
        bounds = qr_code.getBounds()
        width = bounds[2] - bounds[0]
        height = bounds[3] - bounds[1]
        d = Drawing(40, 40, transform=[70 / width, 0, 0, 70 / height, 0, 0])
        d.add(qr_code)

        sku = sku if sku else ' '

        upc = upc if upc else ' '

        labelStr = "#"+str(self.getLabelCount())
        d.add(shapes.String(90, 70, str(productName), fontName="Helvetica", fontSize=10))
        d.add(shapes.String(210, 70, str(labelStr), fontName="Helvetica", fontSize=10))
        d.add(shapes.String(90, 55, str('XSPACE ID: '+uniqueID), fontName="Helvetica", fontSize=10))
        d.add(shapes.String(90, 35, str('SKU: '+ sku), fontName="Helvetica", fontSize=10))
        d.add(shapes.String(90, 15, str('UPC: ' + upc), fontName="Helvetica", fontSize=10))

        return d

    def createDocFromExcel(self, excelSheet):

        # Optional arguments; missing ones will be computed later.
        left_margin = 1
        column_gap = 0
        right_margin = 1
        top_margin = 1
        row_gap = 0
        bottom_margin = 10.668

        #Optional arguments with default values.
        #self._left_padding =
        #self._right_padding =
        #self._top_padding =
        #self._bottom_padding =
        #self._corner_radius =
        #self._padding_radius =
        #self._background_image =
        #self._background_filename =
        no_of_rows = 8
        no_of_columns = 3
        corner_radius = 2
        page_size = 'Plain Paper'

        width,height = self.selectPaperSize(page_size)

        #specs = labels.Specification(215.9, 279.4, 3, 8, 66.802, 25.4, corner_radius=2, row_gap=row_gap, top_margin =12.7)
        specs = labels.Specification(width, height, no_of_columns, no_of_rows, 66.802, 25.4, corner_radius=corner_radius, row_gap=row_gap, top_margin =12.7)

        # Create the sheet.
        sheet = labels.Sheet(specs, self.draw_image, border=True)

        # Add a couple of labels.
        import random

        # ProductOrderList
        rando = random.randint(1,100)

        i = 0
        print("ITERABLE")
        from openpyxl import load_workbook
        myfile = excelSheet
        wb = load_workbook(filename=myfile, read_only=True)
        sheets = wb.sheetnames
        ws = wb['Sheet1']  # ws is now an IterableWorksheet
        temp = []

        row_num = 0

        for index, row in enumerate(ws.rows):

            if index == 0:
                continue
            else:
                row_num += 1
                row_list = []

                #Row 0 - Name, Row 2 - UPC, Row 3 - UPC
                print(row[0].value)
                newUpc = str(row[4].value)
                newForm = row[6].value

                orderID = str(uuid.uuid4())

                #createBarCode(uniqueID, productName, upc, sku):
                #barcode = self.createBarCode('123456789', 'Cool Product Here', '54366547424', 'COOLSKU-320', )
                barcode = self.createBarCode2(orderID,  str(row[0].value),  str(row[2].value),  str(row[3].value))
                sheet.add_label(barcode)
                self.labelCount = self.labelCount + 1
                i = i+1

            # if row_num > 1 and newForm != None:
            #     print(row_list)
            #     row_list = [row[0].value, row[1].value, row[2].value, row[3].value, newUpc.lower(), row[5].value,
            #                 newForm, row[7].value, row[8].value, row[9].value, row[10].value, row_num]


        #Iterate the loop unil end of Product Order List
        # for orderItem in productOrder:
        #     #createBarCode(uniqueID, productName, upc, sku):
        #     #barcode = self.createBarCode('123456789', 'Cool Product Here', '54366547424', 'COOLSKU-320', )
        #     barcode = self.createBarCode(orderItem.uniqueID,  orderItem.product.name, orderItem.product.upccode, orderItem.product.SKU)
        #     sheet.add_label(barcode)
        #     self.labelCount = self.labelCount + 1
        #     i = i+1

        # sheet.add_label("World")
        # We can also add each item from an iterable.
        # sheet.add_labels(range(3, 22))
        # Note that any oversize label is automatically trimmed to prevent it messing up
        # other labels.
        # sheet.add_label("Oversized label here")
        # Save the file and we are done.



        fileName = 'OrderPrintLabel'+"_"+"Order"+"_"

        # try:
        #    dirpath = os.getcwd()
        #    current_dirpath = dirpath
        #except:
        current_dirpath = "/tmp/"

        pathname = current_dirpath+fileName+'.pdf'
        print("PN--------------------------------")
        print("OK------")
        print(pathname)
        sheet.save(pathname)
        print("File: "+pathname+" / {0:d} label(s) output on {1:d} page(s).".format(sheet.label_count, sheet.page_count))
        return pathname

    def createDoc(self, productOrder, data, user):

        # Optional arguments; missing ones will be computed later.
        left_margin = 1
        column_gap = 0
        right_margin = 1
        top_margin = 1
        row_gap = 0
        bottom_margin = 10.668

        #Optional arguments with default values.
        #self._left_padding =
        #self._right_padding =
        #self._top_padding =
        #self._bottom_padding =
        #self._corner_radius =
        #self._padding_radius =
        #self._background_image =
        #self._background_filename =
        no_of_rows = int(data['numberOfRowsPerPage'])
        no_of_columns = int(data['numberOfColumnsPerPage'])
        corner_radius = int(data['cornerRadiusOfLabels'])
        page_size = data['activeTemplate']

        width,height = self.selectPaperSize(page_size)

        #specs = labels.Specification(215.9, 279.4, 3, 8, 66.802, 25.4, corner_radius=2, row_gap=row_gap, top_margin =12.7)
        specs = labels.Specification(width, height, no_of_columns, no_of_rows, 66.802, 25.4, corner_radius=corner_radius, row_gap=row_gap, top_margin =12.7)

        # Create the sheet.
        sheet = labels.Sheet(specs, self.draw_image, border=True)

        # Add a couple of labels.
        import random

        # ProductOrderList
        rando = random.randint(1,100)

        i = 0

        #Iterate the loop unil end of Product Order List
        for orderItem in productOrder:
            #createBarCode(uniqueID, productName, upc, sku):
            #barcode = self.createBarCode('123456789', 'Cool Product Here', '54366547424', 'COOLSKU-320', )
            barcode = self.createBarCode(orderItem.uniqueID, orderItem.product.name, orderItem.product.upccode, orderItem.product.SKU)
            sheet.add_label(barcode)
            self.labelCount = self.labelCount + 1
            i = i+1

        # sheet.add_label("World")
        # We can also add each item from an iterable.
        # sheet.add_labels(range(3, 22))
        # Note that any oversize label is automatically trimmed to prevent it messing up
        # other labels.
        # sheet.add_label("Oversized label here")
        # Save the file and we are done.


        fileName = 'OrderPrintLabel'+"_"+data['orderId'].split('-')[0]+"_"+user.first_name+"_"+user.last_name+"_"+datetime.now().strftime('%Y.%m.%d.%H.%M.%S')
        print(fileName)
        #try:
        #     dirpath = os.getcwd()
        #     current_dirpath = dirpath
        #except:
        current_dirpath = "/tmp/"
        print("--------------from file") 
        pathname = current_dirpath+fileName+'.pdf'
        print(pathname)
        #pathname = os.path.join(current_dirpath,sheetFile)

        sheet.save(pathname)

        print("File: "+pathname+" / {0:d} label(s) output on {1:d} page(s).".format(sheet.label_count, sheet.page_count))
        return pathname

    def selectPaperSize(self, paperType):
        try:
            if paperType == 'A6':
                width = 105
                height = 148
                return width, height
            elif paperType == 'A5':
                width = 148
                height = 210
                return width, height
            elif paperType == 'A4':
                width = 216
                height = 279
                return width, height
            elif paperType == 'A3':
                width = 279
                height = 432
                return width, height
            elif paperType == 'A2':
                width = 432
                height = 559
                return width, height
            elif paperType == 'A1':
                width = 559
                height = 864
                return width, height
            elif paperType == 'Plain Paper':
                width = 215.9
                height = 279.9
                return width, height
        except:
            width = 216
            height = 279
            return  width, height
#x = GenerateQRLabelDoc()
#x.createDocFromExcel('/Users/dommuscatella/Downloads/XSPACE_287 (1).xlsx')