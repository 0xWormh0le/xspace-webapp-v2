from app.core.constants import get_access_token_url, get_all_products_url, SHOPIFY_CLIENT_ID, SHOPIFY_CLIENT_SECRET
from app.accounts.models import Permission, Profile, Company
from rest_framework.exceptions import AuthenticationFailed
from rest_framework import exceptions
from rest_framework import status
from rest_framework.response import Response
import uuid
import requests
import os
import re
from datetime import datetime
from django.template.defaultfilters import slugify
from app.core.elasticsearch import *
from openpyxl import Workbook
from django.contrib.auth.models import AnonymousUser
from rest_framework.authtoken.models import Token
from django.contrib.auth.models import User
from django.utils.html import strip_tags
from rest_framework.authentication import TokenAuthentication, get_authorization_header
from rest_framework.exceptions import AuthenticationFailed
from rest_framework import exceptions
import chargebee
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from rest_framework import permissions
from django.contrib.auth import get_user_model
import boto3
from django.conf import settings

User = get_user_model()
# from xspace.upload_drive import uploadFile
import ipdb
from zenpy import Zenpy
from zenpy.lib.api_objects import User as ZenUser, Organization as ZenOrganization, OrganizationField, Comment, Ticket
import zenpy.lib.exception


class Struct:
    def __init__(self, **entries):
        self.__dict__.update(entries)


perm_dict = {'ALL': 'ALL',
             'ADMIN': 'ADMIN',
             'WHITELABEL': 'WHITELABEL',
             'READ_ONLY': 'READ_ONLY',
             'GENERIC': 'GENERIC',
             'READ': 'READ',
             'Create_Order': 'CREATE_ORDER',
             'Update_Order': 'UPDATE_ORDER',
             'Delete_Order': 'DELETE_ORDER',
             'Create_Category': 'CREATE_CATEGORY',
             'Update_Category': 'UPDATE_CATEGORY',
             'Delete_Category': 'DELETE_CATEGORY',
             'Create_Tag': 'CREATE_TAG',
             'Update_Tag': 'UPDATE_TAG',
             'Delete_Tag': 'DELETE_TAG',
             'Create_Product': 'CREATE_PRODUCT',
             'Update_Product': 'UPDATE_PRODUCT',
             'Delete_Product': 'DELETE_PRODUCT',
             'Create_ProductAsset': 'CREATE_PRODUCTASSET',
             'Update_ProductAsset': 'UPDATE_PRODUCTASSET',
             'Delete_ProductAsset': 'DELETE_PRODUCTASSET',
             'Create_Application': 'CREATE_APPLICATION',
             'Update_Application': 'UPDATE_APPLICATION',
             'Delete_Application': 'DELETE_APPLICATION',
             'Change_API_key': 'CHANGE_API_KEY',
             'View_org_product_list': 'VIEW_ORG_PRODUCT_LIST',
             'Create_a_shipping_method': 'CREATE_A_SHIPPING_METHOD',
             'Update_a_shipping_method': 'UPDATE_A_SHIPPING_METHOD',
             'Enter/Adjust_3rd_party_API_key': 'ENTER/ADJUST_3RD_PARTY_API_KEY',
             'Import_products_from_oAuth_API': 'IMPORT_PRODUCTS_FROM_OAUTH_API',
             'Sync_products_from_oAuth_API': 'SYNC_PRODUCTS_FROM_OAUTH_API'}

reverse_dict = dict([(value, key) for key, value in perm_dict.items()])

master_permissions_list = Struct(**reverse_dict)


# print(master_permissions_list.SYNC_PRODUCTS_FROM_OAUTH_API)

# outputs:
# In [1]: import app.core.utility
# Sync_products_from_oAuth_API


def has_permission(p):
    def wrap(fn):
        def perform_authentication(self, request, *args, **kwargs):
            if hasattr(request, 'user'):
                if str(request.user) != 'AnonymousUser':
                    user = User.objects.get(pk=request.user.pk)
            if hasattr(request, 'query_params'):
                if 'user_id' in request.query_params:

                    user_id = request.query_params['user_id']
                    try:
                        user = User.objects.get(pk=user_id)
                        request.user = user
                    except User.DoesNotExist:
                        raise exceptions.AuthenticationFailed('Invalid User')
            if hasattr(request, 'data'):
                if 'userid' in request.data:
                    user_id = request.data['userid']
                    try:
                        user = User.objects.get(pk=user_id)
                        request.user = user
                    except User.DoesNotExist:
                        raise exceptions.AuthenticationFailed('Invalid User')
                if 'userId' in request.data:
                    user_id = request.data['userId']
                    try:
                        user = User.objects.get(pk=user_id)
                        request.user = user
                    except User.DoesNotExist:
                        raise exceptions.AuthenticationFailed('Invalid User')

                if 'user' in request.data:
                    user_id = request.data['user']
                    try:
                        user = User.objects.get(pk=user_id)
                        request.user = user
                    except User.DoesNotExist:
                        raise exceptions.AuthenticationFailed('Invalid User')
                if 'permission' in request.data:
                    user_id = request.data['permission']['user']
                    try:
                        user = User.objects.get(pk=user_id)
                        request.user = user
                    except User.DoesNotExist:
                        raise exceptions.AuthenticationFailed('Invalid User')
            else:
                raise exceptions.AuthenticationFailed('malformed data')
            try:
                profile = Profile.objects.get(user=user.pk)
            except Profile.DoesNotExist:
                raise exceptions.AuthenticationFailed('Invalid Profile')
            try:
                company = Company.objects.get(pk=profile.companyProfile.id)
            except (Company.DoesNotExist, AttributeError) as e:
                raise exceptions.AuthenticationFailed('Invalid Company')
            try:
                permission = Permission.objects.filter(user_id=user.pk)
            except Permission.DoesNotExist:
                raise exceptions.AuthenticationFailed('User has no permissions assigned')
            try:
                for userP in permission:
                    if (
                            userP.perm == p or userP.perm == 'ALL' or userP.perm == 'ADMIN') and userP.company.pk == company.pk:  # permission.company == company.pk and
                        return fn(self, request, *args, **kwargs)

                return Response({"result": "Invalid permissions."}, status=status.HTTP_401_UNAUTHORIZED)
            except Exception as e:
                return Response({"result": "Invalid request.", "Error:": str(e)}, status=status.HTTP_400_BAD_REQUEST)

        return perform_authentication

    return wrap


def parse_objects(request, raw_json):
    from app.products.models import Product
    from app.core.models import Category
    product_list = raw_json['products']
    update = 0
    skipped = 0
    added = 0
    for product in product_list:
        name = product['title']
        ecommerce_id = 'shopify-' + str(product['id'])
        upc_sku = product['variants'][0]['sku']
        price = product['variants'][0]['price']
        description = strip_tags(product['body_html'])
        category = product['product_type']

        try:

            category_object = Category.objects.filter(name__iexact=category).first()
            if not category_object:
                category_object = Category.objects.filter(name__iexact='Uncategorized').first()
            try:
                prod_obj = Product.objects.get(ecommerce_id=ecommerce_id)
                prod_obj.name = name
                prod_obj.description = description
                prod_obj.slug = slugify(name)
                prod_obj.category = category_object
                prod_obj.SKU = upc_sku
                prod_obj.price = price
                prod_obj.save()
                update += 1
            except:
                prod_obj = Product.objects.create(profile=request.user.profile, uniqueID=uuid.uuid4(),
                                                  name=name, slug=slugify(name),
                                                  description=description,
                                                  category=category_object, SKU=upc_sku,
                                                  price=price, ecommerce_id=ecommerce_id)
                added += 1
                inject_product_data(prod_obj)


        except:
            skipped += 1

    return added, skipped, update


def iterate_and_save_products(request, obj, shop):
    get_access_token = requests.post('https://' + shop + get_access_token_url, data={'client_id': SHOPIFY_CLIENT_ID,
                                                                                     'client_secret': SHOPIFY_CLIENT_SECRET,
                                                                                     'code': obj.code})
    # print(get_access_token.status_code, get_access_token.reason)
    result = get_access_token.json()
    access_token = result['access_token']

    response = requests.get('https://' + shop + get_all_products_url, headers={'X-Shopify-Access-Token': access_token})
    product_list = response.json()
    added, skipped, update = parse_objects(request, product_list)
    return added, skipped, update


class AuthenticatedUserMixin(object):
    """
    Purpose: Mixin for authenticating user based on token provided
    """

    def perform_authentication(self, request):
        from rest_framework_jwt.settings import api_settings
        jwt_decode_handler = api_settings.JWT_DECODE_HANDLER
        if 'HTTP_AUTHORIZATION' in request.META:
            key = request.META.get('HTTP_AUTHORIZATION').split(' ')[1]
            try:
                payload = jwt_decode_handler(key)

                user = Token.objects.get(user_id=payload['user_id']).user
                request.user = user
            except:
                raise exceptions.AuthenticationFailed('Invalid token')
        elif 'user_id' in request.query_params:
            user_id = request.query_params['user_id']
            try:
                user = Token.objects.get(user_id=user_id).user
                request.user = user
            except User.DoesNotExist:
                raise exceptions.AuthenticationFailed('Invalid User')
        elif 'userid' in request.data:
            user_id = request.data['userid']
            try:
                user = Token.objects.get(user_id=user_id).user
                request.user = user
            except User.DoesNotExist:
                raise exceptions.AuthenticationFailed('Invalid User')
        else:
            raise exceptions.AuthenticationFailed('Access Denied')


class ChargebeeAPI:

    def __init__(self, request):
        # testing = False
        testing = False
        if testing:
            # self.site_api_key = 'test_F2pYUouEFVuhINEdOxNl9UlWNgbb8ymw'
            self.site_api_key = 'test_blgq94WQccucuLcuWhajT7NaRtEKS2fRhj5'
            self.site = 'xspaceapp-test'
        else:
            self.site_api_key = 'live_TpNiyM4Qvhky9JwTwnQIqdq8tgRUrSJk'
            self.site = 'xspaceapp'

        # time in UTC seconds for quotes to expire. Currently:  6 months.
        self.quote_expiry_utc = time.time() + 6*30*24*60*60

        self.invoice_id = "None"
        self.hostname = str(request.META['HTTP_HOST'])
        if (request.META.get('HTTP_X_API_SUBDOMAIN')):
            self.subdomain = str(request.META['HTTP_X_API_SUBDOMAIN'])
        else:
            self.subdomain = 'localhost:3000'
        chargebee.configure(self.site_api_key, self.site)


    def create_invoice(self, request, productList):
        try:
            twoDPhotoCount = 0
            threeSixtyPhotoCount = 0
            videoCount = 0
            threeDModelCount = 0
            singlePhotoServicesCount = 0
            fourPackServicesCount = 0
            fivePackPackServicesCount = 0
            standard360ServicesCount = 0
            integrated360ServicesCount = 0
            standard3DModelServicesCount = 0
            advanced3DModelServicesCount = 0
            cinematic3DModelServicesCount = 0
            hdProductVideoServicesCount = 0
            fourKProductServicesCount = 0
            videoWalkthroughServicesCount = 0
            addonsList = []

            for product in productList:

                # New Fields
                if product['singlePhotoServicesRequired'] == True:
                    singlePhotoServicesCount = singlePhotoServicesCount + 1
                if product['fourPackServicesRequired'] == True:
                    fourPackServicesCount = fourPackServicesCount + 1
                if product['fivePackPackServicesRequired'] == True:
                    fivePackPackServicesCount = fivePackPackServicesCount + 1

                # 360 Models
                if product['standard360ServicesRequired'] == True:
                    standard360ServicesCount = standard360ServicesCount + 1
                if product['integrated360ServicesRequired'] == True:
                    integrated360ServicesCount = integrated360ServicesCount + 1

                if product['standard3DModelServicesRequired'] == True:
                    standard3DModelServicesCount = standard3DModelServicesCount + 1
                if product['advanced3DModelServicesRequired'] == True:
                    advanced3DModelServicesCount = advanced3DModelServicesCount + 1
                if product['cinematic3DModelServicesRequired'] == True:
                    cinematic3DModelServicesCount = cinematic3DModelServicesCount + 1

                if product['hdProductVideoServicesRequired'] == True:
                    hdProductVideoServicesCount = hdProductVideoServicesCount + 1
                if product['fourKProductServicesRequired'] == True:
                    fourKProductServicesCount = fourKProductServicesCount + 1
                if product['videoWalkthroughServicesRequired'] == True:
                    videoWalkthroughServicesCount = videoWalkthroughServicesCount + 1

            # 2D Services
            if singlePhotoServicesCount > 0:
                keyList = {"id": "single-2d-photo", "quantity": singlePhotoServicesCount}
                addonsList.append(keyList)

            if fourPackServicesCount > 0:
                keyList = {"id": "2d-four-pack", "quantity": fourPackServicesCount}
                addonsList.append(keyList)

            if fivePackPackServicesCount > 0:
                keyList = {"id": "2d-five-pack", "quantity": fivePackPackServicesCount}
                addonsList.append(keyList)

            # 360 Services
            if standard360ServicesCount > 0:
                keyList2 = {"id": "360-photo-pack", "quantity": standard360ServicesCount}
                addonsList.append(keyList2)

            if integrated360ServicesCount > 0:
                keyList3 = {"id": "int-360-photo-pack", "quantity": integrated360ServicesCount}
                addonsList.append(keyList3)

            # 3D Services
            if threeDModelCount > 0:
                keyList4 = {"id": "standard-3d-model", "quantity": threeDModelCount}
                addonsList.append(keyList4)

            if advanced3DModelServicesCount > 0:
                keyList = {"id": "Performance-3d-model", "quantity": advanced3DModelServicesCount}
                addonsList.append(keyList)

            if cinematic3DModelServicesCount > 0:
                keyList2 = {"id": "advanced-3d-model", "quantity": cinematic3DModelServicesCount}
                addonsList.append(keyList2)

            # Video Services
            if hdProductVideoServicesCount > 0:
                keyList4 = {"id": "hd-video", "quantity": hdProductVideoServicesCount}
                addonsList.append(keyList4)

            if fourKProductServicesCount > 0:
                keyList = {"id": "4k-video", "quantity": fourKProductServicesCount}
                addonsList.append(keyList)

            if videoWalkthroughServicesCount > 0:
                keyList2 = {"id": "video_walkthrough", "quantity": videoWalkthroughServicesCount}
                addonsList.append(keyList2)

            result = chargebee.Invoice.create({
                "customer_id": request.user.profile.companyProfile.uniqueID,
                "addons": addonsList,
                "recurring": False,
            })

            # print(twoDPhotoCount)

            invoice = result.invoice
            self.invoice_id = invoice.id
            return invoice
        except:
            return "Empty"

    def create_editing_estimate(self, profile, fileCount, hasBackgroundRemoval, hasAutoCropAndCenter, hasRename,
                        hasBlemishReduction, hasBarcodeDetection):

        addonsList = []

        # 2D Services
        if hasBackgroundRemoval:
            keyList = {"id": "background-removal", "quantity": fileCount}
            addonsList.append(keyList)

        # 360 Services
        if hasAutoCropAndCenter:
            keyList2 = {"id": "auto-crop-and-center", "quantity": fileCount}
            addonsList.append(keyList2)

        if hasRename:
            keyList3 = {"id": "rename", "quantity": fileCount}
            addonsList.append(keyList3)

        if hasBlemishReduction:
            keyList4 = {"id": "blemish-reduction", "quantity": fileCount}
            addonsList.append(keyList4)

        if hasBarcodeDetection:
            keyList5 = {"id": "barcode-detection", "quantity": fileCount}
            addonsList.append(keyList5)

        try:
            result = chargebee.Estimate.create_invoice(
                {"invoice": {
                             "customer_id": str(profile.companyProfile.uniqueID)
                            },
                 "addons": addonsList,
                 # "expires_at": self.quote_expiry_utc
                 })

            return result
        except Exception as e:
            print("create_editing_estimate error", e, type(e))
            return None

    def create_fullorder_estimate(self, profile, counters, many=False):


        # # 2D Services
        # if hasBackgroundRemoval:
        #     keyList = {"id": "background-removal", "quantity": fileCount}
        #     addonsList.append(keyList)
        #
        # # 360 Services
        # if hasAutoCropAndCenter:
        #     keyList2 = {"id": "auto-crop-and-center", "quantity": fileCount}
        #     addonsList.append(keyList2)
        #
        # if hasRename:
        #     keyList3 = {"id": "rename", "quantity": fileCount}
        #     addonsList.append(keyList3)
        #
        # if hasBlemishReduction:
        #     keyList4 = {"id": "blemish-reduction", "quantity": fileCount}
        #     addonsList.append(keyList4)
        #
        # if hasBarcodeDetection:
        #     keyList5 = {"id": "barcode-detection", "quantity": fileCount}
        #     addonsList.append(keyList5)
        if many:

            addonsList = dict()
            returnList = dict()

            for order in counters:

                try:
                    addonsList[order] = []

                    for key, value in counters[order]['counters'].items():
                        count = value['count']
                        # print(count)
                        # print(value['chargebeeServiceID'])
                        if count > 0:
                            addonsList[order].append({'id': value['chargebeeServiceID'], 'quantity': str(count)})

                    # print(addonsList[order])

                    result = chargebee.Estimate.create_invoice(
                        {"invoice": {
                            "customer_id": str(profile.companyProfile.uniqueID)
                        },
                            "addons": addonsList[order],
                            # "expires_at": self.quote_expiry_utc
                        })

                    returnList[order] = result
                    # print(result)
                except Exception as e:
                    print(e, type(e), order)

            try:

                alladdons = {}
                for addonss in  [ addonsList[order][x] for order in addonsList for x in range(0, len(addonsList[order])) ]:
                    if addonss['id'] in alladdons:
                        alladdons[addonss['id']] += int(addonss['quantity'])
                    else: alladdons[addonss['id']] = int(addonss['quantity'])

                alladdons = [ {'id': x, 'quantity': str(alladdons[x]) }  for x in alladdons]

                # print('BIG', alladdons)
                result = chargebee.Estimate.create_invoice(
                    {"invoice": {
                        "customer_id": str(profile.companyProfile.uniqueID)
                    },
                        "addons": alladdons,
                        # "expires_at": self.quote_expiry_utc
                    })

                returnList['combined'] = result
                # print('BIG', result)

            except Exception as e:
                print(e, type(e))

            return returnList, alladdons



        else:
            addonsList = []

            try:

                for key, value in counters.items():
                    count = value['count']
                    # print(count)
                    # print(value['chargebeeServiceID'])
                    if count > 0:
                        addonsList.append( {'id': value['chargebeeServiceID'], 'quantity': str(count)})

                # print(addonsList)

                result = chargebee.Estimate.create_invoice(
                    {"invoice": {
                        "customer_id": str(profile.companyProfile.uniqueID)
                    },
                        "addons": addonsList,
                        # "expires_at": self.quote_expiry_utc
                    })


                return result
            except Exception as e:
                print("create_editing_estimate error", e, type(e))
                return None



    def get_all_quotes(self,obj): pass

    def create_order(self, obj):

        try:
            result = chargebee.Order.create(
                {"id": str(obj.uniqueID), "invoice_id": self.invoice_id, "fulfillment_status": "Awaiting Shipment", })
            order = result.order
            return order
        except:
            return "Empty"

    def create_customer(self, request, obj, companyProfile):
        try:
            result = chargebee.Customer.create({
                "id": companyProfile.uniqueID,
                "first_name": obj.first_name,
                "last_name": obj.last_name,
                "email": obj.email,
                "company": companyProfile.companyName,
            })

            customer = result.customer
            return customer
        except:
            return "Empty"

    def migrateExisting(self, request, obj, companyProfile):
        # Function runs only if no customer UID match.
        try:
            result = chargebee.Customer.create({
                "id": companyProfile.uniqueID,
                "first_name": obj.first_name,
                "last_name": obj.last_name,
                "email": obj.email,
                "company": companyProfile.companyName,
            })

            customer = result.customer

        except:
            return "Empty"

        try:
            result = chargebee.Subscription.create_for_customer(companyProfile.uniqueID, {
                "id": companyProfile.subscription_id,
                "plan_id": "free-plan",
                "cf_currentstoragesize": 0,
            })
            subscription = result.subscription

        except:
            return "Empty"

        try:
            result = chargebee.Subscription.create_for_customer(companyProfile.uniqueID, {
                "plan_id": "capture-services",
            })
            subscription2 = result.subscription

        except:
            return "Empty"

    def create_Subscription(self, request, obj):
        try:
            result = chargebee.Subscription.create_for_customer(obj.uniqueID, {
                "id": obj.subscription_id,
                "plan_id": "free-plan",
                "cf_currentstoragesize": 0,

            })
            subscription = result.subscription
            return subscription
        except:
            return "Empty"

    def customerExistsCheck(self, id):
        try:
            result = chargebee.Customer.retrieve(id)
            if result:
                return True
            else:
                return False
        except chargebee.InvalidRequestError:
            return False

    def getBillingInfo(self, request, profile):

        # Perform Chargee Get Customer. If False, perform migrations, if true skip.
        # Check To See If Company Exists in Chargebee or Not

        if self.customerExistsCheck(profile.companyProfile.uniqueID) == False:
            company = profile.companyProfile
            user = profile.user
            self.migrateExisting(request, user, company)

        entries = chargebee.Subscription.list({
            "limit": 2,
            "customer_id[is]": str(profile.companyProfile.uniqueID)
        })

        # print(entries)
        return (entries)

    def getAddUserPlan(self, request, profile):
        entries = chargebee.Subscription.list({
            "limit": 1,
            "plan_id[in]": ["additional-user-plan"],
            "customer_id[is]": request.user.profile.companyProfile.uniqueID
        })
        # print(entries)
        return (entries)

    def getSubscriptionObject(self, request, subID):
        sub = chargebee.Subscription.retrieve(subID)
        # print(entries)
        return (sub)

    def getInvoicePDFObject(self, request, invoiceID):
        pdfObject = chargebee.Invoice.pdf(invoiceID)
        # print(entries)
        return (pdfObject)

    def getPlanObject(self, request, planID):
        plan = chargebee.Plan.retrieve(planID)
        # print(entries)
        return (plan)

    def createNewAddUserSubscriptionPlan(self, request, profile):

        redirect_host = ''
        prod_url = "https://" + self.subdomain + ".xspaceapp.com/#/myaccount/billing"
        dev_url = "https://dev-us1.xspaceapp.com/#/myaccount/billing"

        if self.hostname == "127.0.0.1:8000":
            redirect_host = dev_url
        else:
            redirect_host = prod_url

        result = chargebee.HostedPage.checkout_new({
            "subscription": {
                "plan_id": str(request.data["planType"]),
            },
            "customer": {
                "id": profile.companyProfile.uniqueID
            },
            "redirect_url": redirect_host
        })
        return result.hosted_page

    def updateSubscriptionPlan(self, request, subID):
        redirect_host = ''
        prod_url = "https://" + self.subdomain + ".xspaceapp.com/#/myaccount/billing"
        dev_url = "https://dev-us1.xspaceapp.com/#/myaccount/billing"

        if self.hostname == "127.0.0.1:8000":
            redirect_host = dev_url
        else:
            redirect_host = prod_url

        result = chargebee.HostedPage.checkout_existing({
            "subscription": {
                "id": subID,
                "plan_id": str(request.data["planType"]),
            },
            "redirect_url": redirect_host
        })
        return result.hosted_page

    def getCustomerPortal(self, request, profile):

        redirect_host = ''
        prod_url = "https://" + self.subdomain + ".xspaceapp.com/#/myaccount/billing"
        dev_url = "https://dev-us1.xspaceapp.com/#/myaccount/billing"

        if self.hostname == "127.0.0.1:8000":
            redirect_host = dev_url
        else:
            redirect_host = prod_url

        result = chargebee.PortalSession.create({
            "redirect_url": redirect_host,
            "customer": {
                "id": profile.companyProfile.uniqueID
            }
        })
        # print(self.hostname)
        portal_session = result.portal_session
        # print(portal_session)
        return (portal_session)

    def getCustomerInvoices(self, request, profile):
        entries = chargebee.Invoice.list({
            "limit": 100,
            "status[in]": '["paid","not_paid"]',
            "sort_by[desc]": "date",
            "customer_id[is]": profile.companyProfile.uniqueID,
        })

        for entry in entries:
            invoice = entry.invoice
            print(invoice)
        return entries

    def fullCheckout(self, profile, serviceCounts, orderIDs=None, many=False):

        # print(serviceCounts)

            if many: addonsList = serviceCounts

            else:

                addonsList = []

                for key, value in serviceCounts.items():
                    count = value['count']
                    if count > 0:
                        addonsList.append({'id': value['chargebeeServiceID'], 'quantity': str(count)})

            captureServiceSubID = chargebee.Subscription.list({
                "limit": 1,
                "plan_id[in]": ["capture-services"],
                "customer_id[is]": profile.companyProfile.uniqueID
            })

            captureSub = captureServiceSubID.__dict__['response']
            captureSubID = captureSub[0]["subscription"]["id"]

            redirect_host = ''
            prod_url = "https://" + self.subdomain + ".xspaceapp.com"
            dev_url = "http://localhost:3000"


            if self.hostname == "127.0.0.1:8000":
                redirect_host = dev_url
            else:
                redirect_host = prod_url

            try:
                result = chargebee.HostedPage.checkout_existing({
                    # Grab User's existing billing address

                    # Grab User's existing subscription plan
                    # we will redirect to live
                    "subscription": {
                        "id": captureSubID,
                        "plan_id": 'capture-services'
                    },
                    "redirect_url": redirect_host + "/#/shoppingcart",
                    "cancel_url": redirect_host + "/#/shoppingcart",
                    "addons": addonsList,
                    "pass_thru_content": ','.join(orderIDs),
                    "customer": {
                        "id": profile.companyProfile.uniqueID
                    },
                })

            except Exception as e:
                print(e, type(e))

            hosted_page = result.hosted_page
            # print hosted_page
            return hosted_page


    def editingCheckout(self, profile, fileCount, hasBackgroundRemoval, hasAutoCropAndCenter, hasRename,
                        hasBlemishReduction, hasBarcodeDetection):

        addonsList = []

        # 2D Services
        if hasBackgroundRemoval:
            keyList = {"id": "background-removal", "quantity": fileCount}
            addonsList.append(keyList)

        # 360 Services
        if hasAutoCropAndCenter:
            keyList2 = {"id": "auto-crop-and-center", "quantity": fileCount}
            addonsList.append(keyList2)

        if hasRename:
            keyList3 = {"id": "rename", "quantity": fileCount}
            addonsList.append(keyList3)

        if hasBlemishReduction:
            keyList4 = {"id": "blemish-reduction", "quantity": fileCount}
            addonsList.append(keyList4)

        if hasBarcodeDetection:
            keyList5 = {"id": "barcode-detection", "quantity": fileCount}
            addonsList.append(keyList5)

        captureServiceSubID = chargebee.Subscription.list({
            "limit": 1,
            "plan_id[in]": ["capture-services"],
            "customer_id[is]": profile.companyProfile.uniqueID
        })

        try:
            captureSub = captureServiceSubID.__dict__['response']
            print(captureSub)
            captureSubID = captureSub[0]["subscription"]["id"]
        except Exception as e:
            print(e, type(e))
        # for x in chargebee.Addon.list({ "limit": 100}):
        #     print(x)
        redirect_host = ''
        prod_url = "https://" + self.subdomain + ".xspaceapp.com/#/products/orders"
        dev_url = "http://localhost:3000/#/products/orders"

        if self.hostname == "127.0.0.1:8000":
            redirect_host = dev_url
        else:
            redirect_host = prod_url

        result = chargebee.HostedPage.checkout_existing({
            # Grab User's existing billing address

            # Grab User's existing subscription plan
            # we will redirect to live
            "subscription": {
                "id": captureSubID,
                "plan_id": 'capture-services'
            },
            "redirect_url": redirect_host,
            "cancel_url": redirect_host + "/#/order/wizard/confirm",
            "addons": addonsList,
            "customer": {
                "id": profile.companyProfile.uniqueID
            },
        })
        hosted_page = result.hosted_page
        # print hosted_page
        return hosted_page

    def checkout(self, request, profile, productList):
        twoDPhotoCount = 0
        threeSixtyPhotoCount = 0
        videoCount = 0
        threeDModelCount = 0
        singlePhotoServicesCount = 0
        fourPackServicesCount = 0
        fivePackPackServicesCount = 0
        standard360ServicesCount = 0
        integrated360ServicesCount = 0
        standard3DModelServicesCount = 0
        advanced3DModelServicesCount = 0
        cinematic3DModelServicesCount = 0
        hdProductVideoServicesCount = 0
        fourKProductServicesCount = 0
        videoWalkthroughServicesCount = 0
        addonsList = []

        for product in productList:
            # Capture Fields
            # 2D
            if product['singlePhotoServicesRequired'] == True:
                singlePhotoServicesCount = singlePhotoServicesCount + 1
            if product['fourPackServicesRequired'] == True:
                fourPackServicesCount = fourPackServicesCount + 1
            if product['fivePackPackServicesRequired'] == True:
                fivePackPackServicesCount = fivePackPackServicesCount + 1

            # 360 Models
            if product['standard360ServicesRequired'] == True:
                standard360ServicesCount = standard360ServicesCount + 1
            if product['integrated360ServicesRequired'] == True:
                integrated360ServicesCount = integrated360ServicesCount + 1

            if product['standard3DModelServicesRequired'] == True:
                standard3DModelServicesCount = standard3DModelServicesCount + 1
            if product['advanced3DModelServicesRequired'] == True:
                advanced3DModelServicesCount = advanced3DModelServicesCount + 1
            if product['cinematic3DModelServicesRequired'] == True:
                cinematic3DModelServicesCount = cinematic3DModelServicesCount + 1

            if product['hdProductVideoServicesRequired'] == True:
                hdProductVideoServicesCount = hdProductVideoServicesCount + 1
            if product['fourKProductServicesRequired'] == True:
                fourKProductServicesCount = fourKProductServicesCount + 1
            if product['videoWalkthroughServicesRequired'] == True:
                videoWalkthroughServicesCount = videoWalkthroughServicesCount + 1

        # 2D Services
        if singlePhotoServicesCount > 0:
            keyList = {"id": "single-2d-photo", "quantity": singlePhotoServicesCount}
            addonsList.append(keyList)

        if fourPackServicesCount > 0:
            keyList = {"id": "2d-four-pack", "quantity": fourPackServicesCount}
            addonsList.append(keyList)

        if fivePackPackServicesCount > 0:
            keyList = {"id": "2d-five-pack", "quantity": fivePackPackServicesCount}
            addonsList.append(keyList)

        # 360 Services
        if standard360ServicesCount > 0:
            keyList2 = {"id": "360-photo-pack", "quantity": standard360ServicesCount}
            addonsList.append(keyList2)

        if integrated360ServicesCount > 0:
            keyList3 = {"id": "int-360-photo-pack", "quantity": integrated360ServicesCount}
            addonsList.append(keyList3)

        # 3D Services
        if threeDModelCount > 0:
            keyList4 = {"id": "standard-3d-model", "quantity": threeDModelCount}
            addonsList.append(keyList4)

        if advanced3DModelServicesCount > 0:
            keyList = {"id": "Performance-3d-model", "quantity": advanced3DModelServicesCount}
            addonsList.append(keyList)

        if cinematic3DModelServicesCount > 0:
            keyList2 = {"id": "advanced-3d-model", "quantity": cinematic3DModelServicesCount}
            addonsList.append(keyList2)

        # Video Services
        if hdProductVideoServicesCount > 0:
            keyList4 = {"id": "hd-video", "quantity": hdProductVideoServicesCount}
            addonsList.append(keyList4)

        if fourKProductServicesCount > 0:
            keyList = {"id": "4k-video", "quantity": fourKProductServicesCount}
            addonsList.append(keyList)

        if videoWalkthroughServicesCount > 0:
            keyList2 = {"id": "video_walkthrough", "quantity": videoWalkthroughServicesCount}
            addonsList.append(keyList2)

        captureServiceSubID = chargebee.Subscription.list({
            "limit": 1,
            "plan_id[in]": ["capture-services"],
            "customer_id[is]": profile.companyProfile.uniqueID
        })

        captureSub = captureServiceSubID.__dict__['response']
        captureSubID = captureSub[0]["subscription"]["id"]

        redirect_host = ''
        prod_url = "https://" + self.subdomain + ".xspaceapp.com/#/products/orders"
        dev_url = "https://dev-us1.xspaceapp.com/#/products/orders"

        if self.hostname == "127.0.0.1:8000":
            redirect_host = dev_url
        else:
            redirect_host = prod_url

        result = chargebee.HostedPage.checkout_existing({
            # Grab User's existing billing address

            # Grab User's existing subscription plan
            # we will redirect to live
            "subscription": {
                "id": captureSubID,
                "plan_id": 'capture-services'
            },
            "redirect_url": redirect_host,
            "cancel_url": redirect_host + "/#/order/wizard/confirm",
            "addons": addonsList,
            "customer": {
                "id": profile.companyProfile.uniqueID
            },

        })
        hosted_page = result.hosted_page
        # print hosted_page
        return hosted_page

    def getPaidOrderIDs(self, chargebeeID):
        result = chargebee.HostedPage.retrieve(chargebeeID)
        hosted_page = result.hosted_page
        return hosted_page.pass_thru_content.split(',')



class IsNotAuthenticated(permissions.BasePermission):
    """
    Restrict access only to unauthenticated users.
    """

    def has_permission(self, request, view):
        return True


def write_asset_order_excel(ordered_data, user_request):
    print("Generating Excel")

    TAG_RE = re.compile(r'<[^>]+>')
    dirname = datetime.now().strftime('%Y.%m.%d')
    try:
        if os.path.exists(os.environ['HOME'] + '/user-order-excel'):
            if not os.path.exists(os.environ['HOME'] + '/user-order-excel/' + dirname):
                os.mkdir(os.path.join(os.environ['HOME'] + '/user-order-excel/', dirname))
                base_path = os.environ['HOME'] + '/user-order-excel/' + dirname + '/'
            else:
                base_path = os.environ['HOME'] + '/user-order-excel/' + dirname + '/'
        else:
            os.mkdir(os.path.join(os.environ['HOME'], 'user-order-excel'))
            os.mkdir(os.path.join(os.environ['HOME'] + '/user-order-excel/', dirname))
            base_path = os.environ['HOME'] + '/user-order-excel/' + dirname + '/'
    except:
        # base_path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        base_path = "/home/ubuntu/user-order-excel/"

    wb = Workbook()
    sheet = wb.active
    sheet.merge_cells('A2:B2')
    sheet.merge_cells('D2:E3')
    sheet.merge_cells('A15:M15')
    # style sheet
    text_font = Font(name='Arial', size=13, bold=True, italic=False, vertAlign=None, underline='none', strike=False,
                     color='FF000000')
    row_font = Font(name='Arial', size=11, bold=True, italic=False, vertAlign=None, underline='none', strike=False,
                    color='FF000000')
    alignment = Alignment(horizontal='center', vertical='center', text_rotation=1, wrap_text=False, shrink_to_fit=True,
                          indent=10)
    # text_alignment = Font(name='Arial',size=15,bold=True,italic=False,vertAlign=None,underline='none',strike=False,color='FF000000')
    # Customer Info Data
    sheet['A2'] = "Customer Info"
    customer_info = sheet['A2']
    customer_info.font = text_font
    customer_info.alignment = alignment
    sheet['A3'] = "Company Name"
    sheet['B3'] = ''
    sheet['A4'] = "First Name"
    sheet['B4'] = ''
    sheet['A5'] = "Last Name"
    sheet['B5'] = ''
    sheet['A6'] = "Address"
    sheet['B6'] = ''
    sheet['A7'] = "Address 2"
    sheet['B7'] = ''
    sheet['A8'] = "City"
    sheet['B8'] = ''
    sheet['A9'] = "State"
    sheet['B9'] = ''
    sheet['A10'] = "Zip Code"
    sheet['B10'] = ''
    sheet['A11'] = "Contact Number"
    sheet['B11'] = ''
    sheet['A12'] = "Contact Email"
    sheet['B12'] = ''
    sheet['A13'] = "Website"

    # Pricing
    sheet['D2'] = "Pricing"
    pricing_info = sheet['D2']
    pricing_info.font = text_font
    pricing_info.alignment = alignment
    sheet['D4'] = "4-Pack 2D Shots"
    sheet['D5'] = "5 Pack 2D Shots"
    sheet['D6'] = "360 Model View"
    sheet['D7'] = "Video"

    # Sheet Header
    sheet['A15'] = "Ordering Form"
    ordering_info = sheet['A15']
    ordering_info.font = text_font
    ordering_info.alignment = alignment
    sheet['A16'] = "Product Name"
    sheet['B16'] = "Product Description"
    sheet['C16'] = "Product Category"
    sheet['D16'] = "Product SKU"
    sheet['E16'] = "Product UPC Code"
    sheet['F16'] = "4 Pack 2D Shots"
    sheet['G16'] = "5 Pack 2D Shots"
    sheet['H16'] = "360 Model View"
    sheet['I16'] = "Video"
    sheet['J16'] = "General Order Notes"
    sheet['K16'] = "2D Order Notes"
    sheet['L16'] = "360 Order Notes"
    sheet['M16'] = "Video Order Notes"

    sheet.freeze_panes = 'A17'
    filename = user_request.user.first_name + "_" + user_request.user.last_name + "_" + datetime.now().strftime(
        '%Y.%m.%d.%H.%M.%S') + "_" + "Order"

    for idx, data in enumerate(ordered_data):
        row = 17 + idx
        sheet['A' + str(row)] = data['product_name']
        sheet['B' + str(row)] = TAG_RE.sub('', data['description'])
        sheet['C' + str(row)] = data['category_name']
        sheet['D' + str(row)] = data['Sku']
        sheet['E' + str(row)] = data['upc_code']
        sheet['F' + str(row)] = 'Yes' if data['fourPackServicesRequired'] else 'No'
        sheet['G' + str(row)] = 'Yes' if data['fourPackServicesRequired'] else 'No'
        sheet['H' + str(row)] = 'Yes' if data['standard360ServicesRequired'] else 'No'
        sheet['I' + str(row)] = 'Yes' if data['hdProductVideoServicesRequired'] else 'No'
        sheet['J' + str(row)] = ""
        sheet['K' + str(row)] = data['notes2d']
        sheet['L' + str(row)] = data['notes360view']
        sheet['M' + str(row)] = data['notesvideo']

    wb.save(base_path + filename + '.xlsx')
    # try:
    #     uploadFile(filename+'.xlsx',base_path+filename+'.xlsx','application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    # except:
    #     print("File not uploaded on drive but saved on server")


from boto3.s3.transfer import S3Transfer


def uploadTos3(filelocation, username, location):
    client = boto3.client('s3', aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
                          aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY)
    transfer = S3Transfer(client)
    try:
        # print(str(filelocation.split('/')[(len(filelocation.split('/'))-1)]))
        keyvalue = str(username) + str(location) + str(filelocation.split('/')[(len(filelocation.split('/')) - 1)])
        transfer.upload_file(filelocation, 'storage.xspaceapp.com', keyvalue)
        return {'status': 'success', 'location': keyvalue}
    except Exception as err:
        print('err')
        return 'failed to upload files'


# This is the initial Boto Folder Creation for a company
def boto3FHC(company_slug):
    bucketname = settings.AWS_STORAGE_BUCKET_NAME
    company = str(company_slug)

    # location = location.split()
    # if company exists
    # return
    # check if company location exists

    co_folders = ["products", "company_assets"]
    prod_folders = ["2D", "3D", "360", "Video", "Miscellaneous"]
    keyvalue = company + "/"
    s3 = boto3.resource('s3', aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
                        aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY)

    s3.Bucket(bucketname).put_object(Key=keyvalue)
    s3.Object(bucketname, keyvalue)

    for i in co_folders:
        subdir = company + "/" + i + "/"
        s3.Bucket(bucketname).put_object(Key=subdir)
        s3.Object(bucketname, subdir)

    return ("https://s3.amazonaws.com/" + bucketname + "/" + keyvalue)


# This is the Boto Folder Creation for a product
def boto3Pierarchy(prod_slug, company_slug):
    bucketname = settings.AWS_STORAGE_BUCKET_NAME
    company = str(company_slug)

    # location = location.split()
    # if company exists
    # return
    # check if company location exists

    # co_folders = ["products", "company_assets"]
    prod_folders = ["2D", "3D", "360", "Video", "Miscellaneous"]
    keyvalue = company + "/products/" + prod_slug + "/"
    s3 = boto3.resource('s3', aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
                        aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY)

    s3.Bucket(bucketname).put_object(Key=keyvalue)
    s3.Object(bucketname, keyvalue)

    for i in prod_folders:
        subdir = keyvalue + i + "/"
        s3.Bucket(bucketname).put_object(Key=subdir)
        s3.Object(bucketname, subdir)

    return ("https://s3.amazonaws.com/" + bucketname + "/" + keyvalue)


# This is the Boto upload file
def boto3Upload3(file, company, location, is_bytes=False, fileName=None, return_key=False):
    print("received byte object as", file)
    bucketname = settings.AWS_STORAGE_BUCKET_NAME
    filename = ""
    filedata = ""
    keyvalue = ""
    s3 = boto3.resource('s3', aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
                        aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY)
    # filepath="media"
    loc = location.split('/')
    print('loc in boto3Upload3:', loc)
    if file:
        if not is_bytes:
            filename = file.name
            # print('filename in boto3Upload3:', filename)
            filedata = file.read()
            # print('filedata in boto3Upload3:', filedata)
        else:
            filedata = file
            if not fileName:
                print("pass a filename if you are uploading a byte obect")
                return None
            filename = fileName
        if loc[0] == company:
            keyvalue = location
        else:
            keyvalue = company + location
    else:
        print('no file given to boto3Upload3')

    if filename[0] == '/':
        filename = filename[1:]
    if keyvalue[-1] == '/':
        keyvalue = keyvalue[:-1]
    key = keyvalue + '/' + filename
    # print("key inside boto3 upload")
    s3.Bucket(bucketname).put_object(Key=key, Body=filedata)
    s3.Object(bucketname, key).wait_until_exists()
    s3.Object(bucketname, key)
    if return_key:
        return ("https://" + bucketname + ".s3.amazonaws.com/" + key, key)
    else:
        return ("https://" + bucketname + ".s3.amazonaws.com/" + key)


def boto3Upload(file, username, location, pid, product_slug):
    bucketname = 'storage.xspaceapp.com'
    filename = ""
    filedata = ""
    keyvalue = ""
    pid = str(pid)
    s3 = boto3.resource('s3', aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
                        aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY)
    filepath = "media"

    if file:
        filename = file.name
        filedata = file.read()
        keyvalue = username + location

    # keyvalue = request.user.username+"/"+p_id+"/"+slug
    key = keyvalue + '/' + filename  # "".join(random.choice(valuestr) for _ in range(6))
    s3.Bucket(bucketname).put_object(Key=key, Body=filedata)
    s3.Object(bucketname, key)
    return ("https://s3.amazonaws.com/" + bucketname + "/" + key)


def boto3Upload2(file, username, location):
    print('inside boto3upload2')
    bucketname = settings.AWS_STORAGE_BUCKET_NAME
    filename = ""
    filedata = ""
    keyvalue = ""
    s3 = boto3.resource('s3', aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
                        aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY)
    filepath = "media"

    if file:
        filename = file.name
        print('got the filename as:', filename)
        filedata = file.read()
        print('got the filedata as:', filedata)
        keyvalue = username + location
        print('got the keyvalue as:', keyvalue)

    # keyvalue = request.user.username+"/"+p_id+"/"+slug
    key = keyvalue + '/' + filename  # "".join(random.choice(valuestr) for _ in range(6))
    s3.Bucket(bucketname).put_object(Key=key, Body=filedata)
    s3.Object(bucketname, key)
    return ("https://s3.amazonaws.com/" + bucketname + "/" + key)


class ProductComparison:
    def __init__(self, ):
        pass

    '''
    Assumes ProductListPayload is a e-commerce product list payload response is passed thru.
    Assumes xspace_products is a query set of product objects are passed thru.
    Assumes platform is the platform of comparing against.
    Returns list of mismatched products
    '''

    def compare(self, ProductListPayload, xspace_products, platform):
        compareList = []

        if platform == 'big-commerce':
            # TODO: [BACKEND][Dom] Convert to Map Array
            for product in ProductListPayload:
                diffObject = {'name': product.name, 'description': product.description, 'SKU': product.sku,
                              'upccode': product.upc,
                              'price': product.price, 'ecommerce_id': product.id,
                              'length': product.depth, 'width': product.width, 'height': product.height}

                for object in xspace_products:
                    didCompare = self.singleCompare(diffObject, object)
                    if didCompare == True:
                        compareList.append(object)

            if len(compareList) == 0:
                return xspace_products

                # print(diffObject)

    def singleCompare(self, diffObj, xspaceproductObj):

        matchFound = False

        if diffObj["name"] == xspaceproductObj.name:
            # Do Something
            matchFound = True
        if diffObj["description"] == xspaceproductObj.description:
            # Do Something
            matchFound = True
        if diffObj["SKU"] == xspaceproductObj.SKU:
            # Do Something
            matchFound = True
        if diffObj["upccode"] == xspaceproductObj.upccode:
            # Do Something
            matchFound = True
        # if diffObj["price"] == xspaceproductObj.price:
        #     # Do Something
        #     matchFound = True
        # if diffObj["ecommerce_id"] == xspaceproductObj.ecommerce_id:
        #     # Do Something
        #     matchFound = True

        # if diffObj.length == xspaceproductObj.length:
        #     # Do Something
        #     matchFound = True
        # if diffObj.width == xspaceproductObj.width:
        #     # Do Something
        #     matchFound = True
        # if diffObj.height == xspaceproductObj.height:
        #     # Do Something
        #     matchFound = True

        return matchFound


class ZendeskAPI():

    def __init__(self, request):

        # VH42QknYkmM2dSv8LTgiWkCseODn1ZES58iLSM52
        # An OAuth token
        self.creds = {
            'email': 'i.dallas@xspaceapp.com',
            'token': '5vE77xm5yPS8FBYtqv250fmBcPolRnJ0YLX1RJ8m',
            'subdomain': 'xapphelp'
        }

        self.zenpy_client = Zenpy(**self.creds)

        # if(request.META.get('HTTP_X_API_SUBDOMAIN')):
        #     self.subdomain = str(request.META['HTTP_X_API_SUBDOMAIN'])
        # else:
        #     self.subdomain = 'localhost:3000'

    def checkUsers(self, user):
        # TODO: [BACKEND][Dom] Add existing user check.

        x = self.zenpy_client.users(user)
        print(x.id)

    def usersOrganization(self, company, user):
        # TODO: [BACKEND][Dom] Add existing user check.
        # for user in zenpy_client.users():
        #     print(user.name)
        name = str(user.first_name) + ' ' + str(user.last_name)
        org = ZenOrganization(name=company.companyName)
        created_org = self.zenpy_client.organizations.create(org)

    def createZendeskUser(self, user):
        # TODO: [BACKEND][Dom] Add existing user check.
        # for user in zenpy_client.users():
        #     print(user.name)
        name = str(user.first_name) + ' ' + str(user.last_name)
        user = ZenUser(name=name, email=user.email)
        created_user = self.zenpy_client.users.create(user)

    def userOrgs(self, user):
        organization = self.zenpy_client.users.organizations(user)
        print(organization.name)

    def viewOrgZendeskTickets(self, company):
        # TODO: [BACKEND][Dom] Add existing user check.
        # user = self.zenpy_client.users(id=user.pk)
        # for user in zenpy_client.users():
        #     print(user.name)
        # ipdb.set_trace()

        try:
            org = self.zenpy_client.organizations(id=int(company.zendeskid))

            if org is None:
                raise zenpy.lib.exception.RecordNotFoundException
            else:
                print("Org Found on Zendesk")
                # ipdb.set_trace()
                org = org

        except zenpy.lib.exception.RecordNotFoundException as e:
            print("Org Not found")
            orgDetails = {"subscription_id": company.subscription_id, "comapny_web_site": company.companyWebsiteURL,
                          "phone_number": company.companyPhoneNumber, "company_slug": company.slug}
            org = ZenOrganization(name=company.companyName, external_id=str(company.pk), shared_comments=True,
                                  shared_tickets=True, organization_fields=orgDetails)
            created_org = self.zenpy_client.organizations.create(org)
            # print(created_org)

        tickets = self.zenpy_client.tickets()
        orgTicketList = []
        ticketList = tickets[::]
        payloadList = []

        if len(ticketList) > 0:
            for ticket in ticketList:
                if ticket.organization == org:
                    orgTicketList.append(ticket)

            # ipdb.set_trace()

            for orgTic in orgTicketList:
                ticketPayload = {"id": orgTic.id, "subject": orgTic.subject, "created_date": orgTic.created_at,
                                 "status": orgTic.status.capitalize(), "type": "Incident"}
                commentList = []
                for comment in self.zenpy_client.tickets.comments(ticket=orgTic.id):
                    commentObject = {"author": comment.author.email, "body": comment.body,
                                     "created_date": comment.created_at}
                    commentList.append(commentObject)
                    # print(comment.body)
                ticketPayload["comments"] = commentList
                payloadList.append(ticketPayload)
        else:
            payloadList = []

        return payloadList

    def createZendeskOrganization(self, company, user):
        # TODO: [BACKEND][Dom] Add existing user check.
        # for user in zenpy_client.users():
        #     print(user.name)

        # name = str(user.first_name)+' '+str(user.last_name)
        org = ZenOrganization(name=company.companyName)
        created_org = self.zenpy_client.organizations.create(org)

        name = str(user.first_name) + ' ' + str(user.last_name)
        user = ZenUser(name=name, email=user.email, organization_id=created_org.id)
        created_user = self.zenpy_client.users.create(user)

    def checkOrganizations(self):
        # TODO: [BACKEND][Dom] Add existing user check.
        for org in self.zenpy_client.organizations():
            print(org.name)

    def createZendeskTicket(self, user, company, prod, subject, msg):
        # TODO: [BACKEND][Dom] Add existing user check.
        # for user in zenpy_client.users():
        #     print(user.name)
        # ipdb.set_trace()

        # uid = us.organization()

        # organization_id = xxx
        # created_ticket = self.zenpy_client.tickets.create(external_id=external_id, organization=organization)

        try:
            org = self.zenpy_client.organizations(id=int(company.zendeskid))

            if org is None:
                raise zenpy.lib.exception.RecordNotFoundException
            else:
                print("Org Found on Zendesk")
                # ipdb.set_trace()
                org = org

        except zenpy.lib.exception.RecordNotFoundException as e:
            print("Org Not found")
            orgDetails = {"subscription_id": company.subscription_id, "comapny_web_site": company.companyWebsiteURL,
                          "phone_number": company.companyPhoneNumber, "company_slug": company.slug}
            org = ZenOrganization(name=company.companyName, external_id=str(company.pk), shared_comments=True,
                                  shared_tickets=True, organization_fields=orgDetails)
            created_org = self.zenpy_client.organizations.create(org)
            # print(created_org)

        name = str(user.first_name) + ' ' + str(user.last_name)
        xxx = ZenUser(name=name, email=user.email)
        # coop =  self.zenpy_client.users.organizations(xxx)
        descrip = "Company: " + company.companyName + " | Company Slug: " + company.slug + " | Product Name: " + prod.name + " | Product Slug: " + prod.slug + "| Filename: " + subject + " | Message: " + msg

        x = self.zenpy_client.tickets.create(Ticket(description=descrip, raw_subject=subject, organization=org,
                                                    requester=ZenUser(name=name, email=user.email)))
        tick = x.ticket
        return tick.id

    def createZendeskTicket2(self, user, company, subject, type, msg):
        # TODO: [BACKEND][Dom] Add existing user check.
        # for user in zenpy_client.users():
        #     print(user.name)
        # ipdb.set_trace()

        # uid = us.organization()

        # organization_id = xxx
        # created_ticket = self.zenpy_client.tickets.create(external_id=external_id, organization=organization)

        org = self.zenpy_client.organizations(id=int(company.zendeskid))
        # zenuser = self.zenpy_client.users(id=user.zendeskid)

        if org is None:
            raise zenpy.lib.exception.RecordNotFoundException
        else:
            print("Org Found on Zendesk")
            # ipdb.set_trace()
            org = org

        name = str(user.first_name) + ' ' + str(user.last_name)
        xxx = ZenUser(name=name, email=user.email)
        # coop =  self.zenpy_client.users.organizations(xxx)

        x = self.zenpy_client.tickets.create(Ticket(description=msg, raw_subject=subject, organization=org,
                                                    requester=ZenUser(name=name, email=user.email), type=type))
        tick = x.ticket
        return tick.id

    def viewTicketCom(self, ticketID):
        for comment in self.zenpy_client.tickets.comments(ticket=ticketID):
            print(comment.body)

    def ticketComment(self, ticket_id, user, message):
        # TODO: [BACKEND][Dom] Add existing user check.
        name = str(user.first_name) + ' ' + str(user.last_name)

        p = ZenUser(name=name, email=user.email)
        ticket = self.zenpy_client.tickets(id=ticket_id)
        ticket.comment = Comment(body=message, author_id=p.id, public=True)
        self.zenpy_client.tickets.update(ticket)

    def zU(self, user):
        name = str(user.first_name) + ' ' + str(user.last_name)
        p = ZenUser(name=name, email=user.email)

    def upZen(self, company, user, prod):

        try:
            org = self.zenpy_client.organizations(external_id=str(company.pk))

            if len(org) == 0:
                raise zenpy.lib.exception.RecordNotFoundException
                # (1) create Organization 
                # (2) add User (or all company users) to Organization

            else:
                # ipdb.set_trace()
                print("Org Found on Zendesk")
                org = org[::]
                ei = org[0]
                organization_id = ei.id
                ex_id = ei.external_id
                name = str(user.first_name) + ' ' + str(user.last_name)
                descrip = company.companyName + " Product ID: " + prod.name + " Product Thread"

                x = self.zenpy_client.tickets.create(
                    Ticket(description=descrip, requester=ZenUser(name=name, email=user.email), external_id=ex_id))
                tick = x.ticket
                return tick.id
        except:
            user = self.zenpy_client.users.me()
            print(user.to_dict(serialize=True))
        # pass user check
