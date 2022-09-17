"""
Django Contrib & Core
"""
from django.core.paginator import Paginator
from django.db.models import Q
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.contrib.auth import get_user_model
from django.utils.datastructures import MultiValueDictKeyError
from django.views.generic.detail import DetailView
from django.template import Context
from django.template.loader import render_to_string, get_template
from django.views.decorators.csrf import csrf_exempt
from itertools import chain
from django.db.models import Q
from django.http import HttpResponse
# DRF
import oauth2 as oauth
# Django REST Framework
from rest_framework import viewsets, mixins, parsers, renderers, status, generics
from rest_framework import generics
from rest_framework.permissions import IsAdminUser, IsAuthenticated
from rest_framework.response import Response
from rest_framework.authentication import SessionAuthentication, BasicAuthentication, TokenAuthentication
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework_jwt.views import obtain_jwt_token
from rest_framework import status
from rest_framework.decorators import api_view
from rest_framework import authentication, permissions
from rest_framework.renderers import JSONRenderer, MultiPartRenderer
from rest_framework.parsers import JSONParser, MultiPartParser, FileUploadParser

from django.template.defaultfilters import slugify
from autoslug import AutoSlugField

# Other
import csv
import urllib.request, urllib.parse, urllib.error, json
import requests
from app.core.utility import iterate_and_save_products, AuthenticatedUserMixin, IsNotAuthenticated, ChargebeeAPI, \
    write_asset_order_excel, has_permission, boto3Upload3
from app.core.utility import master_permissions_list as P
from app.core.QRGenerator import generateProductOrderQRCode, GenerateQRLabelDoc
from elasticsearch import *
import requests
import random
import string
import ast
import base64
import os
from os.path import split, splitext
import datetime
import boto3
from wsgiref.util import FileWrapper

# import simplejson as json
import requests
import sys
from requests.exceptions import ConnectionError
import openpyxl
# from constants import product_search,API2CART_SUCCESS,API2CART_FAILED,
#      UPDATE_CREDS_SUCCESS,UPDATE_CREDS_FAILED,API2CART_DOMAIN,VALIDATE_CART_API,CREATE_CART_API
from app.imageEditingFunctions.manipulateImage import load_image
from PIL import Image
import io

from xapi.settings import AWS_ACCESS_KEY_ID, AWS_SECRET_ACCESS_KEY, AWS_STORAGE_BUCKET_NAME

try:
    import simplejson as json
except ImportError:
    import json
from dateutil.tz import tzutc
import traceback

UTC = tzutc()

from app.accounts.serializers import EmailThread, CompanyUpdateSerializer, AddressSerializer, \
    CompanySerializer, InvitationSerializer, UserDetailSerializer, Roll_phone_UpdateSerializer, PasswordSerializer, \
    NameUpdateSerializer
from app.products.serializers import ManualProductCreationSerializer, ProductSerializer, ProductMessageSerializer, \
    ProductMessageThreadCreateSerializer, ProductMessageThreadListSerializer, UploaderSerializer, ProductListSerializer
from app.orders.serializers import EditOrderSerializer, OrderStatusUpdateSerializer, OrderUpdateSerializer, \
    ProductOrderSerializer, \
    ProductOrderListSerializer, Order, CreateOrderSerializer, FullOrderSerializer, ContentStandardCreateSerializer, \
    ContentStandardListSerializer, BatchOrderSerializer, OrderSerializer, OrderTrackingSerializer, \
    ProductOrderUpdateSerializer, processEditingOrder
from app.orders.models import ContentStandard, Order, OrderTracking, ProductOrder, FullOrder
from app.core.models import UploadAsset
from app.core.utility import boto3Upload2, boto3Upload, uploadTos3
from app.products.models import Product, ProductAsset, ProductMessage, ProductMessageThread, AssetContentStandards
from app.products.views import get_all_prods_for_sub_companies
from app.accounts.models import User, Profile, Company
import time
from notifications.signals import notify


class ContentStandardListAPI(APIView):
    queryset = ContentStandard.objects.all()
    permission_classes = (IsAuthenticated,)
    # authentication_classes = (TokenAuthentication)
    renderer_classes = (renderers.JSONRenderer,)
    serializer_class = ContentStandardListSerializer

    @has_permission(P.READ_ONLY)
    def get(self, request):
        # standardsList = ContentStandard.objects.all()
        # serializer = ContentStandardListSerializer(standardsList, many=True)
        return Response({'success': 'None', 'error': 'No 2D image found'})
        # return Response(serializer.data)

    @csrf_exempt
    def post(self, request, *args, **kwargs):
        serializer = ContentStandardListSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class ContentStandardAPI(AuthenticatedUserMixin, APIView):
    queryset = ContentStandard.objects.all()
    permission_classes = (IsAuthenticated,)
    # authentication_classes = (TokenAuthentication)
    renderer_classes = (renderers.JSONRenderer,)
    serializer_class = ContentStandardCreateSerializer

    @has_permission(P.READ_ONLY)
    def get(self, request):
        print('ContentStandardAPI GET received:', request.data, request.user)

        profile = Profile.objects.get(user=request.user)
        # we have to check the new way and the old way that people created content standards as of 06/18/2020
        concatList = ContentStandard.objects.filter(
            Q(company_id=profile.companyProfile) | Q(profile=profile)).distinct()

        # remove duplicates
        companyNameList = []
        nameList = []
        count = 0
        if concatList.count() > 0:
            for x in concatList:
                if x.company_id in companyNameList:

                    if x.name in nameList:
                        # remove it
                        if count == concatList.count():
                            concatList = concatList[:count]
                        else:
                            concatList = list(chain(concatList[:count], concatList[(count + 1):]))
                    else:
                        nameList += x.name
                else:
                    companyNameList.append(x.company_id)
                    nameList.append(x.name)
                count += 1
            serializer = ContentStandardCreateSerializer(concatList, many=True)
            return Response({'contentStandard': serializer.data}, status=status.HTTP_200_OK)
            # return Response({'API Success': 'None',}, status=status.HTTP_200_OK)
        else:
            lis = []
            return Response({'contentStandard': lis}, status=status.HTTP_200_OK)
            # return Response({'API Error': 'No Content Standards found.'}, status=status.HTTP_404_NOT_FOUND)

    @csrf_exempt
    def post(self, request, *args, **kwargs):
        profile = Profile.objects.get(user=request.user)
        serializer = ContentStandardCreateSerializer(data=request.data, context={"profile": profile})
        print((request.data))
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class AuditAPI(AuthenticatedUserMixin, APIView):
    permission_classes = (IsAuthenticated,)
    renderer_classes = (renderers.JSONRenderer,)

    @csrf_exempt
    def post(self, request, *args, **kwargs):
        """
        this function is to be called when an audit is to be run on a company account for each product it checks the
        prdouct level and product asset level content standards
        :param request:
        :param args:
        :param kwargs:
        :return:
        """
        # generally we want to follow the following steps:
        # 1. get the admin information
        # 2. get all of the products, or if a list of products/ specific product is given, just the specified products
        # 3. get either the assigned content standards or the specified content standards within the request
        # 4. iterate over each product at the product level standards followed by the product asset level standards
        print('!!! received:', request.data)
        profile = Profile.objects.get(user=request.user)
        company = Company.objects.get(id=profile.companyProfile.id)
        if request.data['products'] == 'all':
            # grab all products
            product_list = get_all_prods_for_sub_companies(profile.companyProfile)
        elif request.data['products'] != '':
            if isinstance(request.data['products'], list):
                if len(request.data['products']) == 0:
                    return Response(status=status.HTTP_400_BAD_REQUEST)
                # grab the list of products
                product_list = request.data['products']
                count = 0
                for product in product_list:
                    product_list[count] = Product.objects.get(slug=product['xspaceid'])
                    count += 1
        else:
            print("No products selected for audit")
            return Response(status=status.HTTP_400_BAD_REQUEST)

        if request.data['Content Standard'] == ['']:
            applied_content_standards_only = True
        else:
            applied_content_standards_only = False

        # get the list of content standards if we're not analyzing by pre-existing
        content_standard_list = []
        if applied_content_standards_only is False:
            for cs in request.data['Content Standard']:
                temp_cs = ContentStandard.objects.get(uniqueID=cs['uniqueID'])
                content_standard_list.append(temp_cs)
            content_standard_list = list(set(content_standard_list))
            del temp_cs

        # initialize error collection holders
        error_narrative = ''
        error_list = []
        error_product_list = []
        error_product_asset_list = []
        backgroundColorDict = {"1": 'transparent',
                               "2": 'white',
                               "3": 'black',
                               "4": 'unknown',
                               "5": 'custom'}

        print('grabbed products')

        for product in product_list:

            # get the products assets
            product_assets = ProductAsset.objects.filter(product=product.id).filter(is_thumbnail=False).filter(
                is_mask=False)

            # separate the assets by type
            image_product_assets = product_assets.filter(assetType=1)
            spin_product_assets = product_assets.filter(assetType=2)
            video_product_assets = product_assets.filter(assetType=3)
            model_product_assets = product_assets.filter(assetType=4)
            asset_type_list = [image_product_assets, spin_product_assets, video_product_assets,
                               model_product_assets]

            if applied_content_standards_only:
                acs = AssetContentStandards.objects.filter(product=product.id)
                if int(acs.count()) > 0:
                    for cs in acs:
                        content_standard_list.append(ContentStandard.objects.get(uniqueID=cs.contentStandard.uniqueID))
                    # remove duplicates
                    content_standard_list = list(set(content_standard_list))
                else:
                    print('no preexisting content standards found for the specified product')
                    error_list.append(
                        {'value': 'No preexisting content standards found.',
                         'upc': product.upccode,
                         'slug': product.slug,
                         'name': product.name,
                         'asset name': '',
                         'type': 'product error',
                         })
                    error_product_list.append(product)
                    error_narrative += ' %s No preexisting content standards found.' % (
                        str(product.uniqueID))
                    continue

            # print("product_list, content_standard_list", product_list, content_standard_list)

            print('grabbed assets and standards')

            for content_standard in content_standard_list:
                # evaluate product at the product content standard level first

                # check number of assets
                if content_standard.strictnessLevel == 'strict':
                    local_count = 0
                    for asset_type in asset_type_list:
                        if local_count == 0:
                            asset_type_name = 'images'
                            asset_count_ref = content_standard.image_number
                        elif local_count == 1:
                            asset_type_name = '360s'
                            asset_count_ref = content_standard.threeSixtyImage_number
                        elif local_count == 2:
                            asset_type_name = 'videos'
                            asset_count_ref = content_standard.video_number
                        elif local_count == 3:
                            asset_type_name = '3D models'
                            asset_count_ref = content_standard.video_number
                        else:
                            # don't check misc files
                            continue

                        if asset_count_ref == '' or asset_count_ref == 'null' or asset_count_ref is None:
                            asset_count_ref = 0

                        if int(asset_count_ref) > int(0):
                            if int(asset_type.count()) != int(asset_count_ref):
                                error_list.append(
                                    {'value': 'Invalid Number of %s. Expected %s but counted %s.' % (
                                        asset_type_name, str(int(asset_count_ref)),
                                        str(int(image_product_assets.count()))),
                                     'upc': product.upccode,
                                     'slug': product.slug,
                                     'name': product.name,
                                     'asset name': '',
                                     'type': 'product error'})
                                error_product_list.append(product)
                                error_narrative += ' %s Invalid Number of %s. Expected %s but counted %s.' % (
                                    str(product.uniqueID),
                                    str(asset_type_name),
                                    str(int(asset_count_ref)),
                                    str(int(image_product_assets.count())))
                        local_count += 1

                else:
                    local_count = 0
                    for asset_type in asset_type_list:
                        if local_count == 0:
                            asset_type_name = 'images'
                            asset_count_ref = content_standard.image_number
                        elif local_count == 1:
                            asset_type_name = '360s'
                            asset_count_ref = content_standard.threeSixtyImage_number
                        elif local_count == 2:
                            asset_type_name = 'videos'
                            asset_count_ref = content_standard.video_number
                        elif local_count == 3:
                            asset_type_name = '3D models'
                            asset_count_ref = content_standard.video_number
                        else:
                            # don't check misc files
                            continue

                        if asset_count_ref == '' or asset_count_ref == 'null' or asset_count_ref is None:
                            asset_count_ref = 0

                        if int(asset_count_ref) > int(0):
                            if int(asset_type.count()) < int(asset_count_ref):
                                error_list.append(
                                    {'value': 'Invalid Number of %s. Expected %s but counted %s.' % (
                                        asset_type_name, str(int(asset_count_ref)),
                                        str(int(image_product_assets.count()))),
                                     'upc': product.upccode,
                                     'slug': product.slug,
                                     'name': product.name,
                                     'asset name': '',
                                     'type': 'product error'})
                                error_product_list.append(product)
                                error_narrative += ' %s Invalid Number of %s. Expected %s but counted %s.' % (
                                    str(product.uniqueID),
                                    str(asset_type_name),
                                    str(int(asset_count_ref)),
                                    str(int(image_product_assets.count())))
                        local_count += 1

                print('end product compairson')

                # ################## END PRODUCT COMPARISON ~~~ START PRODUCT ASSET COMPARISON ################ #
                # evaluate product at the product asset content standard level
                #
                # list of things we are checking in this for loop
                #                     # - assets are square
                #                     # - file types**
                #                     # - asset sizes**
                #                     # - file sizes**
                #                     # - background was removed by XSPACE verified source
                #                     # - if background was removed by us; are the margins accurate?**
                #                     # - background color
                #                     # - file naming
                #                     # **lazy strictness will relax or not check the tolerances on these fields

                # print('made it past the product level comparison')

                s3 = boto3.client('s3', aws_access_key_id=AWS_ACCESS_KEY_ID,
                                  aws_secret_access_key=AWS_SECRET_ACCESS_KEY)
                for product_asset in product_assets:
                    if product_asset.assetType == 1:
                        # image/2D specific logic
                        # get file type
                        filetype = str(splitext(product_asset.file_name)[1]).lower()

                        # open asset
                        original_img = load_image(product_asset.url)
                        original_img = Image.open(io.BytesIO(original_img))
                        # get width and height
                        original_img_width, original_img_height = original_img.size

                        # get background type
                        original_img_background_type = int(product_asset.asset_background_type)

                        # get admin info on the asset
                        key = (str(company.slug) + '/products/' + str(product.slug) + '/2D/' +
                               str(product_asset.file_name))
                        response = s3.head_object(Bucket=AWS_STORAGE_BUCKET_NAME, Key=key)
                        # get file size
                        original_img_filesize = response['ContentLength'] / 1000
                        # background removed by XSPACE
                        original_img_bg_removed = product_asset.verified_background_removed

                        # content standard references
                        filetype_ref = str(content_standard.image_file_type).lower()
                        filesize_ref = content_standard.image_ideal_file_size
                        if filesize_ref is None or filesize_ref == '':
                            filesize_ref = None
                        else:
                            filesize_ref = int(filesize_ref)
                        filewidth_ref = int(content_standard.image_width)
                        fileheight_ref = int(content_standard.image_height)
                        filebackground_ref = content_standard.image_background_type
                        # Background identifiers: 1 = transparent; 2 = white; 3 = black; 4 = unknown; 5 = custom
                        if filebackground_ref == 'White':
                            filebackground_ref = int(2)
                        elif filebackground_ref == 'Transparent':
                            filebackground_ref = int(1)
                        elif filebackground_ref == 'Black':
                            filebackground_ref = int(3)
                        elif filebackground_ref == 'Custom':
                            filebackground_ref = int(5)
                        else:
                            filebackground_ref = int(4)
                        fileverified_bg_removed = content_standard.verified_background_removed

                        print('grabbed image info')
                        # print('grabbed necessary image reqs as:', filetype, original_img_width, original_img_height,
                        #       original_img_background_type, key, original_img_filesize, original_img_bg_removed)
                    elif product_asset.assetType == 2:
                        # 360 specific logic
                        # get file type
                        filetype = str(splitext(product_asset.file_name)[1]).lower()

                        # open asset but use the base image as a quick reference
                        original_img = load_image(product_asset.baseImage)
                        original_img = Image.open(io.BytesIO(original_img))

                        # get width and height (use the referenced base image for a quick check... later we can check
                        #                       each individual image)
                        original_img_width, original_img_height = original_img.size

                        # get background type (use the referenced base image for a quick check... later we can check
                        #                      each individual image)
                        original_img_background_type = int(product_asset.asset_background_type)

                        # get admin info on the asset
                        subkey = split(split(str(product_asset.url))[0])[1]
                        key = (str(company.slug) + '/products/' + str(product.slug) + '/360/' + '/' +
                               subkey)
                        response = s3.head_object(Bucket=AWS_STORAGE_BUCKET_NAME, Key=key)

                        # get file size
                        # probably need to iterate
                        original_img_filesize = response['ContentLength']

                        # content standard references
                        filetype_ref = str(content_standard.image_file_type).lower()
                        filesize_ref = content_standard.image_ideal_file_size
                        if filesize_ref is None or filesize_ref == '':
                            filesize_ref = None
                        else:
                            filesize_ref = int(filesize_ref)
                        filewidth_ref = int(content_standard.image_width)
                        fileheight_ref = int(content_standard.image_height)
                        filebackground_ref = content_standard.image_background_type
                        # Background identifiers: 1 = transparent; 2 = white; 3 = black; 4 = unknown; 5 = custom
                        if filebackground_ref == 'White':
                            filebackground_ref = int(2)
                        elif filebackground_ref == 'Transparent':
                            filebackground_ref = int(1)
                        elif filebackground_ref == 'Black':
                            filebackground_ref = int(3)
                        elif filebackground_ref == 'Custom':
                            filebackground_ref = int(5)
                        else:
                            filebackground_ref = int(4)
                        fileverified_bg_removed = content_standard.verified_background_removed

                        print('grabbed 360 info')
                    elif product_asset.assetType == 3:
                        # video specific logic
                        # get file type
                        # TODO: videos
                        filetype = str(splitext(product_asset.file_name)[1]).lower()
                        filetype_ref = str(content_standard.video_file_type).lower()
                        filesize_ref = content_standard.image_ideal_file_size
                        if filesize_ref is None or filesize_ref == '':
                            filesize_ref = None
                        else:
                            filesize_ref = int(filesize_ref)
                        filewidth_ref = int(content_standard.image_width)
                        fileheight_ref = int(content_standard.image_height)
                        filebackground_ref = content_standard.image_background_type
                        # Background identifiers: 1 = transparent; 2 = white; 3 = black; 4 = unknown; 5 = custom
                        if filebackground_ref == 'White':
                            filebackground_ref = int(2)
                        elif filebackground_ref == 'Transparent':
                            filebackground_ref = int(1)
                        elif filebackground_ref == 'Black':
                            filebackground_ref = int(3)
                        elif filebackground_ref == 'Custom':
                            filebackground_ref = int(5)
                        else:
                            filebackground_ref = int(4)
                        fileverified_bg_removed = content_standard.verified_background_removed

                    elif product_asset.assetType == 4:
                        # 3D specific logic
                        # get file type
                        # TODO: 3D
                        filetype = str(splitext(product_asset.file_name)[1]).lower()
                        filetype_ref = str(content_standard.threeD_file_type).lower()
                    else:
                        # misc specific logic
                        # set variables to None to help prevent bleed over from previous logic
                        filetype = None
                        filetype_ref = None
                        filesize_ref = None
                        filewidth_ref = None
                        fileheight_ref = None
                        filebackground_ref = None
                        fileverified_bg_removed = None

                    # print('checking file type')
                    if filetype is not None:
                        # check file type
                        if content_standard.strictnessLevel == 'strict':
                            if filetype != filetype_ref:
                                error_list.append(
                                    {'value': 'Invalid File Type. Expected %s but received %s.' % (
                                        str(filetype_ref), str(filetype)),
                                     'upc': product.upccode,
                                     'slug': product.slug,
                                     'name': product.name,
                                     'asset name': product_asset.file_name,
                                     'type': 'product asset error'})
                                error_product_list.append(product)
                                error_product_asset_list.append(product_asset)
                                error_narrative += ' %s Invalid File Type %s. Expected %s but counted %s.' % (
                                    str(product_asset.uniqueID),
                                    str(filetype),
                                    str(filetype_ref),
                                    str(filetype))

                        # check file size (in kB)
                        # print('checking file size')
                        passFlag = True
                        if filesize_ref is not None:
                            if content_standard.strictnessLevel == 'strict':
                                if filesize_ref < original_img_filesize:
                                    passFlag = False
                            else:
                                # we are lenient with file sizing with the 'lazy' flag
                                if filesize_ref * 1.5 < original_img_filesize:
                                    passFlag = False

                            if passFlag is False:
                                error_list.append(
                                    {'value': 'Invalid File Size. Expected %s kilobytes'
                                              ' but received %s kilobytes.' % (
                                                  str(filesize_ref), str(original_img_filesize)),
                                     'upc': product.upccode,
                                     'slug': product.slug,
                                     'name': product.name,
                                     'asset name': product_asset.file_name,
                                     'type': 'product asset error'})
                                error_product_list.append(product)
                                error_product_asset_list.append(product_asset)
                                error_narrative += ' %s Invalid File Size: %s kiloBytes. Expected %s but measured %s.' % (
                                    str(product_asset.uniqueID),
                                    str(original_img_filesize),
                                    str(filesize_ref),
                                    str(original_img_filesize))

                        # check asset width
                        # print('checking asset width')
                        passFlag = True
                        if content_standard.strictnessLevel == 'strict':
                            if original_img_width != filewidth_ref:
                                passFlag = False
                        else:
                            # we need to check if the width is at least that amount
                            if original_img_width < filewidth_ref:
                                passFlag = False

                        if passFlag is False:
                            error_list.append(
                                {'value': 'Invalid Asset Width. Expected %s pixels'
                                          ' but received %s pixels.' % (
                                              str(filewidth_ref), str(original_img_width)),
                                 'upc': product.upccode,
                                 'slug': product.slug,
                                 'name': product.name,
                                 'asset name': product_asset.file_name,
                                 'type': 'product asset error'})
                            error_product_list.append(product)
                            error_product_asset_list.append(product_asset)
                            error_narrative += ' %s Invalid Asset Width: %s pixels. Expected %s pixels ' \
                                               ' but measured %s pixels.' % (
                                                   str(product_asset.uniqueID),
                                                   str(original_img_width),
                                                   str(filewidth_ref),
                                                   str(original_img_width))

                        # check asset height
                        # print('checking asset height')
                        passFlag = True
                        if content_standard.strictnessLevel == 'strict':
                            if original_img_height != fileheight_ref:
                                passFlag = False
                        else:
                            # we need to check if the height is at least that amount
                            if original_img_height < fileheight_ref:
                                passFlag = False

                        if passFlag is False:
                            error_list.append(
                                {'value': 'Invalid Asset height. Expected %s pixels but received'
                                          ' %s pixels.' % (
                                              str(fileheight_ref), str(original_img_height)),
                                 'upc': product.upccode,
                                 'slug': product.slug,
                                 'name': product.name,
                                 'asset name': product_asset.file_name,
                                 'type': 'product asset error'})
                            error_product_list.append(product)
                            error_product_asset_list.append(product_asset)
                            error_narrative += ' %s Invalid Asset height: %s pixels. Expected %s pixels but' \
                                               ' measured %s pixels.' % (
                                                   str(product_asset.uniqueID),
                                                   str(original_img_height),
                                                   str(fileheight_ref),
                                                   str(original_img_height))

                        # check square-iness
                        # print('checking asset square-iness')
                        passFlag = True
                        if content_standard.force_asset_square is True:
                            if original_img_height != original_img_width:
                                passFlag = False

                        if passFlag is False:
                            error_list.append(
                                {'value': 'Invalid Squared Asset. Received (%s px [width], %s px [height]).' % (
                                    str(original_img_width), str(original_img_height)),
                                 'upc': product.upccode,
                                 'slug': product.slug,
                                 'name': product.name,
                                 'asset name': product_asset.file_name,
                                 'type': 'product asset error'})
                            error_product_list.append(product)
                            error_product_asset_list.append(product_asset)
                            error_narrative += ' %s Invalid Asset Is Not Square. Measured (%s px [width], %s px [' \
                                               'height]).' % (
                                                   str(product_asset.uniqueID),
                                                   str(original_img_width),
                                                   str(original_img_height))

                        # check background was removed
                        # print('checking background was removed')
                        passFlag = True
                        if fileverified_bg_removed is True:
                            if original_img_bg_removed is False:
                                passFlag = False

                        if passFlag is False:
                            error_list.append(
                                {'value': "Asset's background was not removed by a verified source.",
                                 'upc': product.upccode,
                                 'slug': product.slug,
                                 'name': product.name,
                                 'asset name': product_asset.file_name,
                                 'type': 'product asset error'})
                            error_product_list.append(product)
                            error_product_asset_list.append(product_asset)
                            error_narrative += " %s Invalid Asset background. Asset's background was not removed by a " \
                                               "verified source." % (
                                                   str(product_asset.uniqueID))

                        # check background type
                        # print('checking background type')
                        passFlag = True
                        if filebackground_ref != original_img_background_type:
                            passFlag = False

                        if passFlag is False:
                            # Background identifiers: 1 = transparent; 2 = white; 3 = black; 4 = unknown; 5 = custom
                            backgroundColor_ref = backgroundColorDict.get(str(filebackground_ref))
                            backgroundColor = backgroundColorDict.get(str(original_img_background_type))
                            error_list.append(
                                {'value': "Invalid Asset background color. Expected %s,"
                                          " but received %s" % (str(backgroundColor_ref),
                                                                str(backgroundColor)),
                                 'upc': product.upccode,
                                 'slug': product.slug,
                                 'name': product.name,
                                 'asset name': product_asset.file_name,
                                 'type': 'product asset error'})
                            error_product_list.append(product)
                            error_product_asset_list.append(product_asset)
                            error_narrative += " %s Invalid Asset background color. Expected %s, but received %s" % (
                                str(product_asset.uniqueID), str(backgroundColor_ref),
                                str(backgroundColor))

                        # check file naming.... BUT I DONT WANT TOOOOOOOOOOOOO.... you owe me a beer
                        # TODO: noticed that there's no way to add your own custom string to the filename
                        # print('checking file naming')
                        passFlag = True
                        if content_standard.strictnessLevel == 'strict':
                            count = 0
                            filename = splitext(product_asset.file_name)[0]
                            if content_standard.filename_delimiter not in filename:
                                print(content_standard.filename_delimiter, 'not found in ', filename)
                                error_list.append(
                                    {'value': "Invalid File Name Delimiter."
                                              " Expected %s" % (
                                                  str(content_standard.filename_delimiter)),
                                     'upc': product.upccode,
                                     'slug': product.slug,
                                     'name': product.name,
                                     'asset name': product_asset.file_name,
                                     'type': 'product asset error'})
                                error_product_list.append(product)
                                error_product_asset_list.append(product_asset)
                                error_narrative += " %s Invalid File Name Delimiter. Expected %s" % (
                                    str(product_asset.uniqueID), str(content_standard.filename_delimiter))

                                # check other variants
                                if content_standard.filename_delimiter == '-':
                                    if '_' in filename:
                                        print('_ found in ', filename, ' using it instead')
                                        filename_parts = filename.split('_')
                                    else:
                                        print('_ not found in ', filename)
                                        if ' ' in filename:
                                            print("' ' found in ", filename, " using it instead")
                                            filename_parts = filename.split(' ')
                                        else:
                                            filename_parts = filename
                                else:
                                    if '-' in filename:
                                        print('- found in ', filename, ' using it instead')
                                        filename_parts = filename.split('-')
                                    else:
                                        print('- not found in ', filename)
                                        if ' ' in filename:
                                            print("' ' found in ", filename, " using it instead")
                                            filename_parts = filename.split(' ')
                                        else:
                                            filename_parts = filename
                            else:
                                filename_parts = filename.split(content_standard.filename_delimiter)

                            # check prefix
                            if content_standard.filename_prefix != '$NONE':
                                if content_standard.filename_prefix == '$PRODUCT_NAME':
                                    if filename_parts[count] != product.name:
                                        ref_value = product.name
                                        passFlag = False
                                elif content_standard.filename_prefix == '$PRODUCT_SKU':
                                    if filename_parts[count] != product.SKU:
                                        ref_value = product.SKU
                                        passFlag = False
                                elif content_standard.filename_prefix == '$PRODUCT_UPC':
                                    if filename_parts[count] != product.upccode:
                                        ref_value = product.upccode
                                        passFlag = False
                                elif content_standard.filename_prefix == '$XSPACE_ID':
                                    if filename_parts[count] != product.uniqueID:
                                        ref_value = product.uniqueID
                                        passFlag = False

                                if passFlag is False:
                                    error_list.append(
                                        {'value': "Invalid File Name Prefix. Expected %s,"
                                                  " but received %s" % (str(ref_value),
                                                                        str(filename_parts[count])),
                                         'upc': product.upccode,
                                         'slug': product.slug,
                                         'name': product.name,
                                         'asset name': product_asset.file_name,
                                         'type': 'product asset error'})
                                    error_product_list.append(product)
                                    error_product_asset_list.append(product_asset)
                                    error_narrative += " %s Invalid File Name Prefix. Expected %s, but received %s" % (
                                        str(product_asset.uniqueID), str(ref_value),
                                        str(filename_parts[count]))
                                count += 1
                            if content_standard.filename_base != '$NONE':
                                try:
                                    if content_standard.filename_base == '$NAME':
                                        if filename_parts[count] != product.name:
                                            ref_value = product.name
                                            passFlag = False
                                    elif content_standard.filename_base == '$UPC':
                                        if filename_parts[count] != product.upccode:
                                            ref_value = product.upccode
                                            passFlag = False
                                    elif content_standard.filename_base == '$SKU':
                                        if filename_parts[count] != product.SKU:
                                            ref_value = product.SKU
                                            passFlag = False
                                    elif content_standard.filename_base == '$XSPACE_ID':
                                        if filename_parts[count] != product.uniqueID:
                                            ref_value = product.uniqueID
                                            passFlag = False

                                    if passFlag is False:
                                        error_list.append(
                                            {'value': "Invalid File Name Base. Expected %s,"
                                                      " but received %s" % (str(ref_value),
                                                                            str(filename_parts[
                                                                                    count])),
                                             'upc': product.upccode,
                                             'slug': product.slug,
                                             'name': product.name,
                                             'asset name': product_asset.file_name,
                                             'type': 'product asset error'})
                                        error_product_list.append(product)
                                        error_product_asset_list.append(product_asset)
                                        error_narrative += " %s Invalid File Name Base. Expected %s, but received %s" % (
                                            str(product_asset.uniqueID), str(ref_value),
                                            str(filename_parts[count]))
                                except Exception as e:
                                    print('AUDIT BASE ERROR: ', e)
                                    error_list.append(
                                        {'value': "Invalid File Name Base.",
                                         'upc': product.upccode,
                                         'slug': product.slug,
                                         'name': product.name,
                                         'asset name': product_asset.file_name,
                                         'type': 'product asset error'})
                                    error_product_list.append(product)
                                    error_product_asset_list.append(product_asset)
                                    error_narrative += " %s Invalid File Name Base." % (
                                        str(product_asset.uniqueID))

                                count += 1
                            if content_standard.filename_suffix != '$NONE':
                                try:
                                    if content_standard.filename_suffix == '$NAME':
                                        if filename_parts[count] != product.name:
                                            ref_value = product.name
                                            passFlag = False
                                    elif content_standard.filename_suffix == '$UPC':
                                        if filename_parts[count] != product.upccode:
                                            ref_value = product.upccode
                                            passFlag = False
                                    elif content_standard.filename_suffix == '$SKU':
                                        if filename_parts[count] != product.SKU:
                                            ref_value = product.SKU
                                            passFlag = False
                                    elif content_standard.filename_suffix == '$XSPACE_ID':
                                        if filename_parts[count] != product.uniqueID:
                                            ref_value = product.uniqueID
                                            passFlag = False

                                    if passFlag is False:
                                        error_list.append(
                                            {'value': "Invalid File Name Suffix. Expected %s,"
                                                      " but received %s" % (str(ref_value),
                                                                            str(filename_parts[
                                                                                    count])),
                                             'upc': product.upccode,
                                             'slug': product.slug,
                                             'name': product.name,
                                             'asset name': product_asset.file_name,
                                             'type': 'product asset error'})
                                        error_product_list.append(product)
                                        error_product_asset_list.append(product_asset)
                                        error_narrative += " %s Invalid File Name Suffix. Expected %s, but received %s" % (
                                            str(product_asset.uniqueID), str(ref_value),
                                            str(filename_parts[count]))
                                except Exception as e:
                                    print('AUDIT SUFFIX ERROR: ', e)
                                    error_list.append(
                                        {'value': "Invalid File Name Suffix.",
                                         'upc': product.upccode,
                                         'slug': product.slug,
                                         'name': product.name,
                                         'asset name': product_asset.file_name,
                                         'type': 'product asset error'})
                                    error_product_list.append(product)
                                    error_product_asset_list.append(product_asset)
                                    error_narrative += " %s Invalid File Name Suffix." % (
                                        str(product_asset.uniqueID))
                                count += 1

        # print('error_list', error_list)
        # print('error_product_list', list(set(error_product_list)))
        # print('error_product_asset_list', list(set(error_product_asset_list)))
        # print('error_narrative', error_narrative)

        # create excel sheet
        # TODO: to save on memory write the excel sheet at the end of each product
        # print(1)
        wb = openpyxl.Workbook()
        # print(2)
        # YYYY-MM-DD HHmmss
        current_datetime = datetime.datetime.now()
        date = str(current_datetime.year) + '-' + str(current_datetime.month) + '-' + str(current_datetime.day) + \
               '_' + str(current_datetime.hour) + '-' + str(current_datetime.minute) + '-' + \
               str(current_datetime.second)
        dest_filename = 'app/orders/templates/audit_%s.xlsx' % date

        product_sheet = wb.active
        product_sheet.title = 'Product Errors'
        # print(3)
        # print(4)
        # format header for the product errors sheet
        product_sheet_headers = ['Error #', 'Product UPC', 'Product Name', 'XSpaceID', 'Error Description']
        for col in range(len(product_sheet_headers)):
            _ = product_sheet.cell(column=col+1, row=1, value="%s" % product_sheet_headers[col])
        # print(5)
        product_asset_sheet = wb.create_sheet(title="Product Asset Errors")

        # print(6)
        # format header for the product asset errors sheet
        product_asset_sheet_headers = ['Error #', 'Product UPC', 'Product Name', 'XSpaceID', 'Product Asset Name',
                                       'Error Description']
        for col in range(len(product_sheet_headers)):
            _ = product_asset_sheet.cell(column=col+1, row=1, value="{0}".format(product_asset_sheet_headers[col]))
        # print(7)
        product_sheet_index_v = 2
        # product_sheet_index_h = 1
        product_asset_sheet_index_v = 2
        # product_asset_sheet_index_h = 1
        # print(8)
        for error in error_list:
            # from pprint import pprint
            # pprint(error)
            if error['type'] == 'product error':
                # print('product error')
                values_list = [str(int(product_sheet_index_v - 1)), error['upc'], error['name'],
                               error['slug'], error['value']]
                for col in range(len(product_sheet_headers)):
                    _ = product_sheet.cell(column=col+1, row=product_sheet_index_v, value="{0}".format(values_list[col]))
                product_sheet_index_v += 1
                # print('complete')
            else:
                # print('product asset error')
                values_list = [str(int(product_asset_sheet_index_v - 1)), error['upc'], error['name'],
                               error['slug'], error['asset name'], error['value']]
                for col in range(len(product_asset_sheet_headers)):
                    _ = product_asset_sheet.cell(column=col+1, row=product_asset_sheet_index_v,
                                                 value="{0}".format(values_list[col]))
                product_asset_sheet_index_v += 1
                # print('complete')
        # print('8a')
        if os.name == 'posix':
            from tempfile import NamedTemporaryFile
            with NamedTemporaryFile() as tmp:
                wb.save(tmp.name)
                tmp.seek(0)
                # stream = tmp.read()
                key = str(company.slug) + "/audit/"
                dest_filename = boto3Upload3(tmp, company.slug, key, is_bytes=True, fileName=str(split(dest_filename)[1]))
            # print('path exists and moving onto saving', dest_filename)
        else:
            try:
                wb.save(filename=dest_filename)
            except Exception as e:
                # print('2', e)
                return Response(status=status.HTTP_200_OK)
        # print(9)
        # send email
        user = User.objects.get(id=profile.user_id)
        # masterAccount = User.objects.get(id=291)  # Jacob Dallas's Xspace Account
        recipientsList = [user]
        # print(10)
        recipient = recipientsList[0]
        ctx = {
            'subject': 'XSPACE | Audit Complete',
            'rep_firstname': recipient.first_name,
            'rep_lastname': recipient.last_name,
        }
        message = get_template('AuditComplete.html').render(ctx)

        mail_subject = '[XSPACE Web App] Your Audit is Complete '
        to_email = str(recipient.email)
        # print(11)
        EmailThread(mail_subject, message, to_email, bcc_list=['j.dallas@xspaceapp.com'], file=dest_filename).start()
        # print(12)
        return Response(status=status.HTTP_200_OK)


class OrdersAPI(AuthenticatedUserMixin, APIView):
    # authentication_classes = (TokenAuthentication,)
    permission_classes = (IsAuthenticated,)
    renderer_classes = (renderers.JSONRenderer,)

    @csrf_exempt
    @has_permission(P.GENERIC)
    def post(self, request, *args, **kwargs):

        bser = BatchOrderSerializer(data=request.data)
        prod_list = []
        checkoutURL = ''
        chargebee_obj = ChargebeeAPI(request)

        if bser.is_valid():
            for order in bser.data['orders']:
                serializer = CreateOrderSerializer(data=order)
                if serializer.is_valid():
                    obj = serializer.create(serializer.validated_data, request)
                    data = serializer.data
                    kintoneData = serializer.data

                    data.update({'products': []})
                    dateString = str(datetime.datetime.now().isoformat())
                    kintoneData.update({
                        'product_orders': [],
                        'numberOfProducts': order['numberOfProducts'],
                        'paid': True,
                        'profile': [],
                        'status': 'Pending',
                        'orderDate': dateString,
                        'chargebeeURL': ''})
                    try:
                        user_profile = {
                            "email": request.user.email,
                            "first_name": request.user.first_name,
                            "last_name": request.user.last_name,
                            "address1": request.data["shippingInfo"]["address1"],
                            "address2": request.data["shippingInfo"]["address2"],
                            "poBoxNum": request.data["shippingInfo"]["poBoxNum"],
                            "city": request.data["shippingInfo"]["city"],
                            "state": request.data["shippingInfo"]["state"],
                            "zipcode": request.data["shippingInfo"]["zipcode"],
                            "country": request.data["shippingInfo"]["country"],
                            "region": request.data["shippingInfo"]["region"]
                        }
                        kintoneData['profile'] = user_profile

                        for product in order['product_list']:
                            product.update({'product': product['slug'], 'order': obj.id})
                            product.update({'singlePhotoServicesRequired': order['hasSingleShot'],
                                            'fourPackServicesRequired': order['hasFourPackShot'],
                                            'fivePackPackServicesRequired': order['hasFivePackShot'],
                                            'standard360ServicesRequired': order['has360View'],
                                            'integrated360ServicesRequired': order['has3DIntegratedModel'],
                                            'standard3DModelServicesRequired': order['hasStandard3DModel'],
                                            'advanced3DModelServicesRequired': order['hasAdvanced3DModel'],
                                            'cinematic3DModelServicesRequired': order['hasCinematic3DModel'],
                                            'hdProductVideoServicesRequired': order['hasStandardHD'],
                                            'fourKProductServicesRequired': order['has4kProductVideo'],
                                            'videoWalkthroughServicesRequired': order['hasVideoWalkthrough'],
                                            'status': 'Pending',
                                            'notes2d': order['text2DNotes'],
                                            'notes360view': order['text360Notes'],
                                            'notes3dmodel': order['text3DNotes']})

                            ordered_product = ProductOrderSerializer(data=product)

                            kintoneData.update({'orderId': str(obj.orderID)})

                            if ordered_product.is_valid():
                                orderProductObj = ordered_product.create(ordered_product.validated_data)

                                product_details = ordered_product.data

                                product_details.update({
                                    'category_name': orderProductObj.product.category.name,
                                    'product_name': orderProductObj.product.name,
                                    'description': orderProductObj.product.description,
                                    'Sku': orderProductObj.product.SKU,
                                    'upc_code': orderProductObj.product.upccode,
                                    'status': 'Pending',
                                    'productId': orderProductObj.product.uniqueID
                                })
                                productWebhook = {
                                    'category_name': orderProductObj.product.category.name,
                                    'product_name': orderProductObj.product.name,
                                    'description': orderProductObj.product.description,
                                    'sku': orderProductObj.product.SKU,
                                    'upc_code': orderProductObj.product.upccode,
                                    'status': 'standby',
                                    'product': orderProductObj.product.uniqueID,
                                    'date_created': str(orderProductObj.product.createdDate.isoformat()),
                                    'date_updated': str(orderProductObj.product.lastUpdatedDate.isoformat()),
                                }

                                kintoneProductData = ordered_product.data
                                kintoneProductData.pop('order', None)
                                kintoneProductData.update({'product_order_id': str(orderProductObj.uniqueID)})
                                kintoneProductData['product'] = productWebhook
                                kintoneData['product_orders'].append(kintoneProductData)
                                data['products'].append(product_details)
                                data['orderId'] = obj.id
                                prod_list.append(product)

                        write_asset_order_excel(data['products'], request)
                        # print(prod_list)

                        headers = {'Content-type': 'application/json'}
                        # res = requests.post(url="https://webhook.site/ad28362d-5415-47c6-962e-82b4c1b64eee", json=kintoneData, headers=headers)

                        # print(kintoneData)

                    except Exception as e:
                        print(e)
                        return Response({"result": e}, status=status.HTTP_400_BAD_REQUEST)
                else:
                    return Response({"result": None, "errors": "serializer error"}, status=status.HTTP_400_BAD_REQUEST)

            # chargebee_obj.create_invoice(request, prod_list)

            chargebee_obj.create_order(request, obj)

            profile = Profile.objects.get(user=request.user)

            # chargebee_obj.getCustomerPortal(request, obj)

            checkoutPage = chargebee_obj.checkout(request, profile, prod_list)

            checkoutURL = checkoutPage.url

            notify.send(request.user, recipient=request.user,
                        verb='Your order has been successfully created.'
                             ' Check your email for QR Print labels to finish shipping.')

            # ipdb.set_trace()

            nameList = []
            counter = 1

            for prod in prod_list:
                keys = {'counter': counter, 'name': prod['name'], 'upc': prod['upccode'], 'sku': prod['SKU']}
                nameList.append(keys)
                counter += 1

            # ipdb.set_trace()

            ctx = {
                'first_name': request.user.first_name,
                'last_name': request.user.last_name,
                'orderID': obj.orderID,
                'createdDate': obj.order_date,
                'numOfProducts': counter - 1,
                'company': str(profile.companyProfile.companyName),
                'productList': nameList,
            }
            message = get_template('CaptureOrderEmail.html').render(ctx)

            mail_subject = '[XSPACE Web App] Thank you for your order! Here are next steps.'
            to_email = str(request.user.email)

            # ipdb.set_trace()

            EmailThread(mail_subject, message, to_email).start()

            return Response({"status": "OK", 'chargebeeURL': checkoutURL}, status=status.HTTP_201_CREATED)
        else:
            return Response(bser.errors, status=status.HTTP_400_BAD_REQUEST)

    @csrf_exempt
    @has_permission(P.READ_ONLY)
    def get(self, request, *args, **kwargs):
        # bugsnag.notify(Exception('working'))
        # orderId = request.query_params['orderId']

        try:
            profile = Profile.objects.get(user=request.user)
            user_order_list = {}
            # Find Product for selected Order
            if 'orderId' in request.data or 'orderId' in request.query_params:
                orderId = request.query_params['orderId']
                order = Order.objects.get(orderID=orderId)
                product = ProductOrder.objects.filter(order=order)
                serializer = ProductOrderListSerializer(product, many=True)
                user_order_list['product_order'] = serializer.data
            else:
                # Find Order for current user
                # queryset = Order.objects.filter(company=profile.companyProfile).order_by('-order_date')
                queryset = FullOrder.objects.filter(profile=profile).order_by('-createdDate')
                paginator = Paginator(queryset, 100)
                page = request.GET.get('page')
                try:
                    queryset = paginator.page(page)
                except PageNotAnInteger:
                    queryset = paginator.page(1)
                except EmptyPage:
                    queryset = paginator.page(paginator.num_pages)
                # serializer = OrderSerializer(queryset, many=True)
                try:
                    serializer = FullOrderSerializer(queryset, many=True)
                    paginatorcount = paginator.count
                    orderids = set()
                    databack = dict()
                    servicelist = {'singlePhoto', 'fourPack', 'fivePack', 'standard360', 'integrated360',
                                   'standard3DModel', 'advanced3DModel', 'cinematic3DModel', 'hdProductVideo',
                                   'fourKProduct', 'videoWalkthrough', 'backgroundRemoval', 'autoCropAndCenter',
                                   'blemishReduction', 'barcodeDetection', 'applyContentStandard', 'colorCorrection'
                                   }
                    servicenotes = {'notes2d', 'notes360view', 'notes3dmodel', 'notesvideo'}
                    statuslist = {0: 'Pending Payment', 1: 'Paid', 2: 'In Progress', 3: 'Pending Approval',
                                  4: 'Complete', -1: 'cancelled'}

                    for order in serializer.data:
                        orderid = order['order_ID']
                        statuss = {order['status']}
                        createdate = order['createdDate']
                        lastmodified = order['lastModifiedDate']
                        description = {x for x in servicelist if order[x]}
                        servicenote = {x: {order[x], } for x in servicenotes}

                        if orderid in orderids:
                            databack[orderid]['status'] = databack[orderid]['status'].union(statuss)
                            if createdate < databack[orderid]['createdDate']: databack[orderid][
                                'createdDate'] = createdate
                            if lastmodified > databack[orderid]['lastModifiedDate']: databack[orderid][
                                'lastModifiedDate'] = lastmodified
                            databack[orderid]['description'] = databack[orderid]['description'].union(description)
                            databack[orderid]['numberOfProducts'] += 1
                            for x in servicenote:
                                databack[orderid]['serviceNote'][x] = databack[orderid]['serviceNote'][x].union(
                                    servicenote[x])
                        else:
                            databack[orderid] = {
                                'order_ID': orderid, 'status': statuss, 'createdDate': createdate,
                                'lastModifiedDate': lastmodified, 'description': description, 'numberOfProducts': 1,
                                'serviceNote': servicenote
                            }
                            orderids.add(orderid)

                    for k, v in databack.items():
                        statuss = sorted(list(v['status']))
                        statusconvert = statuslist[statuss[-1]]
                        statusconvert = 'Partial Order Status - ' + statusconvert if len(statuss) > 1 else statusconvert
                        v['status'] = statusconvert
                        noservice = servicelist - v['description']
                        yesservice = [x + ' : Yes\n' for x in v['description']]
                        noservice = yesservice + [x + ' : No\n' for x in noservice]
                        noservice[-1] = noservice[-1].rstrip('\n')
                        v['description'] = yesservice
                        v['descriptionList'] = ''.join(noservice)
                        notes = v['serviceNote']
                        for x in notes:
                            notes[x] = x + ':\n' + '\n'.join(list(notes[x]))
                        v['serviceNote'] = '\n'.join(notes.values())

                    user_order_list['product_order'] = databack.values()
                    # user_order_list['totalCount'] = paginator.count
                    user_order_list['totalCount'] = paginatorcount

                except Exception as e:
                    traceback.print_exc()
                    print("OrdersAPI:  get,  serializer error", e, type(e))
            try:
                return Response({"product_order": user_order_list}, status=status.HTTP_200_OK)
            except Exception as e:
                traceback.print_exc()
                print("OrdersAPI:  get,  OK Response", e, type(e))
        except Exception as e:
            return Response({"result": "invalid request", "error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @has_permission(P.GENERIC)
    def put(self, request, *args, **kwargs):
        if (request.META["HTTP_X_TOKEN_ID"] != "6f86fe9b538e08aa0f7c8df44ce8f403"):
            return Response({"error": "INVALID API Key"}, status=status.HTTP_400_BAD_REQUEST)
        orderId = request.query_params['orderId']
        serializer = OrderUpdateSerializer(data=request.data)
        try:
            if serializer.is_valid():
                Order.objects.filter(orderID=orderId).update(**serializer.validated_data)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except:
            return Response({"error": serializer.error}, status=status.HTTP_400_BAD_REQUEST)


class EditingOrdersAPI(AuthenticatedUserMixin, APIView):
    # authentication_classes = (TokenAuthentication,)
    permission_classes = (IsAuthenticated,)
    renderer_classes = (renderers.JSONRenderer,)

    @csrf_exempt
    @has_permission(P.GENERIC)
    def post(self, request, *args, **kwargs):
        # Get Payload
        data = request.data

        profile = Profile.objects.get(user=request.user)

        # Get Variables from Payload.
        hasBackgroundRemoval = data['hasBackgroundRemoval']
        hasAutoCropAndCenter = data['hasAutoCropAndCenter']
        hasRename = data['hasRename']
        hasBlemishReduction = data['hasBlemishReduction']
        hasBarcodeDetection = data['hasBarcodeDetection']
        fileCount = data['fileCount']

        # ipdb.set_trace()

        imageList = data['fileList']

        data['profile'] = profile.id
        data['company'] = profile.companyProfile.id

        # ipdb.set_trace()
        serializer = EditOrderSerializer(data=request.data)

        # ipdb.set_trace()
        #
        checkoutURL = ''
        chargebee_obj = ChargebeeAPI(request)

        if serializer.is_valid():
            # ipdb.set_trace()
            obj = serializer.create(serializer.validated_data, request)

        # chargebee_obj.create_order(request, obj)
        # chargebee_obj.create_order(obj)
        estimate = chargebee_obj.create_editing_estimate(profile, fileCount, hasBackgroundRemoval,
                                                         hasAutoCropAndCenter, hasRename, hasBlemishReduction,
                                                         hasBarcodeDetection)

        # print("QUOOOOOOTES", estimate)
        checkoutPage = chargebee_obj.editingCheckout(profile, fileCount, hasBackgroundRemoval,
                                                     hasAutoCropAndCenter, hasRename, hasBlemishReduction,
                                                     hasBarcodeDetection)

        print("QUOOOOOOTES")

        # #checkoutPage = chargebee_obj.checkout(request, profile, prod_list)
        #
        checkoutURL = checkoutPage.url

        nameList = []
        counter = 1

        # ipdb.set_trace()

        ctx = {
            'first_name': request.user.first_name,
            'last_name': request.user.last_name,
            'orderID': obj.orderID,
            'createdDate': obj.order_date,
            'numOfFiles': fileCount,
            'company': str(profile.companyProfile.companyName),
            'fileList': imageList,
        }
        message = get_template('EditOrderEmail.html').render(ctx)

        mail_subject = '[XSPACE Web App] Thank you for your order! Here are next steps.'
        to_email = str(request.user.email)

        # ipdb.set_trace()

        EmailThread(mail_subject, message, to_email).start()

        notify.send(request.user, recipient=request.user,
                    verb='Your editing order has been successfully created. Check My Orders page for downloads when your order is complete.')

        print("BEFOREJSON")
        print(estimate)
        print("AFTERJSON")
        estimate = json.loads(str(estimate.estimate))
        # estimate = estimate.estimate.invoice_estimate
        # estimate = estimate
        # try:
        # estimate['chargebeeURL'] = checkoutURL
        # print(estimate)
        #     estimate = json.dumps(estimate);
        # except Exception as e:
        #     print("AFTERJSON", e, type(e), type(estimate), str(estimate.estimate))
        #     for x in estimate:
        #         print(type(x))
        #         print(x)

        return Response({"status": "OK", 'chargebeeURL': checkoutURL, 'cart_estimate': estimate},
                        status=status.HTTP_201_CREATED)


class FullOrdersAPI(AuthenticatedUserMixin, APIView):
    # authentication_classes = (TokenAuthentication,)
    permission_classes = (IsAuthenticated,)
    renderer_classes = (renderers.JSONRenderer,)

    # request:
    # -data
    # --'orderid'
    # --'fileList'
    # --'orderparts'
    # ----[[assetid1, attribute1, value],
    #      [assetid1, attribute2, value],
    #      [assetid2, attribute1, value],
    #      ...
    # - user

    # TODO: add in subscriptions

    @csrf_exempt
    @has_permission(P.GENERIC)
    def post(self, request, *args, **kwargs):
        """
        create a new order
        :param request:
        :param args:
        :param kwargs:
        :return:
        """
        print('FullOrdersAPI POST received', request.data)

        # Get Payload
        data = request.data

        # Get Profile
        profile = Profile.objects.get(user=request.user)

        # chargebee_obj = ChargebeeAPI(request)

        # Get Variables from Payload.
        # fileCount = len(data)

        fileList = data['fileList']
        fileCount = len(fileList)

        imageList = [x[0] for x in fileList]

        try:
            # create serializer
            serializer = FullOrderSerializer(data=request.data)

            # create order object
            counters, errors, serializer_response, obj, productList = serializer.create(request.data, request)

            if serializer_response == 400:
                return Response({"status": None, 'result': None}, status=status.HTTP_400_BAD_REQUEST)

        except Exception as e:
            print(e, type(e))

        # [Instead of create order (old pipeline to indicate waiting for shipment), create shopping cart estimated
        # invoice as first step.
        # chargebee_obj.create_order(request.data['orderid'])

        ################################## Moved to Get.
        # try:
        #     estimate = chargebee_obj.create_fullorder_estimate( profile, counters)
        #
        #     estimate = json.loads(str(estimate.estimate))
        #     # print(type(estimate))
        #
        #     checkoutPage = chargebee_obj.fullCheckout(profile, counters)
        #     checkoutURL = checkoutPage.url
        #     order_obj = FullOrder.objects.filter(order_ID=obj.order_ID)
        #     # for x in order_obj:
        #     #     x.chargebeeURL = checkoutURL
        #     #     x.save()
        #
        # except Exception as e:
        #     print(e, type(e))

        if productList:
            try:
                ctx = {
                    'first_name': request.user.first_name,
                    'last_name': request.user.last_name,
                    'orderID': obj.order_ID,
                    'createdDate': obj.createdDate,
                    # 'numOfFiles': fileCount,
                    'numOfFiles': len(productList),
                    'company': str(profile.companyProfile.companyName),
                    # 'fileList': imageList,
                    'productList': productList,

                }

                # message = get_template('EditOrderEmail.html').render(ctx)

                message = get_template('EditOrderEmail.html').render(ctx)

                mail_subject = '[XSPACE Web App] Thank you for your order! Here are next steps.'
                to_email = str(request.user.email)
                # bcc_email = "sales@snap36.com"
                bcc_email = "angel.alladina@gmail.com"

                EmailThread(mail_subject, message, to_email, bcc_email).start()

                notify.send(request.user, recipient=request.user,
                            verb='Your PhotoCapture order has been successfully created. Check "My Orders" page for downloads when your order is complete.')

            except Exception as e:
                print(e, type(e))

        # checkoutURL = { 'checkoutURL': checkoutURL, 'fileList': fileList, 'orderID': request.data['orderid'], 'estimate': estimate }

        checkoutURL = {}

        return Response({"status": "OK", 'result': checkoutURL}, status=status.HTTP_201_CREATED)

    def get(self, request, *args, **kwargs):
        # TODO: update chargebee url

        profile = Profile.objects.get(user=request.user)
        # get order_ID
        if 'orderid' in request.data:
            # get the order
            order_obj = FullOrder.objects.filter(order_ID=request.data["orderid"], profile=profile)

            # No such order exists
            if order_obj.count() == 0:
                return Response({"status": None, 'result': None}, status=status.HTTP_404_NOT_FOUND)

            return Response({"status": "OK", 'result': order_obj}, status=status.HTTP_200_OK)

        else:
            # get all active orders:
            order_objs = FullOrder.objects.filter(isActive=True, profile=profile, status=0)
            order_objs_editing, order_objs_capture = order_objs.exclude(isCapture=True), order_objs.exclude(
                isEditing=True)

            if order_objs.count() > 0:
                # active orders exist, structure response
                chargebee_obj = ChargebeeAPI(request)
                serializer = FullOrderSerializer(data=request.data)
                checkoutURLS = {}

                # iterate through active orders
                print('[order_objs_editing, order_objs_capture]', [order_objs_editing, order_objs_capture])
                for idx, order_objs in enumerate([order_objs_editing, order_objs_capture]):
                    print('\nidx, order_objs', idx, order_objs)
                    if order_objs.count() > 0:
                        print("GREATERTHANZERO")

                        allorders = serializer.getAttributes(order_objs)
                        print('order_objs', order_objs)

                        allordersIDs = allorders.keys()
                        print('allorders, allordersIDs', allorders, allordersIDs)

                        try:
                            estimates, combined_counter = chargebee_obj.create_fullorder_estimate(profile, allorders,
                                                                                                  many=True)

                            for orders in estimates:
                                estimates[orders] = json.loads(str(estimates[orders].estimate))

                            if idx == 0:
                                checkoutPage = chargebee_obj.fullCheckout(profile, combined_counter,
                                                                          orderIDs=allordersIDs, many=True)
                                checkoutURL = checkoutPage.url
                                print('checkoutURL', checkoutURL)

                            print('estimates1', estimates)
                            print('allorders',  allorders)
                            for orderid in estimates:
                                print('estimates2', orderid)
                                if orderid in allorders:
                                    estimates[orderid]['fileList'] = allorders[orderid]['fileList']
                                else:
                                    print('orderid not in allorders')

                            if idx == 0:
                                checkoutURLS['checkoutURL'] = checkoutURL
                                checkoutURLS['editing'] = estimates
                            else:
                                checkoutURLS['capture'] = estimates

                        except Exception as e:
                            print('ERROR:', e, type(e))
                print('sending forward', checkoutURLS)
                return Response({"status": "OK", 'result': checkoutURLS}, status=status.HTTP_200_OK)

            else:

                return Response({"status": "Empty", 'result': {}}, status=status.HTTP_200_OK)

    def put(self, request, *args, **kwargs):
        """
        update an order
        :param request:
        :param args:
        :param kwargs:
        :return:
        """
        # Get Payload
        data = request.data
        print('PUT data', data)
        # Get Profile
        profile = Profile.objects.get(user=request.user)
        if 'isActive' in data:
            try:
                print(request.data['orderIDs'])
                for orderID in request.data['orderIDs']:
                    print(orderID)
                    order_obj = FullOrder.objects.filter(order_ID=orderID, profile=profile.id, status=0)
                    for obj in order_obj:
                        obj.status = -1
                        obj.isActive = False
                        obj.save()
            except Exception as e:
                print("FullOrderAPI PUT error in IsActive", e, type(e))

            return self.get(request)
        elif 'Complete' in data:
            try:
                for orderID in request.data['orderIDs']:
                    order_obj = FullOrder.objects.filter(order_ID=orderID, profile=profile.id, isCapture=True)
                    for obj in order_obj:
                        obj.status = 4
                        obj.isActive = False
                        obj.save()
            except Exception as e:
                print("FullOrderAPI PUT error in complete", e, type(e))
                return Response(status.HTTP_400_BAD_REQUEST)
            return Response(status.HTTP_200_OK)
        elif 'paid' in data:
            try:
                print(data['paid'])

                chargebee_obj = ChargebeeAPI(request)
                # print(1)
                orderIdList = chargebee_obj.getPaidOrderIDs(data['paid'])
                # print(2)
                editingWorkList = []
                # print(3)
                for order_ID in orderIdList:
                    order_obj = FullOrder.objects.filter(order_ID=order_ID, profile=profile.id, status=0, isActive=True)
                    # print(4, order_obj)
                    for obj in order_obj:
                        obj.isActive = False
                        obj.status = 1
                        obj.paid = True
                        obj.save()
                        # print(5)
                        # if editing is required, we need to process it on a separate thread so let's add it to a
                        # list to be iterated over on a separate thread so that we don't block the impatient frontend
                        # print('obj.backgroundRemoval', obj.backgroundRemoval)
                        if obj.backgroundRemoval:
                            try:
                                productAsset = ProductAsset.objects.get(id=obj.productAsset_id)
                            except Exception as e:
                                print('error getting product asset id', e)
                                continue
                            cs = None
                            try:
                                if obj.contentStandard == '' or obj.contentStandard == 'null' or obj.contentStandard == None:
                                    cs = None
                                else:
                                    cs = obj.contentStandard
                            except Exception as e:
                                print('error getting contentStandard', e)
                            # print(1)
                            # print(productAsset.id)
                            # print(2)
                            # print(productAsset.assetType)
                            editingWorkList.append({'productAssetID': str(productAsset.id),
                                                    'type': 'backgroundRemoval',
                                                    'contentStandard': cs,
                                                    'assetType': productAsset.assetType
                                                    })

                if len(editingWorkList) > 0:
                    # we need to send this off to another thread which is not safely supported as of writing so we'll
                    # send it forward to get sent back again to let django handle the threading

                    return Response({"status": "PAID", 'editingWorkList': editingWorkList}, status=status.HTTP_200_OK)
                else:
                    return Response({"status": "PAID", 'editingWorkList': 'nothing'}, status=status.HTTP_200_OK)
            except Exception as e:
                print(e, type(e))
                return Response({"status": "BAD", 'editingWorkList': 'nothing'}, status=status.HTTP_404_NOT_FOUND)
        else:
            order_ID = request.data['orderid']

            # update the order
            for key in data['orderparts']:
                # get the product asset, attribute, and value to be set
                asset_id, attribute, value = data[key]
                order_obj = FullOrder.objects.get(order_ID=order_ID, productAsset=asset_id, profile=profile.id,
                                                  status=0)
                if attribute == 'singlePhoto':
                    order_obj.singlePhoto = value
                elif attribute == 'fourPack':
                    order_obj.fourPack = value
                elif attribute == 'fivePack':
                    order_obj.fivePack = value
                elif attribute == 'standard360':
                    order_obj.standard360 = value
                elif attribute == 'integrated360':
                    order_obj.integrated360 = value
                elif attribute == 'standard3DModel':
                    order_obj.standard3DModel = value
                elif attribute == 'advanced3DModel':
                    order_obj.advanced3DModel = value
                elif attribute == 'cinematic3DModel':
                    order_obj.cinematic3DModel = value
                elif attribute == 'hdProductVideo':
                    order_obj.hdProductVideo = value
                elif attribute == 'fourKProduct':
                    order_obj.fourKProduct = value
                elif attribute == 'videoWalkthrough':
                    order_obj.videoWalkthrough = value
                elif attribute == 'notes2d':
                    order_obj.notes2d = value
                elif attribute == 'notes360view':
                    order_obj.notes360view = value
                elif attribute == 'notes3dmodel':
                    order_obj.notes3dmodel = value
                elif attribute == 'notesvideo':
                    order_obj.notesvideo = value
                elif attribute == 'backgroundRemoval':
                    order_obj.backgroundRemoval = value
                elif attribute == 'autoCropAndCenter':
                    order_obj.autoCropAndCenter = value
                elif attribute == 'blemishReduction':
                    order_obj.blemishReduction = value
                elif attribute == 'barcodeDetection':
                    order_obj.barcodeDetection = value
                elif attribute == 'applyContentStandard':
                    order_obj.applyContentStandard = value
                else:
                    pass
                order_obj.save()
            return Response(status.HTTP_200_OK)


class OrderTrackingViewSet(AuthenticatedUserMixin, APIView):
    renderer_classes = (renderers.JSONRenderer,)

    @csrf_exempt
    @has_permission(P.GENERIC)
    def post(self, request, *args, **kwargs):
        try:

            orderTracking = OrderTrackingSerializer(data=request.data['orderTrack'])

            if orderTracking.is_valid():
                orderTrackObj = orderTracking.create(orderTracking.validated_data)

            '''
            Added Email to Operations on Updated Tracking Number
            '''

            tracknum = str(orderTracking.validated_data['tracking_number'])
            user = User.objects.get(email=request.user.email)
            orderObj = Order.objects.get(orderID=orderTrackObj.order.orderID)
            profile = Profile.objects.get(user=user)
            company = profile.companyProfile

            ctx = {
                'subject': 'XSPACE | New Message',
                'first_name': user.first_name,
                'last_name': user.last_name,
                'rep_firstname': 'XSPACE',
                'rep_lastname': 'Operations',
                'orderID': orderObj.orderID,
                'company': str(company.companyName),
                'trackingnumber': str(tracknum),
            }
            message = get_template('AddedTrackingNumberEmail.html').render(ctx)

            mail_subject = '[XSPACE Web App] ' + str(user.first_name) + ' ' + str(user.last_name) + ' from ' + str(
                company.companyName) + ' has added a new tracking number'
            to_email = str('operations@xspaceapp.com')

            EmailThread(mail_subject, message, to_email).start()

            '''
            Add Kintone API Push Data
            '''
            trackingNumberData = {'carrierName': '', 'trackingNumber': tracknum}
            trackData = {'orderId': orderObj.orderID, "Recieving": trackingNumberData}
            headers = {'Content-type': 'application/json'}

            # TODO: [BACKEND][Dom] Test Kintone Web Hook for Updating Order Number.
            # res = requests.put(url="https://webhook.site/ad28362d-5415-47c6-962e-82b4c1b64eee", json=trackData,
            # headers=headers)

            return Response({"message": 'You have successfully added your tracking number.'}, status=status.HTTP_200_OK)
        except Exception as e:
            print(e)
            return Response({"message": "Your tracking number could not be added to this order."},
                            status=status.HTTP_400_BAD_REQUEST)

    @has_permission(P.READ_ONLY)
    def get(self, request, *args, **kwargs):
        print('This will get order tracking details')

    @has_permission(P.GENERIC)
    def put(self, request, *args, **kwargs):
        print('Update the Order tracking Number')


class OrderQRCodeViewSet(AuthenticatedUserMixin, APIView):
    renderer_classes = (renderers.JSONRenderer,)

    @csrf_exempt
    @has_permission(P.GENERIC)
    # This function will generate the PDF
    def post(self, request, *args, **kwargs):
        msg = '0'
        try:
            msg = '1'
            data = request.data['orderQr']
            msg = '2'
            barCode = GenerateQRLabelDoc()
            msg = '3'
            order = Order.objects.get(orderID=data['orderId'])
            msg = '4'
            productOrders = ProductOrder.objects.filter(order=order)
            msg = '5'
            user = User.objects.get(email=request.user.email)
            fileName = barCode.createDoc(productOrders, request.data['orderQr'], user)
            msg = '6'

            result = uploadTos3(fileName, user.username, '/orders/')
            if result['status'] == 'success':
                if os.path.exists(fileName):
                    os.remove(fileName)

            """ wrapper = FileWrapper(file(fileName,'rb'))
            response = HttpResponse(wrapper, content_type='application/pdf')
            response['Content-Length'] = os.path.getsize(fileName)
            response['Content-Disposition'] = "attachment; filename=" +fileName """
            return Response(
                {"message": "https://dbrdts7bui7s7.cloudfront.net/" + result['location']})  # FileResponse(response)
        except Exception as e:
            return Response(json.dumps({"message": e}), status=status.HTTP_400_BAD_REQUEST)

    # This function will Download the PDF
    @has_permission(P.READ_ONLY)
    def get(self, request, *args, **kwargs):
        print('This will get order tracking details')

    @has_permission(P.GENERIC)
    def put(self, request, *args, **kwargs):
        print('Update the Order tracking Number')


class OrderExcelToQRCodeViewSet(APIView):
    renderer_classes = (renderers.JSONRenderer,)
    permission_classes = (IsNotAuthenticated,)

    def post(self, request, *args, **kwargs):

        try:
            barCode = GenerateQRLabelDoc()
            excelFile = request.data['file']
            fileName = barCode.createDocFromExcel(excelFile)

            # result = uploadTos3(fileName, user.username, '/orders/')

            """ wrapper = FileWrapper(file(fileName,'rb'))
            response = HttpResponse(wrapper, content_type='application/pdf')
            response['Content-Length'] = os.path.getsize(fileName)
            response['Content-Disposition'] = "attachment; filename=" +fileName """
            return Response(
                {"message": fileName})  # FileResponse(response)
        except Exception as e:
            return Response(json.dumps({"message": e}), status=status.HTTP_400_BAD_REQUEST)

    # This function will Download the PDF
    @has_permission(P.READ_ONLY)
    def get(self, request, *args, **kwargs):
        print('This will get order tracking details')

    @has_permission(P.GENERIC)
    def put(self, request, *args, **kwargs):
        print('Update the Order tracking Number')


# class OrderRequestWebhookAPI(APIView):
#     @csrf_exempt
#     def get(self, request, *args, **kwargs):
#         try:
#             data = {'api_dev_key':"123",
#                     'api_option':'OPTION',
#                     'api_paste_code':'abe',
#                     'api_paste_format':'python'}
#             headers = {'Content-type': 'application/json'}
#             res = requests.post(url="https://webhook.site/5f9b5d55-34cc-4ef3-bdf6-250e54c0c38c", json=data, headers=headers)
#             return Response({"status":"OKK"},status=status.HTTP_200_OK)
#         except:
#             return Response({"status":"error"},status=status.HTTP_400_BAD_REQUEST)


class OrdersRESTAPI(APIView):
    renderer_classes = (renderers.JSONRenderer,)

    @csrf_exempt
    @has_permission(P.UPDATE_ORDER)
    def put(self, request, *args, **kwargs):
        # CLEANUP: transition PSK to env variable to get rid of constants
        if (request.META["HTTP_X_TOKEN_ID"] != "6f86fe9b538e08aa0f7c8df44ce8f403"):
            return Response({"status": "error", "message": "INVALID API Key"}, status=500)
        # check for an order id in params, otherwise evalute product order id
        try:
            order_id = request.query_params['order_id']
            orderSerializer = OrderStatusUpdateSerializer(data=request.data)
            try:
                if orderSerializer.is_valid():
                    Order.objects.filter(orderID=order_id).update(**orderSerializer.validated_data)
                    return Response({"status": "success", "message": "Order status has been updated."},
                                    status=status.HTTP_200_OK)
                else:
                    return Response(
                        {"status": "error", "message": "Order requires a 'status' field in the request body."},
                        status=500)
            except:
                return Response({"status": "error", "message": orderSerializer.error}, status=500)
        except MultiValueDictKeyError:
            try:
                product_order_id = request.query_params['product_order_id']
                poSerializer = ProductOrderUpdateSerializer(data=request.data)
                try:
                    if poSerializer.is_valid():
                        ProductOrder.objects.filter(uniqueID=product_order_id).update(**poSerializer.validated_data)
                        return Response(
                            {"status": "success", "message": "Product order status has been updated successfully."},
                            status=status.HTTP_200_OK)
                    else:
                        return Response({"status": "error",
                                         "message": "Product order requires a 'status' field in the request body."},
                                        status=500)
                except:
                    return Response({"status": "error", "message": poSerializer.error}, status=500)
            except MultiValueDictKeyError:
                return Response({"status": "error", "message": "missing url parameter "}, status=500)


class OrdersByUUIDAPI(APIView):
    renderer_classes = (renderers.JSONRenderer,)

    @csrf_exempt
    @has_permission(P.READ_ONLY)
    def get(self, request, uidb64):
        # bugsnag.notify(Exception('working'))
        # orderId = request.query_params['orderId']
        try:
            user_order_list = {}
            # Find Product for selected Order
            order = Order.objects.get(orderID=uidb64)
            productOrders = ProductOrder.objects.filter(order=order)
            serializer = FullOrderSerializer(order, many=False, context={'productOrders': productOrders})
            #
            # serializer.data['product_orders'] = prodOrderSerializer.data
            # prodOrderSerializer.data

            return Response({"orders": serializer.data}, status=status.HTTP_200_OK)
        except Exception as e:
            print(e)
            return Response({"result": str(e)}, status=status.HTTP_400_BAD_REQUEST)


class AllOrdersAPI(APIView):
    renderer_classes = (renderers.JSONRenderer,)

    @csrf_exempt
    @has_permission(P.READ_ONLY)
    def get(self, request, *args, **kwargs):
        # bugsnag.notify(Exception('working'))
        pageno = request.query_params['pageno']
        try:

            snippets = Order.objects.all()
            paginator = Paginator(snippets, 10)

            totalrow = paginator.count
            # pagecount = [len(totalrow)]

            page = pageno
            # ipdb.set_trace()
            try:
                snippets = paginator.page(page)
            except PageNotAnInteger:
                # If page is not an integer, deliver first page.

                snippets = paginator.page(1)
            except EmptyPage:
                # If page is out of range (e.g. 9999), deliver last page of results.
                snippets = paginator.page(paginator.num_pages)

            # serializer = ProductSerializer(snippets, many=True)

            user_order_list = {}
            # Find Product for selected Order
            # order = Order.objects.all()

            orders = paginator.get_page(page).object_list
            serializer = FullOrderSerializer(orders, many=True)

            # ipdb.set_trace()
            #
            # serializer.data['product_orders'] = prodOrderSerializer.data
            # prodOrderSerializer.data

            return Response({"orders": serializer.data}, status=status.HTTP_200_OK)
        except Exception as e:
            print(e)
            return Response({"result": serializer.data}, status=status.HTTP_400_BAD_REQUEST)


class OrdersByCompanyAPI(APIView):
    renderer_classes = (renderers.JSONRenderer,)

    @csrf_exempt
    @has_permission(P.READ_ONLY)
    def get(self, request, uidb64):
        # bugsnag.notify(Exception('working'))
        # orderId = request.query_params['orderId']
        try:
            user_order_list = {}
            # Find Product for selected Order
            company = Company.objects.get(uniqueID=uidb64)
            order = Order.objects.filter(company=company)
            productOrders = ProductOrder.objects.filter(order=order)
            serializer = FullOrderSerializer(order, many=True, context={'productOrders': productOrders})
            #
            # serializer.data['product_orders'] = prodOrderSerializer.data
            # prodOrderSerializer.data

            return Response({"orders": serializer.data}, status=status.HTTP_200_OK)
        except Exception as e:
            print(e)
            return Response({"result": str(e)}, status=status.HTTP_400_BAD_REQUEST)


class OrdersByUsernameAPI(APIView):
    renderer_classes = (renderers.JSONRenderer,)

    @csrf_exempt
    # @has_permission(P.READ_ONLY)
    def get(self, request, username, *args, **kwargs):
        # bugsnag.notify(Exception('working'))
        # orderId = request.query_params['orderId']
        try:
            user_order_list = {}
            # Find Product for selected Order
            user = User.objects.get(username=username)
            profile = Profile.objects.get(user=user)
            order = Order.objects.filter(profile=profile)

            productOrders = ProductOrder.objects.filter(order=order)
            serializer = FullOrderSerializer(order, many=True, context={'productOrders': productOrders})
            #
            # serializer.data['product_orders'] = prodOrderSerializer.data
            # prodOrderSerializer.data

            return Response({"orders": serializer.data}, status=status.HTTP_200_OK)
        except Exception as e:
            print(e)
            return Response({"result": str(e)}, status=status.HTTP_400_BAD_REQUEST)


class UploadByUsernameAPI(APIView):
    renderer_classes = (renderers.JSONRenderer,)

    @csrf_exempt
    # @has_permission(P.READ_ONLY)
    def get(self, request, username, *args, **kwargs):
        # bugsnag.notify(Exception('working'))
        # orderId = request.query_params['orderId']
        try:
            user_order_list = {}
            # Find Product for selected Order
            user = User.objects.get(username=username)
            profile = Profile.objects.get(user=user)
            products = Product.objects.filter(profile=profile)
            serializer = UploaderSerializer(products, many=False, context={'user': user})
            #
            # serializer.data['product_orders'] = prodOrderSerializer.data
            # prodOrderSerializer.data

            return Response({"uploadData": serializer.data}, status=status.HTTP_200_OK)
        except Exception as e:
            print(e)
            return Response({"result": str(e)}, status=status.HTTP_400_BAD_REQUEST)


class GetAllOrderDownloads(AuthenticatedUserMixin, APIView):
    permission_classes = (IsAuthenticated,)
    # authentication_classes = (TokenAuthentication)
    renderer_classes = (renderers.JSONRenderer,)

    @has_permission(P.READ_ONLY)
    def get(self, request, orderID):

        try:
            profile = Profile.objects.get(user=request.user)
            try:
                company = profile.companyProfile
                order = Order.objects.get(orderID=orderID)

                files = []
                file_size = []
                created_date = []
                p_id = []
                # key = []

                key = str(company.slug) + "/orders/downloads/" + str(orderID)

                bucketname = 'storagev2'
                s3 = boto3.resource('s3', aws_access_key_id='AKIAJKV6HI4YIS35FTVA',
                                    aws_secret_access_key='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')
                data = s3.Bucket(bucketname)

                list_product = []

                # ipdb.set_trace()

                for item in data.objects.filter(Prefix=key):
                    test = ''
                    for k in range(len(item.key)):
                        if k == (len(item.key) - 1):
                            test += item.key[k]
                    if test != '/':
                        # print str(item.size/1024)+' kb'
                        file_size.append(item.size / 1024)
                        files.append('https://storage.xspaceapp.com/' + item.key)
                    created_date.append(item.last_modified)

                fileObjects = []
                # ipdb.set_trace()
                for file in files:
                    stringArr = file.split('/')
                    print(stringArr)
                    payload = {"link": file, "file_name": stringArr[7]}
                    fileObjects.append(payload)
                return Response({'success': 'true', 'error': 'None', 'files': fileObjects})

            except Order.DoesNotExist:
                return Response({'success': 'None', 'error': 'No order found for this ID.'})

        except Profile.DoesNotExist:
            return Response({'success': 'None', 'error': 'User not found.'})

        lis = []
        return Response({'contentStandard': lis}, status=status.HTTP_200_OK)


class ProcessEditingOrderAPI(AuthenticatedUserMixin, APIView):
    """
    class for processing editing orders
    """

    @csrf_exempt
    def post(self, request, *args, **kwargs):
        """
        process an order
        :param request:
        :param args:
        :param kwargs:
        :return:
        """
        pk = request.user.id
        user = User.objects.get(pk=pk)
        profile = Profile.objects.get(user_id=user.id)
        company = Company.objects.get(id=profile.companyProfile.id)
        workload = request.data
        print('workload', workload)
        errorList = []
        for work in workload:
            worker = processEditingOrder(work)
            value = worker.run(company_object=company)
            if value == 'success':
                continue
            else:
                errorList.append(work)
        del value
        if len(errorList) > 0:
            if len(errorList) == len(workload):
                return Response({'errors': errorList}, status=status.HTTP_400_BAD_REQUEST)
            else:
                return Response({'errors': errorList}, status=status.HTTP_206_PARTIAL_CONTENT)
        else:
            return Response({'errors': 'nothing'}, status=status.HTTP_200_OK)
