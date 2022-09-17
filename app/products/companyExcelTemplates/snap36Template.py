import codecs
import decimal
import re
import csv
from babel.numbers import format_currency
from app.core.models import Category
from openpyxl import load_workbook


def snap36Template(excelFile):
    error_list = []
    res = []
    un_upload_product = []
    fieldnames = ("Base required information", "Misc", "Vendor management", "Product Description", "Internal SKU#",
                  "UPC", "Product Category", "Item Description", "Does the product stand on its own?",
                  "Photography notes", "Item Length", "Item Width", "Item Height", "Item Weight")
    row1fieldnamesRefList = ["Base required information", "Misc", "Vendor management"]
    row2fieldnamesRefList = ["Internal SKU#", "UPC", "Product Category", "Item Description",
                             "Does the product stand on its own?", "Photography notes", "Item Length", "Item Width",
                             "Item Height", "Item Weight"]
    newFieldNames = ("Internal SKU#", "UPC", "UPC Type", "Product Category", "Item Description",
                     "Does the product stand on its own?", "Photography notes", "Item Length", "Item Width",
                     "Item Height", "Item Weight", 'ACDelco', 'Advanced Auto', 'Advantage', 'Amazon', 'Grainger SKU',
                     'Home Depot', 'Johnstone', "O'Reilly", 'Walmart', 'VendorCode1', 'VendorCode2', "VendorCode3")
    # TODO: add in csv/text parser
    if excelFile.content_type == 'text/csv':
        reader = csv.reader(codecs.iterdecode(excelFile, 'utf-8'), fieldnames)

        # determine if we are using the default or not
        count = 0

        for fieldname in reader[0]:
            if fieldname == row1fieldnamesRefList[count]:
                # matches
                if row1fieldnamesRefList[count] == "Item Weight":
                    break
                count += 1
                continue
            else:
                return False, False, False

        print('using Snap36 template excel template')
        row_num = 0
        for line in reader:
            row_num += 1
            if row_num > 1:
                if (not re.match(r'\d+(?:[.]\d{2})?$', line[6])):
                    error_list.append(
                        "Warning. Changed format of price for " + line[0] + " in row " + str(row_num) + "\n")
                    line[6] = float(format_currency(decimal.Decimal(line[6]), "USD").replace('$', ''))
                if Category.objects.get(name=line[7]):
                    continue
                else:
                    line[7] = "uncategorized"
                if line[8] is None:
                    line[8] = 0
                if line[9] is None:
                    line[9] = 0
                if line[10] is None:
                    line[10] = 0

            res.append(line)

        res.remove(res[0])
    else:
        wb = load_workbook(filename=excelFile, read_only=True, data_only=True)
        sheets = wb.sheetnames
        ws = wb[sheets[0]]  # grab the first sheet
        temp = []
        row_num = 0
        row_count = 0
        vmFields = []
        print('testing Snap36 excel template')
        for row in ws.rows:
            count = 0
            for x in range(len(row)):
                print(row_count, x, count)
                if row_count == 0 and count > 0:
                    continue
                if count > 9 and row_count == 1:
                    # print('adding', row[count].value, 'to', vmFields)
                    vmFields.append(str(row[count].value))
                    # print('added')
                    count += 1
                    continue

                if row_count == 0:
                    # print('row1 matching:', row1fieldnamesRefList[count], 'to', str(row[count].value))
                    if str(row1fieldnamesRefList[count]) == str(row[count].value):
                        # print('matched:', row1fieldnamesRefList[count], 'to', str(row[count].value))
                        count += 1
                    else:
                        return False, False, False
                else:
                    # print('row2 matching:', row2fieldnamesRefList[count], 'to', str(row[count].value))
                    if str(row2fieldnamesRefList[count]) == str(row[count].value):
                        # print('matched:', row2fieldnamesRefList[count], 'to', str(row[count].value))
                        count += 1
                    else:
                        return False, False, False
            if row_count == 1:
                break
            else:
                row_count += 1
        # print('vmFields', vmFields)
        print('using snap36 excel template')
        for row in ws.rows:
            row_num += 1
            row_length = len(row)
            if row_num > 2:
                # starts from 0
                if row[0].value is None:
                    internalSKU = ""
                else:
                    internalSKU = str(row[0].value)
                # print('internalSKU', internalSKU)
                if row[1].value is None:
                    upccode = ""
                    upctype = "upc-a"
                else:
                    upccode = str(row[1].value)
                    upctype = "upc-a"
                # print('UPC:', upccode,  upctype)
                if Category.objects.filter(name=row[2].value).count() > 0:
                    cat = str(row[2].value)
                else:
                    cat = "uncategorized"
                # print('category', cat)
                if row[2].value is None:
                    description = ""
                else:
                    description = str(row[2].value)
                # print('description', description)
                if row[3].value is None:
                    standalone = False
                else:
                    if str(row[3].value).lower() == 'true' or str(row[3].value).lower() == 'yes' or \
                            str(row[3].value).lower() == 'y' or str(row[3].value) == str(1):
                        standalone = True
                    elif str(row[3].value).lower() == 'false' or str(row[3].value).lower() == 'no' or \
                            str(row[3].value).lower() == 'n' or str(row[3].value) == str(0):
                        standalone = False
                    else:
                        standalone = False
                # print('standalone', standalone)
                if row[4].value is None:
                    photographyNotes = "No Photography Notes Given"
                else:
                    photographyNotes = str(row[4].value)
                # print('photographyNotes', photographyNotes)
                if row[5].value is None:
                    length = 0
                else:
                    length = str(row[5].value)
                # print('length', length)
                if row[6].value is None:
                    width = 0
                else:
                    width = str(row[6].value)
                # print('width', width)
                if row[7].value is None:
                    height = 0
                else:
                    height = str(row[7].value)
                # print('height', height)
                if row[8].value is None:
                    weight = 0
                else:
                    weight = str(row[8].value)
                # print('weight', weight)
                acDelcoID = ''
                advancedAutoID = ''
                advantageID = ''
                amazonID = ''
                graingerID = ''
                homeDepotID = ''
                johnstoneID = ''
                oReillyID = ''
                walmartID = ''
                vendorCode1 = ''
                vendorCode2 = ''
                vendorCode3 = ''
                # print('made it to the variables')
                if len(vmFields) > 0:
                    count = 9
                    for vmField in vmFields:
                        # print('judging')
                        if vmField == '' or vmField is None or vmField == ' ':
                            count += 1
                            continue
                        elif vmField == 'ACDelco':
                            # print('found ACDelco')
                            acDelcoID = str(row[count].value)
                        elif vmField == 'Advanced Auto':
                            # print('found Advanced Auto')
                            advancedAutoID = str(row[count].value)
                        elif vmField == 'Advantage':
                            # print('found Advantage')
                            advantageID = str(row[count].value)
                        elif vmField == 'Amazon':
                            # print('found Amazon')
                            amazonID = str(row[count].value)
                        elif vmField == 'Grainger SKU':
                            # print('found Grainger SKU')
                            graingerID = str(row[count].value)
                        elif vmField == 'Home Depot':
                            # print('found Home Depot')
                            homeDepotID = str(row[count].value)
                        elif vmField == 'Johnstone':
                            # print('found Johnstone')
                            johnstoneID = str(row[count].value)
                        elif vmField == "O'Reilly":
                            # print("found O'Reilly")
                            oReillyID = str(row[count].value)
                        elif vmField == 'Walmart':
                            # print("found Walmart")
                            walmartID = str(row[count].value)
                        elif vmField == 'VendorCode1':
                            # print("found VendorCode1")
                            vendorCode1 = str(row[count].value)
                        elif vmField == 'VendorCode2':
                            # print("found VendorCode2")
                            vendorCode2 = str(row[count].value)
                        elif vmField == "VendorCode3":
                            # print("found VendorCode3")
                            vendorCode3 = str(row[count].value)
                            # else:
                            print('could not define:', vmField)
                        count += 1

                row_list = [internalSKU, upccode, upctype, cat, description, standalone, photographyNotes,
                            length, width, height, weight, acDelcoID, advancedAutoID, advantageID, amazonID, graingerID,
                            homeDepotID, johnstoneID, oReillyID, walmartID, vendorCode1, vendorCode2, vendorCode3]
            else:
                continue

            if all(row is not None for row in row_list):
                row_json = dict(list(zip(newFieldNames, row_list)))
                if row_json != {}:
                    temp.append(row_json)
            elif not all(row is None for row in row_list):
                un_upload_row_json = dict(list(zip(newFieldNames, row_list)))
                un_upload_product.append(un_upload_row_json)

        res = temp
    # print("error list:", error_list)
    print(res)
    return res, error_list, un_upload_product
