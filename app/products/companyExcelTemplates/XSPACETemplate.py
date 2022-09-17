import codecs
import decimal
import re
import csv
from babel.numbers import format_currency
from app.core.models import Category
from openpyxl import load_workbook

def XSPACETemplate(excelFile):
    error_list = []
    res = []
    un_upload_product = []
    # TODO: [BACKEND][Jacob] add in weight
    fieldnames = ("Product Name", "Product Description", "Product UPC", "Product SKU", "UPC Type", "Manufacturer",
                  "Price", "Product Category", "Product Length", "Product Width", "Product Height")
    fieldnamesRefList = ["Product Name", "Product Description", "Product UPC", "Product SKU", "UPC Type", "Manufacturer",
                  "Price", "Product Category", "Product Length", "Product Width", "Product Height"]
    newFieldNames = ("name", "description", "upccode", "SKU", "UPCType", "manufacturer", "price",
                     "category", "length", "width", "height", "row")


    if excelFile.content_type == 'text/csv':
        reader = csv.reader(codecs.iterdecode(excelFile, 'utf-8'), fieldnames)

        # determine if we are using the default or not
        count = 0
        for fieldname in reader[0]:
            if fieldname == fieldnamesRefList[count]:
                # matches
                count += 1
                continue
            else:
                return False, False, False

        print('using XSPACE excel template')
        row_num = 0
        for line in reader:
            row_num += 1
            if row_num > 1:
                if (not re.match(r'\d+(?:[.]\d{2})?$', line[6])):
                    error_list.append(
                        "Warning. Changed format of price for " + line[0] + " in row " + str(row_num) + "\n")
                    line[6] = float(format_currency(decimal.Decimal(line[6]), "USD").replace('$', ''))
                if Category.objects.get(name=line[7]):
                    continue
                else:
                    line[7] = "uncategorized"
                if line[8] is None:
                    line[8] = 0
                if line[9] is None:
                    line[9] = 0
                if line[10] is None:
                    line[10] = 0

            res.append(line)

        res.remove(res[0])
    else:
        wb = load_workbook(filename=excelFile, read_only=True, data_only=True)
        sheets = wb.sheetnames
        ws = wb[sheets[0]]  # grab the first sheet
        temp = []
        row_num = 0
        count = 0
        for row in ws.rows:
            count = 0
            for x in range(len(row)):
                print(x)
                if str(fieldnamesRefList[x]) == str(row[x].value):
                    # print('matched:', fieldnamesRefList[x], 'to', str(row[x].value))
                    count += 1
                else:
                    return False, False, False
            break
        print('using XSPACE excel template')
        for row in ws.rows:
            row_num += 1
            row_list = []
            newUpc = str(row[4].value)
            price = row[6].value
            if row_num > 1 and price is not None:
                if not re.match(r"[\d+]+[.]+\d{2}$", str(price)):
                    error_list.append(
                        "Warning. changed format of price for " + str(row[0].value) + " in row " + str(
                            row_num) + "\n")
                price = float(format_currency(float(decimal.Decimal(price)), 'USD',
                                              locale='en_US').replace('$', ''))
                if Category.objects.filter(name=row[7].value).count() > 0:
                    cat = row[7].value
                else:
                    cat = "uncategorized"
                if row[5].value is None:
                    l5 = "N/A"
                else:
                    l5 = row[8].value
                if row[8].value is None:
                    l8 = 0
                else:
                    l8 = row[8].value
                if row[9].value is None:
                    l9 = 0
                else:
                    l9 = row[9].value
                if row[10].value is None:
                    l10 = 0
                else:
                    l10 = row[10].value

                if "upc-a" not in newUpc.lower() and "upc-b" not in newUpc.lower() \
                        and "upc-e" not in newUpc.lower():
                    error_list.append(
                        "Warning: Changed format of UPC for " + str(row[0].value) + " in row " + str(
                            row_num) + "\n")
                    newUpc = "upc-a"

                row_list = [row[0].value, row[1].value, row[2].value, row[3].value, newUpc.lower(), l5,
                            price, cat, l8, l9, l10, row_num]

            if all(row is not None for row in row_list):
                row_json = dict(list(zip(newFieldNames, row_list)))
                if row_json != {}:
                    temp.append(row_json)
            elif not all(row is None for row in row_list):
                un_upload_row_json = dict(list(zip(newFieldNames, row_list)))
                un_upload_product.append(un_upload_row_json)

        res = temp
    # print("error list:", error_list)
    return res, error_list, un_upload_product