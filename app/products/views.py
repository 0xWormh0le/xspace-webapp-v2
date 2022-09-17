"""
Products/views
"""
# Django Contrib & Core
from django.db.models import Q
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.contrib.auth import get_user_model
from django.views.generic.detail import DetailView
from django.template import Context
from django.template.loader import render_to_string, get_template
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone
from app.imageEditingFunctions.contentStandardFunctions import load_image
from cv2 import normalize, NORM_MINMAX, CV_64FC1, CV_8UC1

# DRF
import oauth2 as oauth
# Django REST Framework
from rest_framework import viewsets, mixins, parsers, renderers, generics
from rest_framework import generics
from rest_framework.permissions import IsAdminUser, IsAuthenticated
from rest_framework.response import Response
from rest_framework.authentication import SessionAuthentication, BasicAuthentication, TokenAuthentication
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework_jwt.views import obtain_jwt_token
from rest_framework import status
from rest_framework.decorators import api_view
from rest_framework import authentication, permissions
from rest_framework.renderers import JSONRenderer, MultiPartRenderer
from rest_framework.parsers import JSONParser, MultiPartParser, FileUploadParser

from pprint import pprint
from xapi.settings import LOCAL_STORAGE_LOCATION, USE_LOCAL_STORAGE, AWS_STORAGE_BUCKET_NAME, \
    AWS_ACCESS_KEY_ID, AWS_SECRET_ACCESS_KEY, BASE_DIR, LOCAL_STORAGE_URL, AWS_S3_CUSTOM_DOMAIN, CLOUDFRONT_DOMAIN

from django.template.defaultfilters import slugify
from autoslug import AutoSlugField
import ipdb

# Other
import csv
import numpy as np
import urllib.request, urllib.parse, urllib.error, json
from app.core.utility import iterate_and_save_products, AuthenticatedUserMixin, IsNotAuthenticated, ChargebeeAPI, \
    write_asset_order_excel, ZendeskAPI
from app.core.QRGenerator import generateProductOrderQRCode, GenerateQRLabelDoc
from app.orders.models import ContentStandard
from app.imageEditingFunctions.threesixty_file_functions import compile360Bundle, prep360
from elasticsearch import *
import random
import string
import ast
import base64
from PIL import Image
import io
import os
from os.path import normcase
import datetime
import boto3
from wsgiref.util import FileWrapper
from botocore.client import ClientError
from .utility import local_find_files, is_asset, is_company, is_product, is_profile, is_content_standard
from app.imageEditingFunctions.manipulateImage import ManipulateImage
from app.imageEditingFunctions.file_name_functions import delete_file_boto3, findFiles

# templates
from .companyExcelTemplates.snap36Template import snap36Template
from .companyExcelTemplates.XSPACETemplate import XSPACETemplate

from app.accounts.serializers import UploadProductSerializer, EmailThread, CompanyUpdateSerializer, AddressSerializer, \
    CompanySerializer, InvitationSerializer, UserDetailSerializer, Roll_phone_UpdateSerializer, PasswordSerializer, \
    NameUpdateSerializer
from app.products.serializers import EditUploadSerializer, ProductUploadSerializer, ManualProductCreationSerializer, \
    ProductSerializer, ProductMessageSerializer, ProductMessageThreadCreateSerializer, \
    ProductMessageThreadListSerializer, UploaderSerializer, ProductUploadFolderSerializer, \
    ProductListSerializer
from app.products.utility import ProductUploadFolderParser
from app.orders.serializers import FullOrderSerializer
from app.core.models import Upload2Ditem, Category
from app.products.models import Product, ProductMessage, ProductMessageThread, MessageParticipant, ProductAsset
from app.accounts.models import User, Profile, Company, Permission, PermissionAssignment
from rest_framework.exceptions import AuthenticationFailed
from rest_framework import exceptions
from app.core.utility import has_permission
from app.core.utility import master_permissions_list as P
from app.imageEditingFunctions.contentStandardFunctions import product_to_content_standard, \
    product_list_to_content_standard
from notifications.signals import notify
from pathlib import WindowsPath, Path
from app.core.utility import boto3Upload3

try:
    import simplejson as json
except ImportError:
    import json
from dateutil.tz import tzutc

UTC = tzutc()


class ProductsAPI(AuthenticatedUserMixin, APIView):
    permission_classes = (IsAuthenticated,)
    # authentication_classes = (BasicAuthentication,)
    """
    This API is used to render template having Product list, Create, Update and Delete
    if query parameter like 'id' is sent the it will render Product details
    """

    parser_classes = (MultiPartParser, JSONParser)
    renderer_classes = (renderers.JSONRenderer,)

    def parse_file_old(self, myfile):
        """
        parse an excel file for products to create in the database; used with the old format
        :param myfile: path to excel file
        :return: the result of parsing the data in JSON format
        """
        print("!!!!!!!!!!!!!!!!!!!!!!     ProductsAPI.parse_file_old Function   !!!!!!!!!!!!!!!!!!!!!")

        if myfile.content_type == 'text/csv':
            fieldnames = (
                "Product Name", "Product Description", "Product UPC", "Product SKU", "UPC Type", "Manufacturer",
                "Price", "Product Category", "Product Length", "Product Width", "Product Height")
            reader = csv.DictReader(myfile, fieldnames)
            res = [row for row in reader]
            res.remove(res[0])
        else:
            from openpyxl import load_workbook
            wb = load_workbook(filename=myfile, read_only=True)
            sheets = wb.sheetnames
            ws = wb['Sheet1']  # ws is now an IterableWorksheet
            temp = []
            for row in ws.rows:
                row_json = {'Product Name': row[0].value,
                            'Product Description': row[1].value,
                            'Product UPC': row[2].value,
                            'Product SKU': row[3].value,
                            'UPC Type': row[4].value,
                            'Manufacturer': row[5].value,
                            'Price': row[6].value,
                            'Product Category': row[7].value,
                            'Product Length': row[8].value,
                            'Product Width': row[9].value,
                            'Product Height': row[10].value}
                temp.append(row_json)
            temp.remove(temp[0])
            res = temp
        return res

    def parse_file_new(self, excelFile):
        """
        parse an excel/csv file for products to create in the database; used with the new format
        :param excelFile: path to excel file
        :return: [JSON result, List of errors, un_upload_product,]
        """
        print("!!!!!!!!!!!!!!!!!!!!!!     ProductsAPI.parse_file_new Function   !!!!!!!!!!!!!!!!!!!!!")

        # determine which template we are using and act on it if there is a match
        print('parse_file_new received:', excelFile)
        result, error_list, un_upload_product = XSPACETemplate(excelFile)
        if result is False:
            print('not an XSPACE template')
            result, error_list, un_upload_product = snap36Template(excelFile)
            if result is False:
                return result, error_list, un_upload_product, None
            else:
                templateUsed = 'snap36'
        else:
            templateUsed = 'xspace'
        return result, error_list, un_upload_product, templateUsed

    def pretty_request(self, request):
        print("!!!!!!!!!!!!!!!!!!!!!!     ProductsAPI.pretty_request Function   !!!!!!!!!!!!!!!!!!!!!")

        headers = ''
        for header, value in list(request.META.items()):
            if not header.startswith('HTTP'):
                continue
            header = '-'.join([h.capitalize() for h in header[5:].lower().split('_')])
            headers += '{}: {}\n'.format(header, value)

        return (
            '{method} HTTP/1.1\n'
            'Content-Length: {content_length}\n'
            'Content-Type: {content_type}\n'
            '{headers}\n\n'
            '{body}'
        ).format(
            method=request.method,
            content_length=request.META['CONTENT_LENGTH'],
            content_type=request.META['CONTENT_TYPE'],
            headers=headers,
            body=request.body,
        )

    @csrf_exempt
    @has_permission(P.READ_ONLY)
    def get(self, request, *args, **kwargs):
        """
        this gets the product information from the database
        :param request:
        :param args:
        :param kwargs:
        :return:
        """
        print("!!!!!!!!!!!!!!!!!!!!!!     ProductsAPI.GET Function   !!!!!!!!!!!!!!!!!!!!!")

        userID = int(request.user.id)
        username = User.objects.get(pk=userID)
        product_slug = request.query_params['product_slug']
        obj = Product.objects.get(slug=product_slug)
        profile = Profile.objects.get(user=username)
        company = profile.companyProfile
        com = company.slug

        try:
            sort_by = request.query_params['sort']
            if sort_by == 'asc':
                created = 'created'
            elif sort_by == 'des':
                created = '-created'
        except:
            created = '-created'

        try:
            product_object = Product.objects.get(slug=product_slug).as_json()
        except Exception as e:
            return Response({"result": "Invalid request.", "Error:": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        urls = {'url': None, 'success': 'false', 'error': 'None'}
        threeSixtyData = []
        try:
            print("getting 2D assets")
            print('!!!!!!!!!!!!!!!!!obj id', obj.name)
            dbproduct2d = ProductAsset.objects.filter(product=obj.id, assetType=1,
                                                      deletedon='').exclude(thumbnail_url='')
            print("with result:", dbproduct2d)
            if dbproduct2d.count() > 0:
                url = []
                for geturl in dbproduct2d:
                    if USE_LOCAL_STORAGE:
                        res = os.path.join(LOCAL_STORAGE_URL, os.path.relpath(geturl.url, start=LOCAL_STORAGE_LOCATION))
                        if geturl.thumbnail_url:
                            res_tb = os.path.join(LOCAL_STORAGE_URL,
                                                  os.path.relpath(geturl.thumbnail_url, start=LOCAL_STORAGE_LOCATION))
                        else:
                            res_tb = None
                        if geturl.mask_url:
                            res_mask = geturl.mask_url
                        else:
                            res_mask = None
                    else:
                        res = geturl.url
                        if geturl.thumbnail_url:
                            res_tb = geturl.thumbnail_url
                        else:
                            res_tb = None
                        if geturl.mask_url:
                            res_mask = geturl.mask_url
                        else:
                            res_mask = None
                    url.append({'link': res, 'size': geturl.size, 'lastModified': geturl.lastmodified,
                                'thumbnailLink': res_tb, 'maskLink': res_mask})
                urls = {'url': url, 'success': 'true', 'error': 'None'}

            else:
                urls = {'success': 'None', 'error': 'No 2D image found'}
            # TODO: [BACKEND][Jacob] grab videos, miscellaneous, and 3D model assets
            # Get Product 360s

            try:
                print("getting 360 assets")
                print('!!!!!!!!!!!!!!!!!obj id', obj.name)
                dbproduct360 = ProductAsset.objects.filter(product=obj.id, assetType=2, deletedon='')
                print("with result:", dbproduct360)
                if dbproduct360.count() > 0:
                    html_url = []
                    for geturl in dbproduct360:
                        if USE_LOCAL_STORAGE:
                            res = os.path.join(LOCAL_STORAGE_URL,
                                               os.path.relpath(geturl.url, start=LOCAL_STORAGE_LOCATION))
                            if geturl.thumbnail_url:
                                res_tb = os.path.join(LOCAL_STORAGE_URL,
                                                      os.path.relpath(geturl.thumbnail_url,
                                                                      start=LOCAL_STORAGE_LOCATION))
                            else:
                                res_tb = None
                        else:
                            res_html = 'https://' + AWS_S3_CUSTOM_DOMAIN + '/' + str(geturl.html)
                            res_csspath = 'https://' + AWS_S3_CUSTOM_DOMAIN + '/' + str(geturl.css)
                            res_jquerypath = 'https://' + AWS_S3_CUSTOM_DOMAIN + '/' + str(geturl.jquery)
                            res_imagerotatorpath = 'https://' + AWS_S3_CUSTOM_DOMAIN + '/' + str(geturl.imagerotator)
                            res_xmlpath = 'https://' + AWS_S3_CUSTOM_DOMAIN + '/' + str(geturl.xml)
                            res_baseimagepath = geturl.baseImage
                            # due to a previous existing bug, the html and baseimage was sometimes not recorded in the
                            # DB, the folowing two if statements corrects for that
                            if res_baseimagepath == '' or res_baseimagepath is None:
                                path = geturl.css.split('/')
                                commonPart = ''
                                count = 0
                                for part in path:
                                    print("judging", part, "vs '360'")
                                    if part == '360':
                                        commonPart += part + '/' + path[count + 1]
                                        break
                                    else:
                                        commonPart += part + '/'
                                    count += 1
                                total, files = findFiles(commonPart, subd=False)
                                if files is not None:
                                    if len(files) > 0:
                                        res_baseimagepath = files[0]
                                        geturl.baseImage = 'https://' + AWS_S3_CUSTOM_DOMAIN + '/' + str(
                                            geturl.res_baseimagepath)
                                        geturl.save()
                                    else:
                                        res_baseimagepath = None
                                else:
                                    res_baseimagepath = None
                            if res_html == '' or res_html is None:
                                path = geturl.css.split('/')
                                commonPart = ''
                                count = 0
                                for part in path:
                                    print("judging", part, "vs '360'")
                                    if part == '360':
                                        commonPart += part + '/' + path[count + 1]
                                        break
                                    else:
                                        commonPart += part + '/'
                                total, files = findFiles(commonPart, ext='html')
                                if files is not None:
                                    if len(files) > 0:
                                        res_html = files[0]
                                        geturl.html = 'https://' + AWS_S3_CUSTOM_DOMAIN + '/' + str(
                                            geturl.res_html)
                                        geturl.save()
                                    else:
                                        res_html = None
                                else:
                                    res_html = None

                            res_filename = geturl.file_name
                            if geturl.thumbnail_url:
                                res_tb = geturl.thumbnail_url
                            else:
                                res_tb = None
                            html_url.append({'html': res_html,
                                             'css': res_csspath,
                                             'jquery': res_jquerypath,
                                             'imagerotator': res_imagerotatorpath,
                                             'xml': res_xmlpath,
                                             'baseImage': res_baseimagepath,
                                             'size': geturl.size,
                                             'lastModified': geturl.lastmodified,
                                             'thumbnailLink': res_tb,
                                             'filename': res_filename
                                             })

                    threeSixtyData = {'Success': 'True', 'error': 'None', 'html': html_url}
                    print('threeSixtyData:', threeSixtyData)
                else:
                    threeSixtyData = {'Success': 'True', 'error': 'None', 'html': 'None'}
                    print('threeSixtyData:', threeSixtyData)

            except:
                threeSixtyData = {'Success': 'True', 'error': 'None', 'html': 'None'}
            # print("returning statement")
            # print({"threeSixtyAssets": threeSixtyData, "url2d": urls, "productData": product_object})
            print('response', "threeSixtyAssets:", threeSixtyData, "url2d:", urls, "productData:", product_object)
            return Response({"threeSixtyAssets": threeSixtyData, "url2d": urls, "productData": product_object},
                            status=status.HTTP_200_OK)
        except Exception as e:
            # print("returning exception")
            return Response({"url2d": urls, "productData": product_object,
                             "threeSixtyAssets": threeSixtyData,
                             "Error:": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @csrf_exempt
    @has_permission(P.GENERIC)
    def post(self, request, *args, **kwargs):
        print("!!!!!!!!!!!!!!!!!!!!!!     ProductsAPI.POST Function   !!!!!!!!!!!!!!!!!!!!!")
        print('request:')
        print(request.data)
        print(len(request.data))
        print('is a list?:', isinstance(request.data, list))
        excelFile = request.FILES['file'] if request.FILES else None
        # create product(s) using Excel Template
        if excelFile:
            try:
                userID = int(request.user.id)
                # print('userID Angel', userID)
                username = User.objects.get(pk=userID)
                profile = Profile.objects.get(user=username)
                uploaded_product, errors, un_upload_list, templateUsed = self.parse_file_new(excelFile)
                print('template used', templateUsed)
                print('uploaded_product', uploaded_product)
                if uploaded_product is False:
                    return Response({"productInformationList": '',
                                     "errorList": '',
                                     "faultyProductList": '',
                                     "templateUsed": ''},
                                    status=status.HTTP_400_BAD_REQUEST)
                else:
                    return Response({"productInformationList": uploaded_product,
                                     "errorList": errors,
                                     "faultyProductList": un_upload_list,
                                     "templateUsed": templateUsed},
                                    status=status.HTTP_200_OK)
            except IndexError:
                return Response({"type": "unknown",
                                 "message": "Unknown error. Please use template from our Import Excel Sheet page"
                                            " or contact our support team at contact@xspaceapp.com with your"
                                            " spreadsheet attached if you have any further issues."},
                                status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        # create product(s) using a product list
        elif isinstance(request.data, list):
            print("inside list")
            product_error_data = []
            print(1)
            count = 0
            print(2)
            template = request.data[1]
            print(3)
            list_data = request.data[0]
            print("made it past setup")
            for product in list_data:
                mpcs = ManualProductCreationSerializer(data=product)
                product, productUploadStatus = mpcs.create_product(validated_data=product, request=request,
                                                                   template=template)
                print(productUploadStatus)
                if str(productUploadStatus) != str(201):
                    product_error_data.append({count: productUploadStatus})
                count += 1
            if product_error_data == []:
                return Response(status=status.HTTP_201_CREATED)
            else:
                return Response({'error': product_error_data, 'results': mpcs.data,
                                 'message': 'Product(s) Was Not Added On Xspace Platform. Contact your administrator'},
                                status=status.HTTP_400_BAD_REQUEST)

        # manually create the product
        else:
            print("in the manual create")
            serializer = ManualProductCreationSerializer(data=request.data)

            if serializer.is_valid():
                try:
                    # print("Checking if product already exists")
                    product = Product.objects.get(slug=request.data['uniqueID'])
                    return Response(serializer.data, status=status.HTTP_200_OK)
                except:
                    # print("trying to create Product a different way")
                    obj = serializer.create_product(serializer.validated_data, request, '')
                return Response(serializer.data, status=status.HTTP_201_CREATED)
            else:
                # print("Serializer Errors:", serializer.errors)
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @csrf_exempt
    @has_permission(P.GENERIC)
    def put(self, request, *args, **kwargs):
        print("!!!!!!!!!!!!!!!!!!!!!!     ProductsAPI.PUT Function   !!!!!!!!!!!!!!!!!!!!!")
        try:
            product = Product.objects.get(slug=request.data['id'])
            name = request.data['name']
            SKU = request.data['SKU']
            price = request.data['price']
            UPCcode = request.data['upccode']
            description = request.data['description']
            manufacturer = request.data['manufacturer']
            UPCType = request.data['upcType']
            length = request.data['length']
            width = request.data['width']
            height = request.data['height']
            category = Category.objects.get(name=request.data['category'])

            product.name = name
            product.SKU = SKU
            product.price = price
            product.upccode = UPCcode
            product.UPCType = UPCType
            product.description = description
            product.manufacturer = manufacturer
            product.length = length
            product.width = width
            product.height = height
            product.category_id = category.id
            product.save()

            return Response({'response': 'successful', 'error': 'None', 'data': request.data},
                            status=status.HTTP_200_OK)
        except Product.DoesNotExist:
            return Response({'response': 'None', 'error': 'Falied'}, status=status.HTTP_201_CREATED)

        except product.DoesNotExist:
            return Response(status=status.HTTP_400_BAD_REQUEST)

    @csrf_exempt
    @has_permission(P.DELETE_PRODUCT)
    def delete(self, request, *args, **kwargs):
        """
        deletes a product from the user's account
        :param request:
        :param args:
        :param kwargs:
        :return:
        """
        print("!!!!!!!!!!!!!!!!!!!!!!     ProductsAPI.DELETE Function   !!!!!!!!!!!!!!!!!!!!!")

        try:
            print("received Info", request.data)
            user_id = request.data['userid']
            print('userid', user_id)
            profile_obj = Profile.objects.get(user=user_id)
            print(" got the profile as", profile_obj)
            print("got the profile company as", profile_obj.companyProfile.id)
            company = Company.objects.get(id=profile_obj.companyProfile.id)
            print("got the company as", company)
            product_uniqueID = request.data['data']['uniqueID']
            print("working on", product_uniqueID)
            # get the product from the db
            allInstancesOfProduct = Product.objects.filter(uniqueID=product_uniqueID)
            productObjects = Product.objects.filter(uniqueID=product_uniqueID).filter(company_id=company.id)
            if productObjects.count() > 0:
                for productObject in productObjects:
                    print("found", productObject)
                    # unlink the product assets from the company
                    print('searching product assets with', productObject.id, company.id)
                    productAsset_objects = ProductAsset.objects.filter(product=productObject.id).filter(
                        company=company.id)
                    print('product has', productAsset_objects.count(), "product assets")
                    if productAsset_objects.count() > 0:
                        count = 0
                        for productAsset_object in productAsset_objects:
                            print("deleting messageThread")
                            try:
                                ProductMessageThread.objects.get(productAsset_id=productAsset_object.id).delete()
                                print("messageThread successfully deleted")
                            except Exception as e:
                                print('did not delete message thread', e)
                            try:
                                a = productAsset_object.delete()
                                print("product asset was successfully deleted", a)
                            except Exception as e:
                                print('PRODUCT ASSET delete error', e)

                            count += 1
                        print("deleted", count, " product assets")
                    try:
                        # print('product Object', str(productObject))
                        # print('product Object uniqueID', str(productObject.uniqueID))
                        # a = Product.objects.filter(uniqueID=str(productObject.uniqueID)).delete()
                        a = productObject.delete()
                        print('successfully deleted product', a)
                    except Exception as e:
                        print('PRODUCT delete error', e)

            try:
                company.update_storage()
                print("company storage updated")
            except:
                print("company storage not updated")
            return Response({'message': 'Product deleted Successfully', 'status': status.HTTP_200_OK},
                            status=status.HTTP_200_OK)
        except:
            return Response({'message': 'Product not deleted Successfully', 'status': status.HTTP_400_BAD_REQUEST},
                            status=status.HTTP_400_BAD_REQUEST)

    @csrf_exempt
    @has_permission(P.CREATE_PRODUCTASSET)
    def apply_content_standard(self, request, *args, **kwargs):
        """
        applies a content standard to a list of products
        :param request:
        :param args:
        :param kwargs:
        :return:
        """
        try:
            product_list = request.data['product_ids']
            content_standard = request.data['content_standard']
            # let's verify that we are working with a legitimate content standard
            if is_content_standard(content_standard):
                # let's verify that we are working with a 'product list' or 'asset list'
                if is_product(product_list[0]):
                    if len(product_list) > 1:
                        # TODO: [FRONTEND/BACKEND][Angel][Jacob] Since this might take a second, should we just send an
                        #  email when the process is done? or have a progress bar show up at the bottom of the screen
                        #  showing upload time left
                        product_list_to_content_standard(product_list, content_standard, request,
                                                         applyContentStandard=True)
                        return Response({'message': 'Content Standard Successfully Added to Products'},
                                        status=status.HTTP_200_OK)
                    else:
                        product_to_content_standard(product_list, content_standard, request, applyContentStandard=True)
                        return Response({'message': 'Content Standard Successfully Added to Product'},
                                        status=status.HTTP_200_OK)

                else:
                    # we're dealing with specific product assets so we're going to adjust one of the parameters
                    product_to_content_standard(content_standard, request, applyContentStandard=True,
                                                specificAssets=product_list)
                    return Response({'message': 'Content Standard Successfully Added to Product'},
                                    status=status.HTTP_200_OK)
        except:
            return Response({'message': 'Failed to Add Content Standard to Product'},
                            status=status.HTTP_400_BAD_REQUEST)


class ProductThumbnailAPI(AuthenticatedUserMixin, APIView):

    @csrf_exempt
    @has_permission(P.READ_ONLY)
    def get(self, request):
        try:
            product_id_list = request.data['product_ids']
            out = []
            for product_id in product_id_list:
                try:
                    out.append(Product.objects.get(id=product_id).thumbnail)
                except:
                    continue
            if len(out) > 0:
                return Response({'data': out}, status=status.HTTP_200_OK)
            else:
                return Response({'data': None}, status=status.HTTP_404_NOT_FOUND)
        except:
            return Response({'data': None}, status=status.HTTP_404_NOT_FOUND)

    @csrf_exempt
    @has_permission(P.CREATE_PRODUCTASSET)
    def post(self, request):
        product_id = request.data['product_id']
        product_asset_id = request.data['product_asset_id']
        product = Product.objects.get(id=product_id)
        productAsset = ProductAsset.objects.get(id=product_asset_id)
        if productAsset.thumbnail_url is not None:
            try:
                product.thumbnail_url = productAsset.thumbnail_url
                product.save()
                return Response({'result': productAsset.thumbnail_url}, status=status.HTTP_200_OK)
            except:
                return Response({'message': 'Could not add new thumbnail'}, status=status.HTTP_304_NOT_MODIFIED)
        else:
            return Response({'message': 'Thumbnail for product asset does not exist'}, status=status.HTTP_404_NOT_FOUND)

    @csrf_exempt
    @has_permission(P.CREATE_PRODUCTASSET)
    def create(self, request):
        product_id = request.data['product_id']
        product_asset_id = request.data['product_asset_id']
        product_slug = request.query_params['product_slug']
        product_slug = str(product_slug)

        obj = Product.objects.get(slug=product_slug)
        pk = obj.profile.user.id
        username = obj.profile.user.username
        # let's post it
        try:
            product_object = ProductUploadSerializer(data=request.data, context={'product_asset_id': product_asset_id,
                                                                                 'product_id': product_id,
                                                                                 'pk': pk,
                                                                                 'product_slug': product_slug,
                                                                                 'pid': obj.id,
                                                                                 'username': username})
            # let's check to do see if it is correct
            if product_object.is_valid():
                product_object.save()
                return Response(product_object.data, status=status.HTTP_201_CREATED)
            else:
                return Response(product_object.errors, status=status.HTTP_400_BAD_REQUEST)
        except:
            return Response(product_object.errors, status=status.HTTP_400_BAD_REQUEST)


class ProductsSearchAPI(AuthenticatedUserMixin, APIView):
    """
    This API is used to render template having Product list or
    if query parameter like 'id' is sent then it will render Product details
    """
    renderer_classes = (renderers.JSONRenderer,)

    @has_permission(P.READ_ONLY)
    def get(self, request, *args, **kwargs):
        print("!!!!!!!!!!!!!!!!!!!!!!     ProductsSearchAPI.GET Function   !!!!!!!!!!!!!!!!!!!!!")
        try:
            query_string = request.query_params['query_string'] if 'query_string' in request.query_params else None
            user = User.objects.get(pk=request.user.pk)
            profile = Profile.objects.get(user=user)
            company = Company.objects.get(pk=profile.companyProfile.pk)
            obj = Product.objects.filter(company_id=company.pk)
            queryset = obj.filter(Q(name__icontains=str(query_string)) | Q(SKU__icontains=str(query_string)) | Q(
                uniqueID__icontains=str(query_string)))
            serializer = ProductSerializer(queryset, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except:
            return Response({"result": "No Result Found."}, status=status.HTTP_200_OK)


class ProductUploadView(AuthenticatedUserMixin, APIView):
    queryset = Upload2Ditem.objects.all()
    serializer_class = ProductUploadSerializer

    @has_permission(P.READ_ONLY)
    def get(self, request):
        print("!!!!!!!!!!!!!!!!!!!!!!     ProductUploadView.get Function   !!!!!!!!!!!!!!!!!!!!!")
        pk = request.data['userid']
        username = User.objects.filter(pk=pk)
        for user in username:
            username = str(user.username)
        snippets = Upload2Ditem.objects.filter(user_name=username)
        serializer = ProductUploadSerializer(snippets, many=True)
        return Response(serializer.data)

    @csrf_exempt
    @has_permission(P.GENERIC)
    def post(self, request):
        """
        This function adds productAssets (2D, 360, 3D, ...) to a product or if the asset is a zip file/folder it returns
         a dictionary for the frontend to verify with the user before uploading assets to specific locations.
         This function is also the "media upload function"
        :param request:
        :return:
        """
        print("!!!!!!!!!!!!!!!!!!!!!!     ProductUploadView.post Function   !!!!!!!!!!!!!!!!!!!!!")
        # print('{}\n{}\r\n{}\r\n\r\n{}'.format(
        #     '-----------START-----------',
        #     request.method + ' ' + request.url,
        #     '\r\n'.join('{}: {}'.format(k, v) for k, v in request.headers.items()),
        #     request.body,
        #     ))
        print('-----------START-----------')
        try:
            print('METHOD', request.method)
        except:
            pass
        # print('URL',request.url)
        try:
            print('USER', request.user, request.user.id, request.user.username)
        except:
            pass

        try:
            print('QUERYPARAMS', request.query_params)
        except:
            pass

        try:
            print('DATA', request.data)
        except:
            pass
        try:
            print('DATA', request.data['file'])
        except:
            pass

        try:
            print('DATA', type(request.data['file']))
        except:
            pass

        try:
            print('DATA', request.data['file'].content_type)
        except:
            pass

        try:
            # print('DATA', request.data['file'].temporary_file_path())
            # print('DATA', request.data.getList('file'))
            print('FILES', request.FILES)
        except:
            pass

        try:
            print('FILES', request.FILES['file'])
        except:
            pass

        try:
            # print('FILES', request.FILES['file'].temporary_file_path())
            print('FILES', request.FILES['file'].content_type)
        except:
            pass

        try:
            print('FILES', request.FILES['file'].content_type_extra)
        except:
            pass

        try:
            print('FILES', request.FILES.getlist('file'))
        except:
            pass

        try:
            print(request.headers)
        except:
            pass

        try:
            product_slug = request.query_params['product_slug']
            product_slug = str(product_slug)
            using_slug = True
            if product_slug == '' or product_slug is None:
                using_slug = False
        except:
            using_slug = False
        if using_slug:
            if request.data['file'] is not None:
                # print("grabbing single file")
                upload_file = request.data['file']
                # print("request data:", upload_file.read())
                obj = Product.objects.get(slug=product_slug)
                pk = obj.profile.user.id
                username = obj.profile.user.username
                # Let's catch the tricky zip files
                if upload_file.name.endswith('.zip'):
                    folder_parser = ProductUploadFolderParser()
                    zipDictionaryResult = folder_parser.create(
                        validated_data=[request.data,
                                        {'file': upload_file,
                                         'pk': pk,
                                         'product_slug': product_slug,
                                         'pid': obj.id,
                                         'username': username,
                                         'zip': True}])
                    # pprint(zipDictionaryResult)
                    zipDictionaryResult = json.dumps(zipDictionaryResult, default=serialize_WindowsPath)
                    # print('\n was passed forward to the frontend')
                    return Response(zipDictionaryResult, status=status.HTTP_200_OK)
                else:
                    # let's post it to the product we know it goes to
                    product_object = ProductUploadSerializer(data=request.data, context={'file': upload_file,
                                                                                         'pk': pk,
                                                                                         'product_slug': product_slug,
                                                                                         'pid': obj.id,
                                                                                         'username': username})
                # let's check to do see if it is correct
                if product_object.is_valid():
                    product_object.save()

                    return Response(product_object.data, status=status.HTTP_201_CREATED)
                else:
                    return Response(product_object.errors, status=status.HTTP_400_BAD_REQUEST)
        else:
            try:
                # we are likely dealing with the media upload function, let's check that we have a .zip/other file
                # or if it is a folder
                if request.data['file'] is not None and request.data['file'] != '' and \
                        request.data['file'] != ['null'] and request.data['file'] != 'null':
                    print("inside the file uploader")
                    # print("grabbing single file")
                    upload_file = request.data['file']
                    # print("request data:", upload_file.read())
                    # obj = Product.objects.get(slug=product_slug)
                    # pk = obj.profile.user.id
                    pk = request.user.id
                    # username = obj.profile.user.username
                    username = request.user.username
                    # Let's catch the tricky zip files
                    if upload_file.name.endswith('.zip') or 'zip' in upload_file.content_type:
                        print("inside the zip uploader")
                        folder_parser = ProductUploadFolderParser()
                        zipDictionaryResult = folder_parser.create(
                            validated_data=[request.data,
                                            {'file': upload_file,
                                             'pk': pk,
                                             # 'pid': obj.id,
                                             'username': username,
                                             'zip': True}])
                        # print('Did we make it here?')
                        zipDictionaryResult = json.dumps(zipDictionaryResult, default=serialize_WindowsPath)
                        # zipDictionaryResult = json.dumps(zipDictionaryResult)
                        pprint(zipDictionaryResult)
                        print('was passed to the frontend')
                        # for k,v in zipDictionaryResult.items(): print(k,v, "\n\n")
                        return Response(zipDictionaryResult, status=status.HTTP_200_OK)
                    else:
                        # we are doing a media upload, just with no zip file and an arbitrary
                        # number of images as a list... eerily similar to a folder upload...
                        # just with no folder
                        folder_parser = ProductUploadFolderParser()
                        folderDictionaryResult = folder_parser.create(
                            validated_data=[request.data,
                                            {'file': upload_file,
                                             'pk': pk,
                                             # 'pid': obj.id,
                                             'username': username,
                                             'zip': False}])

                        return Response(folderDictionaryResult, status=status.HTTP_201_CREATED)
                elif request.data['FILES'] is not None:
                    # request.FILES should be in the form of the dictionary put out from the zip handler
                    # dictionary from the output of the zip handler generally looks like:
                    #  {'ProductName/UPC/SKU':
                    #             {'uniqueID': [],                    # Product.uniqueID
                    #              'name': [],                        # Product.name
                    #              '2D': ['/image_1.png', ...],       # list of images
                    #              '360': ['/image_1.png', ...],      # list of lists of images each list is a spin
                    #              ...
                    #             },
                    #   'ProductName/UPC/SKU':
                    #             {'uniqueID': [],
                    #              'name': [],
                    #              '2D': ['/image_1.png', ...],
                    #              '360': [['/image_1.png', ...,['/image_1.png']],
                    #              ...
                    #              },
                    # }
                    # TODO: [BACKEND][Jacob] implement folder logic, also need to be allowed to upload folders from the frontend
                    #  NOTE: need to know how angel is passing back a folder
                    print('backend received from media upload frontend:', request.data['FILES'])
                    pk = request.user.id
                    print(pk)
                    username = request.user.username
                    company_identifier = Profile.objects.get(user=pk).companyProfile
                    print(company_identifier)
                    company = Company.objects.get(id=company_identifier.id)
                    print("inside the folder uploader")
                    result = ProductUploadFolderSerializer(data=request.data, context={'files': request.data['FILES'],
                                                                                       'pk': pk,
                                                                                       'username': username,
                                                                                       'company': company
                                                                                       })
                    if result.is_valid():
                        try:
                            result.save()
                            return Response(result.data, status=status.HTTP_201_CREATED)
                        except:
                            return Response(None, status=status.HTTP_400_BAD_REQUEST)
                    else:
                        return Response(None, status=status.HTTP_400_BAD_REQUEST)
                else:
                    return Response("We cannot do anything with this", status=status.HTTP_400_BAD_REQUEST)
                    pass
            except Exception as e:
                print(e)


class EditingUploadView(AuthenticatedUserMixin, APIView):
    queryset = Upload2Ditem.objects.all()
    serializer_class = EditUploadSerializer

    @has_permission(P.GENERIC)
    def post(self, request):
        print("!!!!!!!!!!!!!!!!!!!!!!     EditingUploadView.post Function   !!!!!!!!!!!!!!!!!!!!!")

        try:
            # let's gather the info we need so that the serializer has a fighting chance
            orderID = request.query_params['orderID']
            orderID = str(orderID)
            upload_file = request.data['file']
            pk = request.user.id
            username = request.user.username

            # let's post it
            product_serializer = EditUploadSerializer(data=request.data,
                                                      context={'file': upload_file,
                                                               'pk': pk,
                                                               'orderID': orderID,
                                                               'username': username})
            # let's make sure that what we did is valid
            if product_serializer.is_valid():
                product_serializer.save()
                # if so, let's let the frontend know of our endeavors
                return Response(product_serializer.data, status=status.HTTP_201_CREATED)
            else:
                return Response(product_serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            print(e)


def get_all_prods_for_sub_companies(company):
    print("!!!!!!!!!!!!!!!!!!!!!!     get_all_prods_for_sub_companies Function   !!!!!!!!!!!!!!!!!!!!!")

    firstProds = list(Product.objects.filter(company=company))
    subComp = Company.objects.filter(parentCompany=company)
    all_prods = firstProds
    try:
        for comp in subComp:
            if comp.id == company.id:
                continue
            else:
                all_prods += get_all_prods_for_sub_companies(comp)
    except:
        return all_prods
    return all_prods


def serialize_WindowsPath(obj):
    if isinstance(obj, Path):
        serial = str(obj)
    elif isinstance(obj, WindowsPath):
        serial = str(obj)
        return serial
    else:
        raise TypeError("Type not serializable")


class ProductListViewSet(AuthenticatedUserMixin, viewsets.ModelViewSet):
    serializer_class = ProductListSerializer

    @csrf_exempt
    @has_permission(P.READ_ONLY)
    def list(self, request, *args, **kwargs):
        print("!!!!!!!!!!!!!!!!!!!!!!     Product List View.LIST Function   !!!!!!!!!!!!!!!!!!!!!")

        userID = request.data['userid']
        profile = Profile.objects.get(user=userID)
        # set = Product.objects.filter(company=profile.companyProfile)
        newSet = []
        qs = get_all_prods_for_sub_companies(profile.companyProfile)

        if int(len(qs)) == int(0):
            return Response(newSet, status=status.HTTP_404_NOT_FOUND)
        for t in qs:
            # TODO: [BACKEND][Jacob] there is a better way to do this
            # if type(t) != type([]):
            product_object = Product.objects.get(id=t.id)
            # test for content standard compliance
            standards = ContentStandard.objects.filter(name=product_object.uniqueID)
            # print("found ", standards.count(), "content standards for", product_object.name)
            if standards.count() > 0:
                for standard in standards:
                    try:
                        # print("trying to apply content standard listed as:", standard.name)
                        # product_object.compliance_message = product_object.get2DassetCompliance(standard)
                        needsChanged = True
                        if product_object.compliance_message == '':
                            if product_object is True:
                                needsChanged = False
                            else:
                                product_object.compliance = True
                        else:
                            if product_object is True:
                                product_object.compliance = False
                            else:
                                needsChanged = False
                        if needsChanged:
                            product_object.save()
                    except:
                        product_object.compliance_message = 'Content Standard Applied, but system errored out'
                        product_object.compliance = False
                        product_object.save()
                        # TODO: [BACKEND][Jacob] default to XSPACE standard
                        # print("skipped!")
                        continue
            else:
                needsChanged = False
                # print("compliance_message listed as:", product_object.compliance_message)
                if product_object.compliance_message is None:
                    # print("product has no content standard applied to it, and doesn't have a message")
                    product_object.compliance_message = 'No Content Standard Applied To This Product'
                    product_object.compliance = True
                    needsChanged = True
                else:
                    if 'No Content' in product_object.compliance_message:
                        if product_object.compliance is True:
                            product_object.compliance = False
                            needsChanged = True
                    else:
                        # print("failed second if test")
                        product_object.compliance_message = str(
                            'No Content Standard Applied To This Product, ' + product_object.compliance_message)
                        product_object.compliance = False
                        needsChanged = True
                try:
                    if needsChanged:
                        product_object.save()
                except:
                    print("product updated locally but changes were not saved to the"
                          "DB passing forward local stored info")
            product_object = product_object.as_json()
            # print("passing forward:", product_object['compliance'], product_object['compliance_message'])
            if product_object['thumbnail']:
                if USE_LOCAL_STORAGE:
                    # print("changing", product_object['thumbnail'])
                    product_object['thumbnail'] = os.path.join(LOCAL_STORAGE_URL,
                                                               os.path.relpath(product_object['thumbnail'],
                                                                               start=LOCAL_STORAGE_LOCATION))[
                                                  1:].replace('\\', '/')
                    # print('to', product_object['thumbnail'])
            newSet.append(product_object)
        # print("Output from query search:", newSet['compliance_message'])
        return Response(newSet, status=status.HTTP_200_OK)


# TODO: [BACKEND][Dom] Convert to RSAA Pattern
class multifill_db_2dscript(APIView):
    @has_permission(P.READ_ONLY)
    def get(self, request, username):
        print("!!!!!!!!!!!!!!!!!!!!!!     multifill_db_2dscript.GET Function   !!!!!!!!!!!!!!!!!!!!!")
        try:
            # print("trying to get username")
            get_username = User.objects.get(username=username)
            try:
                profile = Profile.objects.get(user=get_username)
                # print("get_username:", get_username)
                # print("profile:", profile)
                company = Company.objects.get(pk=profile.companyProfile.pk)
                set = Product.objects.filter(company=company.pk)
                valuesSet = list(set.values())
                print('set', set, 'valueset', valuesSet)
                # print("company:", company)
                # print("set:", set)
                # print("valueSet:", valuesSet)

                files = []
                file_size = []
                created_date = []
                p_id = []
                key = []
                fill_list = []

                # product_list = Product.objects.filter(profile=profile)

                # Check if each product has 2D images or not.
                for prod in valuesSet:
                    # images = ProductAsset.objects.filter(assetType=1,product=int(prod["id"])).count()
                    # test this
                    images = ProductAsset.objects.filter(product=int(prod["id"])).count()
                    print(images)
                    if images == 0:
                        print('appending images')
                        fill_list.append(prod)
                    else:
                        print('not appending images')
                    print('########')

                for item in fill_list:
                    product = item
                    p_id.append(product["id"])

                    # print '/'+str(get_username.username)+'/product/'+str(product.pk)+'/'+str(product.slug)+'/360/'
                    key.append(str(get_username.username) + '/product/' + str(product["id"]) + '/' + str(
                        product["slug"]) + '/2D/')

                    if USE_LOCAL_STORAGE:
                        pass
                    else:
                        bucketname = AWS_STORAGE_BUCKET_NAME
                        s3 = boto3.resource('s3', aws_access_key_id=AWS_ACCESS_KEY_ID,
                                            aws_secret_access_key=AWS_SECRET_ACCESS_KEY)
                        data = s3.Bucket(bucketname)

                    list_product = []

                    for i in range(len(key)):
                        for item in data.objects.filter(Prefix=key[i]):
                            test = ''
                            for k in range(len(item.key)):
                                if k == (len(item.key) - 1):
                                    test += item.key[k]
                            if test != '/':
                                # print str(item.size/1024)+' kb'
                                file_size.append(item.size / 1024)
                                files.append('https://dbrdts7bui7s7.cloudfront.net/' + item.key)
                            created_date.append(item.last_modified)

                    for j in range(len(files)):
                        # print files[j]
                        split_files = files[j].split('/')
                        count = ProductAsset.objects.filter(product=split_files[5],
                                                            url=str(files[j]),
                                                            assetType=1).count()

                        prod = Product.objects.get(pk=split_files[5])

                return Response({'success': 'true', 'error': 'None', 'files': files})

            except Product.DoesNotExist:
                return Response({'success': 'None', 'error': 'no products found for this user.'})

        except User.DoesNotExist:
            return Response({'success': 'None', 'error': 'User not found.'})


class fill_db_script(AuthenticatedUserMixin, APIView):
    @has_permission(P.READ_ONLY)
    def get(self, request, slug):
        print("!!!!!!!!!!!!!!!!!!!!!!     fill_db_script.GET Function   !!!!!!!!!!!!!!!!!!!!!")

        try:
            try:
                profile_object = Profile.objects.get(email=request.user)
                # userID = int(request.user.id)
                # username = User.objects.get(pk=userID)
                # profile = Profile.objects.get(user=username)

                company_object = Company.objects.get(uniqueID=profile_object.companyProfile)

                files = []
                p_id = []
                key = []

                product_object = Product.objects.get(company=company_object.uniqueID, slug=slug)

                p_id.append(product_object.uniqueID)

                key.append(str(profile_object.companyProfile.slug) + '/products/' + str(slug) + '/360/')

                # let's locate where the eff this product's assets are supposed to be
                if USE_LOCAL_STORAGE:
                    if not os.path.isdir(os.path.normcase(os.path.join(LOCAL_STORAGE_LOCATION, str(key[0])))):
                        os.makedirs(os.path.normcase(os.path.join(LOCAL_STORAGE_LOCATION, str(key[0]))))
                        return Response({'success': 'None', 'error': 'no assets found for this user.'})
                    else:
                        # let's grab the local files
                        fileCount, out = local_find_files(
                            os.path.normcase(os.path.join(LOCAL_STORAGE_LOCATION, str(key[0]))))

                        # let's filter our results, since we're working with multiple file types
                        for i in range(len(key)):
                            if fileCount > 0:
                                for item in out:
                                    temp_loc = os.path.join(LOCAL_STORAGE_LOCATION, item['key'])
                                    if item['key'].endswith('xml') or item['key'].endswith('css') \
                                            or item['key'].endswith('.js'):
                                        files.append(os.path.normcase(temp_loc))
                else:
                    # looks like we're using s3, different logic to get the files follows
                    bucketname = AWS_STORAGE_BUCKET_NAME
                    s3 = boto3.resource('s3', aws_access_key_id=AWS_ACCESS_KEY_ID,
                                        aws_secret_access_key=AWS_SECRET_ACCESS_KEY)
                    data = s3.Bucket(bucketname)

                    s3Host = 'https://s3.amazonaws.com/storagev2/'

                    for i in range(len(key)):
                        # print(data.objects.filter(Prefix=key[i]))
                        for item in data.objects.filter(Prefix=key[i]):
                            for k in range(len(item.key)):
                                if item.endswith('xml') or item.endswith('.js') or item.endswith('.css'):
                                    files.append(s3Host + item.key)

                if len(files) > 0:
                    xml_list = [sxml for sxml in files if ".xml" in sxml]
                    css_list = [scss for scss in files if "basic.css" in scss]
                    js_list = [sjs for sjs in files if ".js" in sjs]

                    xml_array = []
                    basic_array = []
                    css_array = []
                    js1_array = []
                    js2_array = []
                    s3product_list = []
                    # product360 = Product360Image()

                    for n in range(len(xml_list)):
                        xml_data = xml_list[n].split('/')
                        try:
                            if xml_data[-2] == '360_assets':
                                # print xml_list[n]
                                xml_array.append(xml_list[n])
                        except Exception as e:
                            print('no 360 image for this product')
                            break

                    for m in range(len(css_list)):
                        css_data = css_list[m].split('/')
                        try:
                            if css_data[7] == '360':
                                # print css_list[m]
                                css_array.append(css_list[m])
                                # print str(css_data[0])+'/'+str(css_data[1])+'/'+str(css_data[2])+'/'+str(css_data[3])+'/'+str(css_data[4])+'/'+str(css_data[5])+'/'+str(css_data[6])+'/'+str(css_data[7])+'/'+str(css_data[8])+'/'+str(css_data[9])+'/'+str(css_data[10])+'/'+str(css_data[11])+'/img/basic'
                                basic_array.append(
                                    str(css_data[0]) + '/' + str(css_data[1]) + '/' + str(css_data[2]) + '/' + str(
                                        css_data[3]) + '/' + str(css_data[4]) + '/' + str(css_data[5]) + '/' + str(
                                        css_data[6]) + '/' + str(css_data[7]) + '/' + str(css_data[8]) + '/' + str(
                                        css_data[9]) + '/' + str(css_data[10]) + '/' + str(css_data[11]) + '/img/basic')
                                s3product_list.append(css_data[5])
                        except Exception as e:
                            print('no 360 image for this product')

                    for p in range(len(js_list)):
                        js_data = js_list[p].split('/')
                        try:
                            if js_data[7] == '360':
                                # print js_list[p]
                                if p % 2:
                                    js1_array.append(js_list[p])
                                else:
                                    js2_array.append(js_list[p])

                        except Exception as e:
                            print('no 360 image for this product')

                    for row in range(len(s3product_list)):
                        count = ProductAsset.objects.filter(product_id=s3product_list[row], assetType=2).count()
                        if count == 0:
                            # new_row = Product360Image()
                            # new_row.product_id = s3product_list[row]
                            # new_row.xml = str(xml_array[row])
                            # new_row.baseImage = str(basic_array[row])
                            # new_row.css = str(css_array[row])
                            # new_row.jquery = str(js1_array[row])
                            # new_row.imagerotator = str(js2_array[row])
                            # new_row.size = 20.00
                            # new_row.save()

                            '''
                            print 'Existing Row'
                            print 'xml: '+str(xml_array[row])
                            print 'basic: '+str(basic_array[row])
                            print 'css: '+str(css_array[row])
                            print 'js1: '+str(js1_array[row])
                            print 'js2: '+str(js2_array[row])
                            '''
                        else:
                            replace_value = ProductAsset.objects.get(assetType=2, product_id=s3product_list[row])
                            replace_value.xml = str(xml_array[row])
                            replace_value.baseImage = str(basic_array[row])
                            replace_value.css = str(css_array[row])
                            replace_value.jquery = str(js1_array[row])
                            replace_value.imagerotator = str(js2_array[row])
                            replace_value.size = 20.00
                            replace_value.save()

                            '''
                            print 'New Row'
                            print 'xml: '+str(xml_array[row])
                            print 'basic: '+str(basic_array[row])
                            print 'css: '+str(css_array[row])
                            print 'js1: '+str(js1_array[row])
                            print 'js2: '+str(js2_array[row])
                            '''

                    '''
                    for j in range(len(files)):
                        if j>4:
                            files.pop()
                    '''
                    return Response({'success': 'true', 'error': 'None',
                                     'files': files,
                                     'xml': xml_array,
                                     'basic': basic_array,
                                     'css': css_array,
                                     'js1': js1_array,
                                     'js2': js2_array,
                                     'key': key,
                                     'p_id': p_id,
                                     's3_pid': s3product_list
                                     })
                else:
                    # we didn't find anything so let's just return home with our heads held high
                    return Response({'success': 'None', 'error': 'no products found for this user.'})
            except Product.DoesNotExist:
                return Response({'success': 'None', 'error': 'no products found for this user.'})

        except Exception as e:
            return Response({'success': 'None', 'error': str(e)})


# TODO: [BACKEND][Dom] Convert to RSAA Pattern
class multifill_db_2dscript(APIView):

    @has_permission(P.READ_ONLY)
    def get(self, request, username):
        print("!!!!!!!!!!!!!!!!!!!!!!     multifill_db_2dscript.GET Function   !!!!!!!!!!!!!!!!!!!!!")

        try:
            get_username = User.objects.get(username=username)
            try:

                profile = Profile.objects.get(user=get_username)
                # products = Product.objects.filter(profile=profile)

                company = Company.objects.get(pk=profile.companyProfile.pk)
                set = Product.objects.filter(company_id=company.pk)

                valuesSet = list(set.values())

                files = []
                file_size = []
                created_date = []
                p_id = []
                key = []
                fill_list = []

                # product_list = Product.objects.filter(profile=profile)

                # Check if each product has 2D images or not.
                for prod in valuesSet:
                    images = ProductAsset.objects.filter(assetType=1, product=int(prod["id"])).count()
                    if (images == 0):
                        fill_list.append(prod)

                for item in fill_list:

                    product = item

                    p_id.append(product["id"])

                    # print '/'+str(get_username.username)+'/product/'+str(product.pk)+'/'+str(product.slug)+'/360/'
                    key.append(str(get_username.username) + '/product/' + str(product["id"]) + '/' + str(
                        product["slug"]) + '/2D/')

                    bucketname = AWS_STORAGE_BUCKET_NAME
                    s3 = boto3.resource('s3', aws_access_key_id=AWS_ACCESS_KEY_ID,
                                        aws_secret_access_key=AWS_SECRET_ACCESS_KEY)
                    data = s3.Bucket(bucketname)

                    list_product = []

                    for i in range(len(key)):
                        for item in data.objects.filter(Prefix=key[i]):
                            test = ''
                            for k in range(len(item.key)):
                                if k == (len(item.key) - 1):
                                    test += item.key[k]
                            if test != '/':
                                # print str(item.size/1024)+' kb'
                                file_size.append(item.size / 1024)
                                files.append('https://dbrdts7bui7s7.cloudfront.net/' + item.key)
                            created_date.append(item.last_modified)

                    for j in range(len(files)):
                        # print files[j]
                        split_files = files[j].split('/')
                        count = ProductAsset.objects.filter(assetType=1, product_id=split_files[5],
                                                            url=str(files[j])).count()

                        prod = Product.objects.get(pk=split_files[5])

                        # print ProductAsset.objects.filter(assetType=1,product_id=split_files[5], url=str(files[j]))
                        # print count

                        if count == 0:
                            print(' if')
                            # new_row = Product2DImage()
                            # new_row.product_id = split_files[5]
                            # new_row.url = files[j]
                            # new_row.size = file_size[j]
                            # #currentDT = datetime.now()
                            # new_row.lastmodified = created_date[j]
                            # new_row.save()
                            #
                            # new_thread = ProductMessageThread()
                            # new_thread.product = prod
                            # new_thread.product2DImage = new_row
                            # new_thread.messageThreadName = str(split_files[8])
                            # currentDT = datetime.datetime.now()
                            # new_thread.creationDateTime = currentDT
                            # new_thread.lastUpdateDateTime = currentDT
                            # new_thread.save()

                '''
                for j in range(len(files)):
                    split_files = files[j].split('/')
                    count = ProductAsset.objects.filter(assetType=1,product_id=split_files[5], url=str(files[j])).count()

                    print ProductAsset.objects.filter(assetType=1,product_id=split_files[5], url=str(files[j]))
                    print count

                    if count != 0:
                        print ' else'

                        replace_value = ProductAsset.objects.filter(assetType=1,product_id=split_files[5])#.update(url=str(files[j]),size=str(file_size[j]),lastmodified = created_date[j])
                        print replace_value
                        for rep in replace_value:
                            rep.url = files[j]
                            rep.size = file_size[j]
                            rep.lastmodified = created_date[j]
                            rep.save()
                '''

                return Response({'success': 'true', 'error': 'None', 'files': files})

            except Product.DoesNotExist:
                return Response({'success': 'None', 'error': 'no products found for this user.'})

        except User.DoesNotExist:
            return Response({'success': 'None', 'error': 'User not found.'})


# TODO: [BACKEND][Dom] Convert to RSAA Pattern
class fill_db_2dscript(AuthenticatedUserMixin, APIView):

    @has_permission(P.READ_ONLY)
    def get(self, request, slug):
        """
        grab the 2D images from a database... this is a dumb function, not modular but I didn't write it - JD
        :param request:
        :param slug: Product slug
        :return: Response({'success': BOOLEAN, 'error': str(error), 'files': files})
        """
        print("!!!!!!!!!!!!!!!!!!!!!!     fill_db_2dscript.GET Function   !!!!!!!!!!!!!!!!!!!!!")
        try:
            # let's see who we are dealing with
            profile_object = Profile.objects.get(email=request.user)
            # print(request.user)
            try:
                files = []
                key = []

                # let's define the relative path
                key.append(str(profile_object.companyProfile.slug) + '/products/' + str(slug) + '/2D/')

                # let's locate where the eff this product's assets are supposed to be
                if USE_LOCAL_STORAGE:
                    if not os.path.isdir(os.path.normcase(os.path.join(LOCAL_STORAGE_LOCATION, str(key[0])))):
                        os.makedirs(os.path.normcase(os.path.join(LOCAL_STORAGE_LOCATION, str(key[0]))))
                        return Response({'success': 'None', 'error': 'no assets found for this user.'})
                    else:
                        # let's grab the files
                        fileCount, out = local_find_files(
                            os.path.normcase(os.path.join(LOCAL_STORAGE_LOCATION, str(key[0]))))
                        # print("getting product assets from", os.path.normcase(os.path.join(LOCAL_STORAGE_LOCATION,
                        #                                                                    str(key[0]))))

                        if int(fileCount) > int(0):
                            for i in out:
                                temp_loc = os.path.join(LOCAL_STORAGE_LOCATION, i['key'])
                                files.append(os.path.normcase(temp_loc))
                        else:
                            # print("couldn't find any 2D assets with the fill_db_2dscript")
                            return Response({'success': 'None', 'error': 'no assets found for this user.'})

                else:
                    # since we're not using local storage, let's follow some different logic to grab it from s3
                    bucketname = AWS_STORAGE_BUCKET_NAME
                    s3 = boto3.resource('s3', aws_access_key_id=AWS_ACCESS_KEY_ID,
                                        aws_secret_access_key=AWS_SECRET_ACCESS_KEY)
                    data = s3.Bucket(bucketname)

                    # # check if bucket exists
                    # try:
                    #     s3.meta.client.head_bucket(Bucket=data.name)
                    # except ClientError:
                    #     # The bucket does not exist or you have no access.
                    #     print("bucket does not exist or we don't have access")

                    for i in range(len(key)):
                        for item in data.objects.filter(Prefix=key[i]):
                            # print(item)
                            if item.key[-1] != '/':
                                # file_size.append(str(item.size / 1024) + ' kb')
                                files.append('https://storage.xspaceapp.com/' + item.key)
                return Response({'success': 'true', 'error': 'None', 'files': files})

            except Product.DoesNotExist:
                return Response({'success': 'None', 'error': 'no products found for this user.'})

        except User.DoesNotExist:
            return Response({'success': 'None', 'error': 'User not found.'})


class s3_create_folder(generics.RetrieveAPIView):

    @has_permission(P.ALL)
    def post(self, request):
        print("!!!!!!!!!!!!!!!!!!!!!!     s3_create_folder.post Function   !!!!!!!!!!!!!!!!!!!!!")

        pk = request.data['userId']
        request1 = request.data['request']
        if request1 == 'get':
            try:
                # print((request.data['viewFolder']))
                pk = request.data['userId']
                get_username = User.objects.get(pk=pk)

                username = User.objects.filter(pk=pk)
                for user in username:
                    username = str(user.username)

                bucketname = AWS_STORAGE_BUCKET_NAME
                s3 = boto3.resource('s3', aws_access_key_id=AWS_ACCESS_KEY_ID,
                                    aws_secret_access_key=AWS_SECRET_ACCESS_KEY)
                data = s3.Bucket(bucketname)

                files = []
                key = username + base64.b64decode(request.data['viewFolder'])

                try:
                    for item in data.objects.filter(Prefix=key):
                        test = ''
                        checkName = []
                        for k in range(len(item.key)):
                            if k == (len(item.key) - 1):
                                test += item.key[k]
                        if test == '/':
                            checkName = item.key.split('/')
                            if len(checkName) > 5 and checkName[4] != '2D' and checkName[4] != '360':
                                # print checkName[4]
                                # app_product_miscellaneous_asset
                                files.append('https://dbrdts7bui7s7.cloudfront.net/' + item.key)

                    return Response({'success': 'true', 'error': 'None', 'files': files})
                except Product.DoesNotExist:
                    return Response({'success': 'None', 'error': 'no products found for this user.'})

            except User.DoesNotExist:
                return Response({'success': 'None', 'error': 'User not found.'})
        elif request1 == 'post':
            location = request.data['location']

            username = User.objects.filter(pk=pk)
            for user in username:
                username = str(user.username)

            bucketname = AWS_STORAGE_BUCKET_NAME
            s3 = boto3.resource('s3', aws_access_key_id=AWS_ACCESS_KEY_ID,
                                aws_secret_access_key=AWS_SECRET_ACCESS_KEY)
            data = s3.Bucket(bucketname)

            action = boto3.client('s3', aws_access_key_id=AWS_ACCESS_KEY_ID,
                                  aws_secret_access_key=AWS_SECRET_ACCESS_KEY)
            result = action.put_object(Bucket=bucketname, Key=username + location)
            # result = data.put(key='admin/product/416/rvoxhyuzeh/2D/'+location)

            files = []
            key = username + request.data['viewFolder']
            try:
                get_username = User.objects.get(pk=pk)
                try:
                    username = User.objects.filter(pk=pk)
                    profile = Profile.objects.filter(user=username)
                    products = Product.objects.filter(profile=profile)

                    for item in data.objects.filter(Prefix=key):
                        test = ''
                        for k in range(len(item.key)):
                            if k == (len(item.key) - 1):
                                test += item.key[k]
                        if test == '/':
                            files.append('https://dbrdts7bui7s7.cloudfront.net/' + item.key)

                except Product.DoesNotExist:
                    return Response({'success': 'None', 'error': 'no products found for this user.'})

            except User.DoesNotExist:
                return Response({'success': 'None', 'error': 'User not found.'})

            return Response({'success': 'None', 'error': 'None', 'result': result, 'files': files, 'key': key})


class s3_move_files(AuthenticatedUserMixin, generics.RetrieveAPIView):

    @has_permission(P.ALL)
    def post(self, request):
        print("!!!!!!!!!!!!!!!!!!!!!!     s3_move_files.post Function   !!!!!!!!!!!!!!!!!!!!!")
        # TODO: [BACKEND][Jacob] add in .archive protection
        product_slug = request.query_params['product_slug']
        product_slug = str(product_slug)
        upload_file = request.data['file']
        lastModifiedList = request.data['lastModified']
        product_obj = Product.objects.get(slug=product_slug)
        pk = product_obj.profile.user.id
        username = product_obj.profile.user.username
        profile = product_obj.profile

        key = upload_file.split(',')
        lastModifiedList = lastModifiedList.split(',')
        result = []
        print('here inside s3 post function')
        # s3 = boto3.resource('s3', aws_access_key_id='AKIAJKV6HI4YIS35FTVA',
        #                     aws_secret_access_key='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')
        #
        # action = boto3.client('s3', aws_access_key_id='AKIAI3IFQOJ2PKPBLOVQ',
        #                       aws_secret_access_key='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')

        if USE_LOCAL_STORAGE:
            # locally remove the instance of the image
            for i in range(len(key)):
                # delete the file
                key[i] = key[i].replace(LOCAL_STORAGE_URL, '')
                key[i] = key[i].replace('/', '\\')
                print(key[i], LOCAL_STORAGE_LOCATION, LOCAL_STORAGE_URL)
                STORAGE_KEY = os.path.join(LOCAL_STORAGE_LOCATION, key[i])
                print("trying to remove", STORAGE_KEY)
                os.remove(STORAGE_KEY)
                print("removed", STORAGE_KEY)
                # delete the corresponding thumbnail
                a, b = os.path.split(STORAGE_KEY)
                a = os.path.join(os.path.join(a, 'thumbnails'), b)
                os.remove(a)
                del a
                del b
                try:
                    print("finding record")
                    # delete the record

                    rec = ProductAsset.objects.get(url=normcase(STORAGE_KEY))
                    print("found record")
                    # delete the thumbnail
                    ProductAsset.objects.get(url=rec.thumbnail_url).delete()
                    # if the thumbnail was the product thumbnail, delete association
                    print("deleted thumbnail")
                    print("checking thumbnail url")
                    print(rec)
                    print(rec.thumbnail_url)
                    print(product_obj.thumbnail)
                    if rec.thumbnail_url == product_obj.thumbnail:
                        # "delete" record
                        print("deleted product thumbnail association")
                        product_obj.thumbnail = None
                        product_obj.save()

                    try:
                        # delete the message thread associated with that record
                        print("trying to delete message thread")
                        ProductMessageThread.objects.get(productAsset_id=rec.id).delete()
                        print("deleted message thread")
                    except:
                        print("could not delete the message thread associated with the record")
                    rec.delete()
                    print("product asset successfully deleted")
                except:
                    print("Tried, but did not fully delete:", key[i])
        else:
            # remove the instance of the image on s3 and cloudfront, if the user is the original owner, then it is
            # deleted otherwise it is unlinked from the users account
            count = 0
            negativeResponses = []
            print('delete received:', request, request.data, request.user)
            profile_object = Profile.objects.get(email=request.user)
            print('key', key)
            assetType = key[1]
            print('assetType', assetType)
            key = [key[0]]
            for i in key:
                if assetType == '1':
                    print('wanting to delete a picture')
                    res, errors, convertedLocation = delete_file_boto3([i])
                    if res:
                        print('deleted:', convertedLocation)
                    else:
                        print('error deleting:', errors, i)
                    try:
                        # delete the record
                        rec = ProductAsset.objects.get(url=i.replace('\\', '/'),
                                                       lastmodified=lastModifiedList[count][:-1].replace('T', ' '))
                        # delete the thumbnail
                        print('record name', rec.file_name)
                        try:
                            # if the thumbnail is a default or null then let's not try to delete it
                            if str(rec.thumbnail_url) == 'XSPACE-subtle-color.png' or str(
                                    rec.thumbnail_url) is None or str(rec.thumbnail_url) == '':
                                pass
                            else:
                                # if the thumbnail was the product thumbnail, delete association
                                print('trying to find:', rec.thumbnail_url)
                                if rec.thumbnail_id == 'null' or rec.thumbnail_id == '' or rec.thumbnail_id is None:
                                    # record does not have a thumbnail that has been updated to the new tracking schema
                                    pass
                                elif rec.thumbnail_url == 'null' or rec.thumbnail_url == '' or rec.thumbnail_url is None:
                                    # record does not have a thumbnail
                                    pass
                                else:
                                    thumbnail = ProductAsset.objects.get(id=rec.thumbnail_id)
                                    print('found:', thumbnail.url)
                                    thumbnail_res, errors, convertedLocation = delete_file_boto3([thumbnail.url])
                                    if thumbnail_res:
                                        print('deleted:', convertedLocation)
                                    else:
                                        print('error deleting:', errors, i)
                                    thumbnail.delete()
                                    print('removed thumbnail from the database')
                                    if rec.thumbnail_url == product_obj.thumbnail:
                                        # "delete" record
                                        product_obj.thumbnail = 'XSPACE-subtle-color.png'
                                        product_obj.save()
                        except:
                            # record doesn't have a thumbnail
                            pass
                        try:
                            # delete the message thread associated with that record
                            ProductMessageThread.objects.get(productAsset_id=rec.id).delete()
                        except:
                            print("could not delete the message thread associated with the record")
                            negativeResponses.append(i)
                        try:
                            print('rec', rec)
                            # rec.clear()
                            rec.delete()
                            print("product asset successfully deleted")
                        except Exception as e:
                            print("product asset was not successfully deleted", e)
                            try:
                                ProductAsset.objects.get(id=rec.id).delete()
                            except Exception as e:
                                print('still errored after relookup', e)
                    except:
                        print("Tried, but did not fully delete:", i)
                        negativeResponses.append(i)
                elif assetType == '2':
                    print('wanting to delete a 360')
                    # for now we are just unlinking the 360
                    i = i.replace('\\', '/')
                    rec = ProductAsset.objects.filter(baseImage=i).filter(assetType=2)
                    if rec.count() == 0:
                        if i.startswith('https://' + AWS_S3_CUSTOM_DOMAIN):
                            print('swapping 1')
                            i = i.replace('https://' + AWS_S3_CUSTOM_DOMAIN, str('https://' + CLOUDFRONT_DOMAIN))
                            print(i)
                        elif i.startswith(str('https://' + CLOUDFRONT_DOMAIN)):
                            print('swapping 2')
                            i = i.replace(str('https://' + CLOUDFRONT_DOMAIN), 'https://' + AWS_S3_CUSTOM_DOMAIN)
                            print(i)
                        else:
                            print('bad')
                            negativeResponses.append(i)
                            continue
                        rec = ProductAsset.objects.filter(baseImage=i).filter(assetType=2)

                    if rec.count() != 1:
                        negativeResponses.append(i)
                    else:
                        # we have a valid candidate for deletion
                        rec = rec[0]
                        print(rec)
                        if USE_LOCAL_STORAGE:
                            print('360 deletion not implemented for local storage use case yet')
                        else:
                            directory_parts = rec.xml.split('/')
                            print(directory_parts)
                            lc = 0
                            target_directory = ''
                            target_flag = False
                            for directory_part in directory_parts:
                                if directory_part == '360':
                                    target_directory += directory_part + '/'
                                    target_flag = True
                                else:
                                    target_directory += directory_part + '/'
                                    if target_flag:
                                        target_directory += directory_parts[lc + 1]
                                        break
                                lc += 1

                            if target_directory.startswith(str('https://' + CLOUDFRONT_DOMAIN)):
                                target_directory = target_directory.replace(str('https://' + CLOUDFRONT_DOMAIN),
                                                                            'https://' + AWS_S3_CUSTOM_DOMAIN)
                            print('target_directory', target_directory)
                            # get all the files in the directory
                            _, pathList = findFiles(target_directory, folder=False, subd=True)
                            # delete all of those files
                            success, errors, _ = delete_file_boto3(pathList)
                            print(success, errors)
                            try:
                                rec.delete()
                            except Exception as e:
                                if success == True:
                                    print('did not delete entry, but deleted assets', e)
                                else:
                                    print('did not delete all assets and did not delete entry', e)

                    print('end 360 delete process')

                else:
                    print('wanting to delete a different type of asset')
                count += 1
            print('finished deleting')
            if len(negativeResponses) == 0:
                return Response({'errorList': negativeResponses, 'status': status.HTTP_200_OK},
                                status=status.HTTP_200_OK)
            elif len(negativeResponses) == len(key):
                return Response({'errorList': negativeResponses, 'status': Response(status.HTTP_400_BAD_REQUEST)},
                                status=status.HTTP_400_BAD_REQUEST)
            else:
                return Response({'errorList': negativeResponses, 'status': Response(status.HTTP_206_PARTIAL_CONTENT)},
                                status=status.HTTP_206_PARTIAL_CONTENT)


# TODO: [BACKEND][Jacob] fix the message threading
class ProductMessageThreadListAPI(AuthenticatedUserMixin, APIView):
    queryset = ProductMessageThread.objects.all()
    permission_classes = (IsAuthenticated,)
    # authentication_classes = (TokenAuthentication)
    renderer_classes = (renderers.JSONRenderer,)
    serializer_class = ProductMessageThreadListSerializer

    def get(self, request):
        messageThread = ProductMessageThread.objects.all()
        serializer = ProductMessageThreadListSerializer(messageThread, many=True)
        return Response(serializer.data)

    @csrf_exempt
    def post(self, request, *args, **kwargs):
        serializer = ProductMessageThreadCreateSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class ProductMessageThreadAPI(AuthenticatedUserMixin, APIView):
    queryset = ProductMessageThread.objects.all()
    permission_classes = (IsAuthenticated,)
    # authentication_classes = (TokenAuthentication)
    renderer_classes = (renderers.JSONRenderer,)
    serializer_class = ProductMessageThreadCreateSerializer

    def get(self, request, product, slug, itype, filename):
        # TODO: [FRONTEND][Angel] product is unused work with frontend to get rid of it
        # TODO: [FRONTEND][Angel] frontend needs to send back filename or some sort of data,
        #  right now the backend is not getting anything besides product slug, itype, and an empty request array
        # Retrieve the product in question.
        # print("!!!!!!!!!!!!!!!!!!!!!!", slug)
        dbproduct = Product.objects.get(slug=slug)
        # print("!!!!!!!!!!!!!!!!!!!!!!", dbproduct.id)
        count = ProductAsset.objects.filter(assetType=1, product=dbproduct.id).count()
        # print("!!!!!!!!!!!!!!!!!!!!!!", count)
        # Perform filter if not an empty list.
        if count != 0:
            dbproduct2d = ProductAsset.objects.filter(assetType=1, product=dbproduct.id, deletedon='')
            url = []

            # go through 2d assets
            for geturl in dbproduct2d:

                splitArray = geturl.url.split('/')
                # print("!!!!!!!!!!!!!!!!!!!!!!   splitarray:", splitArray)
                # print("!!!!!!!!!!!!!!!!!!!!!!   filename:", filename)
                # print("!!!!!!!!!!!!!!!!!!!!!!   request:", request.FILES)
                # print("!!!!!!!!!!!!!!!!!!!!!!   product:", product)
                # print("!!!!!!!!!!!!!!!!!!!!!!   itype:", itype)
                # determine if it is the correct 2d asset
                if filename in splitArray:

                    try:
                        # print("!!!!!!!!!!!!!!!!!!!!!!", geturl, geturl.id)
                        productFile = ProductAsset.objects.get(assetType=1, pk=geturl.id)
                        thread = ProductMessageThread.objects.filter(productAsset=productFile)
                        if not thread:
                            raise Exception('Create thread')
                        elif thread == 1:
                            pass
                    except:
                        prod = ProductAsset.objects.get(assetType=1, pk=geturl.id)
                        new_thread = ProductMessageThread()
                        new_thread.product = dbproduct
                        new_thread.product2DImage = prod
                        new_thread.messageThreadName = filename
                        currentDT = datetime.datetime.now()
                        new_thread.creationDateTime = currentDT
                        new_thread.lastUpdateDateTime = currentDT
                        new_thread.save()
                        productFile = ProductAsset.objects.get(assetType=1, pk=geturl.id)

                    # After retrieving productFileObject, create productMessageThread, with productFile
                    thread = ProductMessageThread.objects.filter(productAsset=productFile)
                    serializer = ProductMessageThreadCreateSerializer(thread, many=True)
                    url.append(serializer)
            if len(url) == 0:
                return Response({'success': 'None', 'error': 'No 2D image found'})
            return Response(url, status=status.HTTP_200_OK)
        else:
            return Response({'success': 'None', 'error': 'No 2D image found'})

    @csrf_exempt
    def post(self, request, *args, **kwargs):

        serializer = ProductMessageThreadCreateSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class ProductMessageAPI(AuthenticatedUserMixin, APIView):
    queryset = ProductMessage.objects.all()
    permission_classes = (IsAuthenticated,)
    # authentication_classes = (TokenAuthentication)
    renderer_classes = (renderers.JSONRenderer,)
    serializer_class = ProductMessageSerializer

    def get(self, request, product, slug, itype, filename):

        profile = Profile.objects.get(user=request.user)

        # Retrieve the product in question.
        dbproduct = Product.objects.get(slug=slug)
        count = ProductAsset.objects.filter(assetType=1, product_id=dbproduct.id).count()

        productFile = None
        # Perform filter if not an empty list.
        if count != 0:
            dbproduct2d = ProductAsset.objects.filter(assetType=1, product_id=dbproduct.id, deletedon='')
            url = []

            for geturl in dbproduct2d:
                # File Split Name
                splitArray = geturl.url.split('/')
                fileName = splitArray[9]

                if filename == fileName:
                    productFile = ProductAsset.objects.get(assetType=1, pk=geturl.id)

            # After retrieving productFileObject, create productMessageThread, with productFile
            # as
            thread = ProductMessageThread.objects.get(productAsset=productFile)
            messages = ProductMessage.objects.filter(productMessageThread=thread)
            serializer = ProductMessageSerializer(messages, many=True)
            return Response(serializer.data)

        # Message.objects.filter(sender=user.profile, receiver=other_user.profile) | Message.objects.filter(
        #    receiver=user.profile, sender=other_user.profile)
        # return Response(serializer.data)

    @csrf_exempt
    def post(self, request, product, slug, itype, filename, *args, **kwargs):
        user = request.user
        profile = Profile.objects.get(user=user)
        # adminUser = User.objects.get(pk=1)

        # Retrieve the product in question.
        dbproduct = Product.objects.get(slug=slug)
        dbproduct2d = ProductAsset.objects.filter(assetType=1, product_id=dbproduct.id, deletedon='')
        productFile = None
        # Perform filter if not an empty list.
        if dbproduct2d.count() != 0:
            url = []
            for geturl in dbproduct2d:
                # File Split Name
                splitArray = geturl.url.split('/')
                fileName = splitArray[9]
                if (filename == fileName):
                    productFile = ProductAsset.objects.get(assetType=1, pk=geturl.id)
            # After retrieving productFileObject, create productMessageThread, with productFile
            # as
            thread = ProductMessageThread.objects.get(productAsset=productFile)

        # print userid
        # product = Product.objects.get(profile=profile)
        request.data["productMessageThread"] = thread.id
        request.data["sender"] = profile.user.id
        # ipdb.set_trace()

        serializer = ProductMessageSerializer(data=request.data)

        if serializer.is_valid():
            newMessageObject = serializer.save()
            thread.messages.add(newMessageObject)
            recipientsList = []
            try:
                for x in thread.participants.all():
                    us = x['participant']
                    if (profile.user != us):
                        recipientsList.append(x)
                        new_part = MessageParticipant()
                        new_part.first_name = user.first_name
                        new_part.last_name = user.last_name
                        new_part.participant = user
                        new_part.save()
                        thread.participants.add(new_part)

            except:
                new_part = MessageParticipant()
                new_part.first_name = user.first_name
                new_part.last_name = user.last_name
                new_part.participant = user
                new_part.save()
                thread.participants.add(new_part)

            thread.save()

            message = request.data["message"]
            # Grab First Recipient
            newMessageObject.recipient = recipientsList[0]
            newMessageObject.save()

            sender = profile.user
            for x in recipientsList:
                recipient = x
                ctx = {
                    'subject': 'XSPACE | New Message',
                    'first_name': sender.first_name,
                    'last_name': sender.last_name,
                    'rep_firstname': recipient.first_name,
                    'rep_lastname': recipient.last_name,
                    'comment_link': 'https://xspaceapp.com/',
                    'product_name': dbproduct.name,
                    'product_link': 'https://xspaceapp.com/',
                    'product_upc': dbproduct.upccode,
                    'product_sku': dbproduct.SKU,
                    'file_name': filename,
                    'photoMessage': message
                }
                message = get_template('PhotoReviewEmail.html').render(ctx)

                mail_subject = '[XSPACE Web App] You have a new message from ' + str(
                    profile.user.first_name) + ' ' + str(profile.user.last_name)  # '+str(current_site)+'
                to_email = str(recipient.email)

                EmailThread(mail_subject, message, to_email).start()

            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class ProductMessageThreadByUUIDAPI(AuthenticatedUserMixin, APIView):
    queryset = ProductMessageThread.objects.all()
    permission_classes = (IsAuthenticated,)
    # authentication_classes = (TokenAuthentication)
    renderer_classes = (renderers.JSONRenderer,)
    serializer_class = ProductMessageThreadCreateSerializer

    def get(self, request, uidb64):
        thread = ProductMessageThread.objects.filter(messageThreadID=uidb64)
        serializer = ProductMessageThreadCreateSerializer(thread, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)


class ProductMessageByUUIDAPI(AuthenticatedUserMixin, APIView):
    queryset = ProductMessage.objects.all()
    permission_classes = (IsAuthenticated,)
    # authentication_classes = (TokenAuthentication)
    renderer_classes = (renderers.JSONRenderer,)
    serializer_class = ProductMessageSerializer

    def get(self, request, uidb64, pk=None):
        # ipdb.set_trace()
        thread = ProductMessageThread.objects.get(messageThreadID=uidb64)
        messages = ProductMessage.objects.filter(productMessageThread=thread)
        serializer = ProductMessageSerializer(messages, many=True)
        return Response(serializer.data)

    @csrf_exempt
    def post(self, request, uidb64, *args, **kwargs):

        thread = ProductMessageThread.objects.get(messageThreadID=uidb64)
        dbproduct = thread.product
        filename = thread.product2DImage
        request.data["productMessageThread"] = thread.id
        request.data["sender"] = request.user.id
        profile = Profile.objects.get(user=request.user)
        company = profile.companyProfile
        user = request.user

        serializer = ProductMessageSerializer(data=request.data)

        if serializer.is_valid():
            newMessageObject = serializer.save()

            thread.messages.add(newMessageObject)

            recipientsList = []
            recipientsListus = []
            # ipdb.set_trace()
            threadps = thread.participants.all()
            z = ZendeskAPI(request)

            for x in threadps:
                us = x.participant
                recipientsList.append(us)
                recipientsListus.append(x)

                # p = z.createZendeskUser(x)

            if user in recipientsList:
                pass
            else:
                new_part = MessageParticipant()
                new_part.first_name = user.first_name
                new_part.last_name = user.last_name
                new_part.participant = user
                new_part.save()
                thread.participants.add(new_part)

            thread.save()

            message = request.data["message"]
            # ipdb.set_trace()
            try:
                tt = int(thread.ticket)
                z.ticketComment(tt, user, message)
            except ValueError as e:
                msg = str(message)
                tt = z.createZendeskTicket(user, company, thread.product, str(thread.messageThreadName), msg)
                thread.ticket = tt
                thread.save()

            # Grab First Recipient
            # newMessageObject.recipient = recipientsList[us]
            newMessageObject.save()

            for x in recipientsList:
                recipient = x
                sender = x

                #                check4ZenOrg()
                #                check4ZenUsers()
                #                 if none - addZenUser
                #                 add ZenUserticket 

                ctx = {
                    'subject': 'XSPACE | New Message',
                    'first_name': sender.first_name,
                    'last_name': sender.last_name,
                    'rep_firstname': "XSPACE",
                    'rep_lastname': "Operations",
                    'comment_url': 'https://xspaceapp.com/',
                    'product_name': dbproduct.name,
                    'product_link': 'https://xspaceapp.com/',
                    'product_upc': dbproduct.upccode,
                    'product_sku': dbproduct.SKU,
                    'file_name': filename,
                    'photoMessage': message
                }

                notify.send(user, recipient=user, verb='you recieved a message')

                message = get_template('PhotoReviewEmail.html').render(ctx)

                mail_subject = '[XSPACE Web App] You have a new message from ' + str(
                    request.user.first_name) + ' ' + str(request.user.last_name)  # '+str(current_site)+'
                to_email = str(recipient)
                # ipdb.set_trace()

                EmailThread(mail_subject, message, to_email).start()

            return Response(serializer.data, status=status.HTTP_201_CREATED)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class s3_get_files(generics.RetrieveAPIView):

    def post(self, request):
        print("!!!!!!!!!!!!!!!!!!!!!!     s3_get_files.POST Function   !!!!!!!!!!!!!!!!!!!!!")

        try:
            location = request.data['location']
            product = request.data['product']
            pid = request.data['pid']
            slug = request.data['slug']
            itype = request.data['itype']

            # oldget(request, location, product, pid, slug, itype)
            product = Product.objects.get(pk=pid)
            url = product.image
            return Response({'success': str(url), 'error': 'None'})
        except Product.DoesNotExist:
            return Response({'success': 'None', 'error': 'No product'})


class s3_get_360(generics.RetrieveAPIView):

    def post(self, request):
        print("!!!!!!!!!!!!!!!!!!!!!!     s3_get_360.POST Function   !!!!!!!!!!!!!!!!!!!!!")

        # files = oldget(request, location, product, pid, slug, itype)
        try:
            location = request.data['location']
            product = request.data['product']
            pid = request.data['pid']
            slug = request.data['slug']
            itype = request.data['itype']

            dbproduct = Product.objects.get(pk=pid)
            try:
                dbproduct360 = ProductAsset.objects.get(assetType=2, product_id=dbproduct)
                xmlfile = dbproduct360.xml
                baseimgpath = dbproduct360.baseImage
                csspath = dbproduct360.css
                jquerypath = dbproduct360.jquery
                imagerotatorpath = dbproduct360.imagerotator

                licensepath = 'https://dbrdts7bui7s7.cloudfront.net/product_360_licence/licence.lic'

                return Response({
                    'success': 'true',
                    'csspath': csspath,
                    'jquerypath': jquerypath,
                    'imagerotatorpath': imagerotatorpath,
                    'licensepath': licensepath,
                    'xmlpath': xmlfile,
                    'baseimage': baseimgpath,
                    'error': 'None'
                    # 'files': files
                })
            except Exception as e:
                return Response({'success': 'None', 'error': 'No 360 image found'})

        except Product.DoesNotExist:
            return Response({'success': 'None', 'error': 'No product'})


class s3_get_2d(AuthenticatedUserMixin, generics.RetrieveAPIView):

    # @csrf_exempt
    # @has_permission(P.GENERIC)
    def post(self, request):
        print("!!!!!!!!!!!!!!!!!!!!!!     s3_get_2d.POST Function   !!!!!!!!!!!!!!!!!!!!!")

        # files = oldget(request, location, product, pid, slug, itype)
        if 'listofslugs' in request.data:
            if len(request.data['listofslugs']) > 0:

                success = 'OK'

                urls_all = dict()
                for slug in request.data['listofslugs']:

                    try:

                        dbproduct = Product.objects.get(slug=slug)
                        count = ProductAsset.objects.filter(assetType=1, product=dbproduct).count()
                        if count != 0:
                            print(slug)
                            name = dbproduct.name
                            dbproduct2d = ProductAsset.objects.filter(assetType=1, product=dbproduct, deletedon='')
                            url = []
                            for geturl in dbproduct2d:
                                if geturl.thumbnail_url != '':
                                    url.append({'link': geturl.url,
                                                'size': geturl.size,
                                                'filename': geturl.file_name,
                                                'assetID': geturl.uniqueID,
                                                'name': name})
                                # print(geturl.thumbnail_id)
                                # print geturl.size

                            urls_all[slug] = url

                    except Product.DoesNotExist:
                        urls_all[slug] = 'does not exit'
                        success = 'Problems'

                return Response({
                    'success': success,
                    'urls': urls_all,
                    # 'files': files
                }, status=status.HTTP_200_OK)
            else:
                return Response({'success': 'None', 'error': 'Must supply slugs'}, status=status.HTTP_400_BAD_REQUEST)


        else:
            try:
                location = request.data['location']
                product = request.data['product']
                pid = request.data['pid']
                slug = request.data['slug']
                itype = request.data['itype']

                dbproduct = Product.objects.get(pk=pid)
                count = ProductAsset.objects.filter(assetType=1, product=dbproduct).count()
                if count != 0:
                    dbproduct2d = ProductAsset.objects.filter(assetType=1, product=dbproduct, deletedon='')
                    url = []
                    for geturl in dbproduct2d:
                        url.append({'link': geturl.url, 'size': geturl.size})
                        # print geturl.size

                    return Response({
                        'success': 'true',
                        'url': url,
                        'error': 'None'
                        # 'files': files
                    })
                else:
                    return Response({'success': 'None', 'error': 'No 2D image found'})

            except Product.DoesNotExist:
                return Response({'success': 'None', 'error': 'No product'})


class AssetUploadAPI(AuthenticatedUserMixin, APIView):
    # print("!!!!!!!!!!!!!!!!!!!!!!     AssetUploadAPI Function   !!!!!!!!!!!!!!!!!!!!!")

    permission_classes = (IsAuthenticated,)
    authentication_classes = (BasicAuthentication,)
    parser_classes = (MultiPartParser, FileUploadParser)

    # renderer_classes = (renderers.JSONRenderer,)

    @csrf_exempt
    def post(self, request, *args, **kwargs):
        print("!!!!!!!!!!!!!!!!!!!!!!     AssetUploadAPI.POST Function   !!!!!!!!!!!!!!!!!!!!!")

        try:
            myfile = request.FILES['file'] if request.FILES else None
            if myfile:
                with open('/Users/bburton/Desktop/imgdump/' + myfile.name, 'wb+') as destination:
                    for chunk in myfile.chunks():
                        destination.write(chunk)
                    destination.close()
                return Response({"status": "files are here"}, status=200)
            else:
                return Response({"status": "files are not here"}, status=200)
        except Exception as e:
            return Response({"error": e})


class ProductDetail(AuthenticatedUserMixin, APIView):
    queryset = Product.objects.all()
    serializer_class = ProductSerializer

    @csrf_exempt
    def post(self, request, pagesize=None, pageno=None, format=None):
        print("!!!!!!!!!!!!!!!!!!!!!!     ProductDetail.POST Function   !!!!!!!!!!!!!!!!!!!!!")

        pk = request.user.id

        user = User.objects.filter(pk=pk)
        profile = Profile.objects.get(user=user)

        snippets = Product.objects.filter(profile=profile)
        paginator = Paginator(snippets, pagesize)

        totalrow = paginator.count
        # pagecount = [len(totalrow)]

        page = pageno
        try:
            snippets = paginator.page(page)
        except PageNotAnInteger:
            # If page is not an integer, deliver first page.
            snippets = paginator.page(1)
        except EmptyPage:
            # If page is out of range (e.g. 9999), deliver last page of results.
            snippets = paginator.page(paginator.num_pages)

        serializer = ProductSerializer(snippets, many=True)
        return Response(serializer.data + [{'pagecount': totalrow}])  # 'productid':'0',


########################################################################
class ProductUpdateDetail(generics.RetrieveAPIView):
    queryset = Product.objects.all()
    serializer_class = ProductSerializer

    def post(self, request, pk=None, format=None):
        print("!!!!!!!!!!!!!!!!!!!!!!     ProductUpdateDetail.POST Function   !!!!!!!!!!!!!!!!!!!!!")

        try:
            user = request.data['user']
            user = User.objects.get(username=user)
            profile = Profile.objects.get(user=user)
            product = Product.objects.get(pk=pk, profile=profile)
            name = request.data['data']['name']
            SKU = request.data['data']['sku']
            price = request.data['data']['price']
            UPCcode = request.data['data']['upccode']
            description = request.data['data']['description']
            manufacturer = request.data['data']['manufacturer']
            # UPCType=request.data['data']['UPCType']
            length = request.data['data']['length']
            width = request.data['data']['width']
            height = request.data['data']['height']
            # TODO: [Backend][Jacob] add in weight
            category = request.data['data']['selectedCategory']
            product.name = name
            product.SKU = SKU
            product.price = price
            product.upccode = UPCcode
            # product.UPCType = UPCType
            product.description = description
            product.manufacturer = manufacturer
            product.length = length
            product.width = width
            product.height = height
            # TODO: [Backend][Jacob] add in weight
            product.category_id = category
            product.save()

            return Response({'response': 'successful', 'error': 'None', 'data': request.data},
                            status=status.HTTP_201_CREATED)
        except Product.DoesNotExist:
            return Response({'response': 'None', 'error': 'Falied'}, status=status.HTTP_201_CREATED)


class XEditAPI(AuthenticatedUserMixin, APIView):
    """
    api interface with xedit

    generally:
    POST -  is for saving edited assets
    GET - is for conducting an intensive action on an asset and returning it to the frontend
    PUT - is undefined at this time
    """
    permission_classes = (IsAuthenticated,)
    parser_classes = (MultiPartParser, JSONParser)
    renderer_classes = (JSONRenderer, MultiPartRenderer)

    @csrf_exempt
    @has_permission(P.GENERIC)
    def post(self, request):
        """
        used for saving edited assets in xedit
        :param request:
        :return:
        """
        print('request', request)
        # placeholder = json.loads(request.data)
        # print(placeholder)
        print('request mask_url', request.data['mask_url'])
        print('request original_url', request.data['original_url'])
        print('request product_info', request.data['product_info'])
        print('request width', request.data['width'])
        print('request height', request.data['height'])

        # placeholder for custom background colors
        background_color_r = 255
        background_color_g = 255
        background_color_b = 255
        background_color_a = 255

        try:
            pk = request.user.id
            user = User.objects.get(pk=pk)
            profile = Profile.objects.get(user_id=user.id)
            company = Company.objects.get(id=profile.companyProfile.id)
            productAsset = ProductAsset.objects.get(url=request.data['original_url'])
            product = Product.objects.get(uniqueID=request.data['product_info']['uniqueID'])
        except:
            return Response(status=status.HTTP_400_BAD_REQUEST)

        print('got all the info')
        newMaskFlag = False
        if request.data['mask_url'] is None or productAsset.mask_url == '' or productAsset.mask_url == 'null':
            # we'll want to make a mask entry since it doesn't already exist
            print('creating new mask')
            maskAsset = ProductAsset()
            maskAsset.file_name = productAsset.file_name
            maskAsset.lastmodified = timezone.now()
            maskAsset.asset_height = int(request.data['height'])
            maskAsset.asset_width = int(request.data['width'])
            maskAsset.asset_channels = 1
            maskAsset.asset_background_type = 3
            maskAsset.assetType = 1
            maskAsset.product_id = product.id
            maskAsset.company_id = company.id
            maskAsset.save()
            newMaskFlag = True
        else:
            maskAsset = ProductAsset.objects.get(url=request.data['mask_url'])

        received_mask = np.asarray(request.data['data'])

        received_mask = received_mask.reshape(request.data['height'], request.data['width'])

        # convert to float for accuracy
        received_mask = received_mask.astype('float64')
        # normalize between 1 and 0 for against the original image
        # print('before normalization between 1 and 0', received_mask)
        received_mask /= 255
        # print('after normalization between 1 and 0', received_mask)
        # print('request received_mask', received_mask)
        # print('received mask shape', received_mask.shape)
        original_img = load_image(request.data['original_url'])
        original_img = Image.open(io.BytesIO(original_img))

        IMAGE_FORMAT = original_img.format
        original_img = np.array(original_img)
        for channel in range(original_img.shape[-1]):
            original_img[:, :, channel] = np.multiply(original_img[:, :, channel], received_mask)
        received_mask *= 255
        received_mask = received_mask.astype('uint8')
        # print('after renormalization', received_mask)

        if original_img.shape[-1] == 3:
            # stack the mask so that it has an alpha channel (we'll remove it before resaving so that
            # jpg doesn't freak the f*** out
            original_img = np.dstack((original_img, received_mask))
        # apply the desired background in
        original_img[np.where((original_img == [0, 0, 0, 0]).all(axis=2))] = [background_color_r, background_color_g,
                                                                              background_color_b, background_color_a]
        if original_img.shape[-1] == 3:
            # remove the alpha channel if this was a jpeg
            original_img = original_img[:, :, :2]
        # save the new original image
        saveable_img = Image.fromarray(original_img)
        in_mem_file = io.BytesIO()
        saveable_img.save(in_mem_file, format=IMAGE_FORMAT)
        in_mem_file.seek(0)
        key = (str(company.slug) + '/products/' + str(product.slug) + '/2D')
        new_raster_location = boto3Upload3(in_mem_file, company.slug, key, is_bytes=True,
                                           fileName=productAsset.file_name)

        # save the new mask image
        saveable_img = Image.fromarray(received_mask)
        in_mem_file = io.BytesIO()
        saveable_img.save(in_mem_file, format=IMAGE_FORMAT)
        in_mem_file.seek(0)
        key = (str(company.slug) + '/products/' + str(product.slug) + '/masks/2D/' + str(company.slug))
        new_mask_location = boto3Upload3(in_mem_file, company.slug, key, is_bytes=True,
                                         fileName=productAsset.file_name)
        if newMaskFlag:
            maskAsset.url = new_mask_location
            productAsset.mask_url = maskAsset.url
            maskAsset.is_mask = True

        maskAsset.save()
        productAsset.lastmodified = timezone.now()
        productAsset.save()

        return Response(status=status.HTTP_200_OK)


class Generate360(AuthenticatedUserMixin, APIView):
    """
    api interface with xedit

    generally:
    POST -  is for creating a 360
    GET - unused
    PUT - unused
    """
    permission_classes = (IsAuthenticated,)
    parser_classes = (MultiPartParser, JSONParser)
    renderer_classes = (JSONRenderer, MultiPartRenderer)

    @csrf_exempt
    @has_permission(P.GENERIC)
    def post(self, request):
        """
        create 360 takes 3 arguments product slug, urllist of images, and last modified list
        :param request:
        :return:
        """
        print('Generate360 received:', request.data)
        # get admin info
        pk = request.user.id
        try:
            user = User.objects.get(pk=pk)
        except:
            return Response({'payload': 'Unknown User ID 001'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            profile = Profile.objects.get(user_id=user.id)
        except:
            return Response({'payload': 'Unknown User ID 002'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            company = Company.objects.get(id=profile.companyProfile.id)
        except:
            return Response({'payload': 'Profile not connected to a valid company'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            product = Product.objects.get(slug=request.data['slug'])
        except:
            return Response({'payload': 'Improper Product Slug'}, status=status.HTTP_400_BAD_REQUEST)
        if product.company.id != company.id:
            return Response({'payload': 'Restricted Access, specified company does not have access to this product'},
                            status=status.HTTP_400_BAD_REQUEST)

        # with the gathered information, create the 360
        if request.data['type'] == '1':
            flatten = True
        else:
            flatten = False
        rows = int(request.data['type'])

        # get product url and build targetdirectory/resultDirectory

        outputLocation = company.slug + '/products/' + product.slug + \
                         '/360/1'

        new_asset, spin_list, fc_key = prep360(request.data['urlList'], outputLocation, product, '1')
        print(new_asset, spin_list, fc_key)
        params = {'targetDirectory': os.path.split(str(fc_key))[0],
                  'productAsset': new_asset,
                  'productAssetList': spin_list,
                  'resultDirectory': None,
                  'flatten': flatten,
                  'mediaUploadFlag': True}
        compile360Bundle(params=params)

        return Response({'payload': '360 Created Successfully'}, status=status.HTTP_201_CREATED)
